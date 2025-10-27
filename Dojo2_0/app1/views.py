from tokenize import Comment
from django.conf import settings
from django.shortcuts import render
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

# Create your views here.

from functools import cache
from django.shortcuts import get_list_or_404, render
from .serializers import CompanyLogoSerializer, KeyEventSerializer, MasterTableSerializer, RegisterSerializer, ScoreSerializer, SimpleScoreSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
#(2) Views for the User Login Views for the User Login
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
# #Create your views here.
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import login
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import ValidationError
from .serializers import LoginSerializer

from django.shortcuts import render

# views.py
from django.shortcuts import render, get_object_or_404
from django.contrib.auth import login
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import ValidationError

from .serializers import LoginSerializer
from .models import User

def dojo_app(request):
    return render(request, 'index.html')


from django.contrib.auth import get_user_model, login
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken

User = get_user_model()

class LoginAPIView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        print("Login request data:", request.data)  # ok for debug, remove in production

        serializer = LoginSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            # return validation/auth errors
            return Response(
                {"message": "Authentication failed", "errors": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        # serializer is valid -> get user by id
        user_id = serializer.validated_data.get('user_id')
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return Response(
                {"message": "Authentication failed", "errors": "User not found"},
                status=status.HTTP_401_UNAUTHORIZED
            )

        # Log user in (optional depending on token-only approach)
        login(request, user)

        # Generate JWT tokens
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)

        # Build user payload safely (use getattr with defaults)
        user_payload = {
            'email': getattr(user, 'email', None),
            'first_name': getattr(user, 'first_name', None),
            'last_name': getattr(user, 'last_name', None),
            'employeeid': getattr(user, 'employeeid', None),
            # Prefer returning primitive values (ids or names) for related objects:
            'role': getattr(getattr(user, 'role', None), 'name', None),
            'hq': getattr(getattr(user, 'hq', None), 'name', getattr(user, 'hq', None)),
            'factory': getattr(getattr(user, 'factory', None), 'name', getattr(user, 'factory', None)),
            'department': getattr(getattr(user, 'department', None), 'name', getattr(user, 'department', None)),
            'status': getattr(user, 'status', None),
        }

        return Response({
            'message': 'Login successful',
            'access_token': access_token,
            'refresh_token': str(refresh),
            'user': user_payload
        }, status=status.HTTP_200_OK)





#(1) Views for the User Register

from django.db import IntegrityError
from django.shortcuts import render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .models import CompanyLogo, KeyEvent, MasterTable, Score, TestSession, User

class RegisterView(generics.GenericAPIView):
    serializer_class = RegisterSerializer

    def post(self, request):
        user_data = request.data
        serializer = self.serializer_class(data=user_data)

        try:
            serializer.is_valid(raise_exception=True)

            # Check if user already exists by email
            if User.objects.filter(email=user_data.get("email")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"email": "This email is already registered."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if employee ID already exists
            if User.objects.filter(employeeid=user_data.get("employeeid")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"employeeid": "This employee ID is already in use."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Save the user
            serializer.save()

            return Response({
                "message": "User registered successfully!"
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            # Handle specific validation errors
            return Response({
                "message": "Validation failed",
                "errors": e.detail
            }, status=status.HTTP_400_BAD_REQUEST)

        except IntegrityError:
            # Handle database integrity errors (like duplicate entries)
            return Response({
                "message": "Database error",
                "errors": {"detail": "Duplicate entry or constraint violation."}
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            # Handle unexpected errors
            return Response({
                "message": "Unexpected error occurred",
                "errors": {"detail": str(e)}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import action
from .models import User, Role
from .serializers import RegisterSerializer, RoleSerializer

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all().select_related("role")
    serializer_class = RegisterSerializer

    def get_permissions(self):
        if self.action in ["create"]:
            return [AllowAny()]  # registration allowed for unauthenticated
        return [IsAuthenticated()]

    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def me(self, request):
        """Return current user info"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

class RoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer

    def get_permissions(self):
        if self.action in ["create", "list"]:
            return [AllowAny()]  # Allow unauthenticated users to create and list roles
        return [IsAuthenticated()]


#(3) Views for the User Logout

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import LogoutSerializer

class LogoutAPIView(APIView):
    """
    User Logout API View
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = LogoutSerializer(data=request.data)
        if serializer.is_valid():
            refresh_token = serializer.validated_data["refresh_token"]

            try:
                token = RefreshToken(refresh_token)
                token.blacklist()  # Blacklist the refresh token
                return Response({"message": "Logout successful"}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    





from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from .models import Hq, Factory, Department, Line, SubLine, Station
from .serializers import (
    HqSerializer, FactorySerializer, DepartmentSerializer,
    LineSerializer, SubLineSerializer, StationSerializer
)

class HqViewSet(viewsets.ModelViewSet):
    queryset = Hq.objects.all()
    serializer_class = HqSerializer
    
    def get_queryset(self):
        queryset = Hq.objects.all().order_by('hq_name')
        return queryset

class FactoryViewSet(viewsets.ModelViewSet):
    queryset = Factory.objects.all()
    serializer_class = FactorySerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['hq']
    
    def get_queryset(self):
        queryset = Factory.objects.all().select_related('hq').order_by('factory_name')
        hq_id = self.request.query_params.get('hq', None)
        if hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
        return queryset

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['factory', 'hq']
    
    def get_queryset(self):
        queryset = Department.objects.all().select_related('factory', 'hq').order_by('department_name')
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['department', 'factory', 'hq']
    
    def get_queryset(self):
        queryset = Line.objects.all().select_related('department', 'factory', 'hq').order_by('line_name')
        department_id = self.request.query_params.get('department', None)
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if department_id is not None:
            queryset = queryset.filter(department_id=department_id)
        elif factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

class SubLineViewSet(viewsets.ModelViewSet):
    queryset = SubLine.objects.all()
    serializer_class = SubLineSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['line', 'department', 'factory', 'hq']
    
    def get_queryset(self):
        queryset = SubLine.objects.all().select_related('line', 'department', 'factory', 'hq').order_by('subline_name')
        line_id = self.request.query_params.get('line', None)
        department_id = self.request.query_params.get('department', None)
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if line_id is not None:
            queryset = queryset.filter(line_id=line_id)
        elif department_id is not None:
            queryset = queryset.filter(department_id=department_id)
        elif factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['subline', 'line', 'department', 'factory', 'hq']
    
    def get_queryset(self):
        queryset = Station.objects.all().select_related('subline', 'line', 'department', 'factory', 'hq').order_by('station_name')
        subline_id = self.request.query_params.get('subline', None)
        line_id = self.request.query_params.get('line', None)
        department_id = self.request.query_params.get('department', None)
        factory_id = self.request.query_params.get('factory', None)
        hq_id = self.request.query_params.get('hq', None)
        
        if subline_id is not None:
            queryset = queryset.filter(subline_id=subline_id)
        elif line_id is not None:
            queryset = queryset.filter(line_id=line_id)
        elif department_id is not None:
            queryset = queryset.filter(department_id=department_id)
        elif factory_id is not None:
            queryset = queryset.filter(factory_id=factory_id)
        elif hq_id is not None:
            queryset = queryset.filter(hq_id=hq_id)
            
        return queryset

# Add this to your views.py

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import HierarchyStructure
from .serializers import HierarchyStructureSerializer

# views.py - Updated HierarchyStructureViewSet
from rest_framework.decorators import api_view

@api_view(['GET'])
def get_all_departments(request):
    departments = Department.objects.all()
    serializer = DepartmentReadSerializer(departments, many=True)
    return Response({
        'departments': serializer.data,
    }, status=status.HTTP_200_OK)



class HierarchyByDepartmentView(APIView):
    """
    Fetch hierarchy by department_id with flexible nesting:
    department → line → subline → station
    department → line → station
    department → subline → station
    department → station
    """

    def get(self, request, *args, **kwargs):
        department_id = request.query_params.get("department_id")
        if not department_id:
            return Response(
                {"error": "department_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            department_id = int(department_id)
            structures = HierarchyStructure.objects.filter(department_id=department_id)

            if not structures.exists():
                return Response(
                    {"error": f"No hierarchy found for department id {department_id}"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            department = structures.first().department
            department_data = {
                "department_id": department.department_id,
                "department_name": department.department_name,
                "lines": {},
                "sublines": {},
                "stations": {},
            }

            # build hierarchy flexibly
            for structure in structures:
                line = structure.line
                subline = structure.subline
                station = structure.station

                if line:  # if line exists
                    if line.line_id not in department_data["lines"]:
                        department_data["lines"][line.line_id] = {
                            "line_id": line.line_id,
                            "line_name": line.line_name,
                            "sublines": {},
                            "stations": {},
                        }

                    if subline:  # subline under line
                        if subline.subline_id not in department_data["lines"][line.line_id]["sublines"]:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id] = {
                                "subline_id": subline.subline_id,
                                "subline_name": subline.subline_name,
                                "stations": {},
                            }

                        if station:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id]["stations"][
                                station.station_id
                            ] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }
                    else:  # station directly under line
                        if station:
                            department_data["lines"][line.line_id]["stations"][station.station_id] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }

                elif subline:  # no line, but subline exists
                    if subline.subline_id not in department_data["sublines"]:
                        department_data["sublines"][subline.subline_id] = {
                            "subline_id": subline.subline_id,
                            "subline_name": subline.subline_name,
                            "stations": {},
                        }

                    if station:
                        department_data["sublines"][subline.subline_id]["stations"][station.station_id] = {
                            "station_id": station.station_id,
                            "station_name": station.station_name,
                        }

                elif station:  # no line, no subline → station directly under department
                    department_data["stations"][station.station_id] = {
                        "station_id": station.station_id,
                        "station_name": station.station_name,
                    }

            # convert dicts to lists
            department_data["lines"] = list(department_data["lines"].values())
            for line in department_data["lines"]:
                line["sublines"] = list(line["sublines"].values())
                line["stations"] = list(line["stations"].values())
                for subline in line["sublines"]:
                    subline["stations"] = list(subline["stations"].values())

            department_data["sublines"] = list(department_data["sublines"].values())
            for subline in department_data["sublines"]:
                subline["stations"] = list(subline["stations"].values())

            department_data["stations"] = list(department_data["stations"].values())

            return Response(department_data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
from rest_framework.decorators import api_view 
from rest_framework.response import Response 
from .models import HierarchyStructure 
 
@api_view(['GET']) 
def get_hierarchy_structures(request): 
    structures = HierarchyStructure.objects.all() 
 
    merged_data = {} 
 
    for s in structures: 
        # key for grouping (structure_name + hq_id + factory_id) 
        key = (s.structure_name, s.hq.hq_id if s.hq else None, s.factory.factory_id if s.factory else None) 
 
        if key not in merged_data: 
            merged_data[key] = { 
                "structure_id": s.structure_id,  # you can choose the first id or min id 
                "structure_name": s.structure_name, 
                "hq": s.hq.hq_id if s.hq else None, 
                "hq_name": f"{s.hq.hq_name} (ID: {s.hq.hq_id})" if s.hq else None, 
                "factory": s.factory.factory_id if s.factory else None, 
                "factory_name": f"{s.factory.factory_name} (ID: {s.factory.factory_id})" if s.factory else None, 
                "structure_data": { 
                    "hq_name": s.hq.hq_name if s.hq else None, 
                    "factory_name": s.factory.factory_name if s.factory else None, 
                    "departments": [] 
                } 
            } 
 
        structure_data = merged_data[key]["structure_data"] 
 
        # Departments 
        if s.department: 
            dept_obj = next( 
                (d for d in structure_data["departments"] if d["id"] == s.department.department_id), 
                None 
            ) 
            if not dept_obj: 
                dept_obj = { 
                    "id": s.department.department_id, 
                    "department_name": s.department.department_name, 
                    "lines": [],
                    "stations": []  # Add stations array for departments
                } 
                structure_data["departments"].append(dept_obj) 
 
            # Handle stations directly under department (when no line/subline)
            if s.station and not s.line and not s.subline:
                if not any(st["id"] == s.station.station_id for st in dept_obj["stations"]):
                    dept_obj["stations"].append({
                        "id": s.station.station_id,
                        "station_name": s.station.station_name
                    })
 
            # Lines 
            if s.line: 
                line_obj = next( 
                    (l for l in dept_obj["lines"] if l["id"] == s.line.line_id), 
                    None 
                ) 
                if not line_obj: 
                    line_obj = { 
                        "id": s.line.line_id, 
                        "line_name": s.line.line_name, 
                        "sublines": [],
                        "stations": []  # Add stations array for lines
                    } 
                    dept_obj["lines"].append(line_obj) 
 
                # Handle stations directly under line (when no subline)
                if s.station and not s.subline:
                    if not any(st["id"] == s.station.station_id for st in line_obj["stations"]):
                        line_obj["stations"].append({
                            "id": s.station.station_id,
                            "station_name": s.station.station_name
                        })
 
                # Sublines 
                if s.subline: 
                    subline_obj = next( 
                        (sl for sl in line_obj["sublines"] if sl["id"] == s.subline.subline_id), 
                        None 
                    ) 
                    if not subline_obj: 
                        subline_obj = { 
                            "id": s.subline.subline_id, 
                            "subline_name": s.subline.subline_name, 
                            "stations": [] 
                        } 
                        line_obj["sublines"].append(subline_obj) 
 
                    # Stations under sublines
                    if s.station: 
                        if not any(st["id"] == s.station.station_id for st in subline_obj["stations"]): 
                            subline_obj["stations"].append({ 
                                "id": s.station.station_id, 
                                "station_name": s.station.station_name 
                            }) 
 
    # return merged list 
    return Response(list(merged_data.values()))




from rest_framework import viewsets, status
from rest_framework.response import Response
from django.db import transaction
from .models import HierarchyStructure
from .serializers import HierarchyStructureSerializer
import logging

logger = logging.getLogger(__name__)

class HierarchyStructureViewSet(viewsets.ModelViewSet):
    queryset = HierarchyStructure.objects.all().select_related(
        "hq", "factory", "department", "line", "subline", "station"
    )
    serializer_class = HierarchyStructureSerializer

    def create(self, request, *args, **kwargs):
        """Create hierarchy records from nested structure_data"""
        data = request.data
        print("📥 Incoming POST data:", data)

        structure_name = data.get('structure_name')
        structure_data = data.get('structure_data', {})
        hq_id = data.get('hq')
        factory_id = data.get('factory')

        if not structure_name:
            return Response({'error': 'structure_name is required'}, status=status.HTTP_400_BAD_REQUEST)

        return self._create_hierarchy_records(structure_name, structure_data, hq_id, factory_id)

    def update(self, request, *args, **kwargs):
        """Update hierarchy structure by deleting old and creating new records"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        data = request.data
        
        print(f"📥 Incoming PUT data for structure_id {instance.structure_id}:", data)

        # Get the structure_name from the instance or data
        structure_name = data.get('structure_name', instance.structure_name)
        structure_data = data.get('structure_data', {})
        hq_id = data.get('hq', instance.hq_id if instance.hq else None)
        factory_id = data.get('factory', instance.factory_id if instance.factory else None)

        if 'structure_data' in data:
            # If structure_data is provided, delete old records and create new ones
            print(f"🔄 Updating hierarchy structure: {structure_name}")
            return self._create_hierarchy_records(structure_name, structure_data, hq_id, factory_id)
        else:
            # If no structure_data, do regular update
            serializer = self.get_serializer(instance, data=data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)

    def _create_hierarchy_records(self, structure_name, structure_data, hq_id, factory_id):
        """Helper method to create hierarchy records (used by both create and update)"""
        try:
            with transaction.atomic():
                # Delete existing records with the same structure_name
                deleted_count = HierarchyStructure.objects.filter(structure_name=structure_name).count()
                HierarchyStructure.objects.filter(structure_name=structure_name).delete()
                print(f"🗑 Deleted {deleted_count} existing records for structure: {structure_name}")

                created_records = []

                # Iterate over departments → lines → sublines → stations
                departments = structure_data.get('departments', [])
                for department in departments:
                    dept_id = department.get('id')
                    print(f"🏢 Processing department {dept_id}")

                    # ✅ Case 1: stations directly under department
                    dept_stations = department.get('stations', [])
                    if dept_stations:
                        print(f"📌 Found {len(dept_stations)} stations directly under department {dept_id}")
                        for station in dept_stations:
                            station_id = station.get('id')
                            print(f"💾 Creating record: dept={dept_id}, station={station_id}")
                            
                            try:
                                record = HierarchyStructure.objects.create(
                                    structure_name=structure_name,
                                    hq_id=hq_id,
                                    factory_id=factory_id,
                                    department_id=dept_id,
                                    line_id=None,
                                    subline_id=None,
                                    station_id=station_id
                                )
                                created_records.append(record)
                                print(f"✅ Successfully created record with ID: {record.structure_id}")
                                
                                # Verify the record was actually saved
                                verification = HierarchyStructure.objects.filter(structure_id=record.structure_id).first()
                                if verification:
                                    print(f"✅ Verification: Record {record.structure_id} exists in DB with station_id={verification.station_id}")
                                else:
                                    print(f"❌ Verification failed: Record {record.structure_id} not found in DB")
                                    
                            except Exception as create_error:
                                print(f"❌ Error creating record: {str(create_error)}")
                                raise create_error

                    # ✅ Case 2: lines processing
                    lines = department.get('lines', [])
                    for line in lines:
                        line_id = line.get('id')
                        
                        # ✅ Case 2a: stations directly under line
                        line_stations = line.get('stations', [])
                        if line_stations:
                            print(f"📌 Found {len(line_stations)} stations directly under line {line_id}")
                            for station in line_stations:
                                station_id = station.get('id')
                                print(f"💾 Creating record: dept={dept_id}, line={line_id}, station={station_id}")
                                
                                try:
                                    record = HierarchyStructure.objects.create(
                                        structure_name=structure_name,
                                        hq_id=hq_id,
                                        factory_id=factory_id,
                                        department_id=dept_id,
                                        line_id=line_id,
                                        subline_id=None,
                                        station_id=station_id
                                    )
                                    created_records.append(record)
                                    print(f"✅ Successfully created line-station record with ID: {record.structure_id}")
                                except Exception as create_error:
                                    print(f"❌ Error creating line-station record: {str(create_error)}")
                                    raise create_error

                        # ✅ Case 2b: sublines processing
                        sublines = line.get('sublines', [])
                        if not sublines:
                            # Create line-only record if no sublines and no stations
                            if not line_stations:
                                print(f"📌 Saving line {line_id} under department {dept_id}")
                                try:
                                    record = HierarchyStructure.objects.create(
                                        structure_name=structure_name,
                                        hq_id=hq_id,
                                        factory_id=factory_id,
                                        department_id=dept_id,
                                        line_id=line_id,
                                        subline_id=None,
                                        station_id=None
                                    )
                                    created_records.append(record)
                                except Exception as create_error:
                                    print(f"❌ Error creating line record: {str(create_error)}")
                                    raise create_error
                        else:
                            for subline in sublines:
                                subline_id = subline.get('id')
                                stations = subline.get('stations', [])

                                if not stations:
                                    # Create subline-only record if no stations
                                    print(f"📌 Saving subline {subline_id} under line {line_id}")
                                    try:
                                        record = HierarchyStructure.objects.create(
                                            structure_name=structure_name,
                                            hq_id=hq_id,
                                            factory_id=factory_id,
                                            department_id=dept_id,
                                            line_id=line_id,
                                            subline_id=subline_id,
                                            station_id=None
                                        )
                                        created_records.append(record)
                                    except Exception as create_error:
                                        print(f"❌ Error creating subline record: {str(create_error)}")
                                        raise create_error
                                else:
                                    for station in stations:
                                        station_id = station.get('id')
                                        print(f"📌 Saving station {station_id} under subline {subline_id}")
                                        try:
                                            record = HierarchyStructure.objects.create(
                                                structure_name=structure_name,
                                                hq_id=hq_id,
                                                factory_id=factory_id,
                                                department_id=dept_id,
                                                line_id=line_id,
                                                subline_id=subline_id,
                                                station_id=station_id
                                            )
                                            created_records.append(record)
                                        except Exception as create_error:
                                            print(f"❌ Error creating station record: {str(create_error)}")
                                            raise create_error

                print(f"🎯 Total created records: {len(created_records)}")
                
                # Double-check all records exist before serializing
                for record in created_records:
                    db_record = HierarchyStructure.objects.filter(structure_id=record.structure_id).first()
                    if not db_record:
                        raise Exception(f"Record {record.structure_id} not found in database after creation")

                # Serialize all created records
                serializer = self.get_serializer(created_records, many=True)
                print("✅ Created hierarchy records:", len(serializer.data))
                
                # Final verification before returning
                final_count = HierarchyStructure.objects.filter(structure_name=structure_name).count()
                print(f"🔍 Final count in DB for structure '{structure_name}': {final_count}")
                
                return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            print(f"❌ Exception occurred: {str(e)}")
            logger.error(f"Failed to create hierarchy structure: {str(e)}")
            return Response({'error': 'Failed to create hierarchy structure', 'details': str(e)},
                            status=status.HTTP_400_BAD_REQUEST)

    def list(self, request, *args, **kwargs):
        """Return hierarchy records (including dept-level stations)"""
        queryset = self.filter_queryset(self.get_queryset())

        # Show:
        #   - Full hierarchy (dept + line + subline + station)
        #   - Dept + station (when no line/subline exist)
        #   - Line + station (when no subline exist)
        queryset = queryset.filter(
            station__isnull=False,
            department__isnull=False
        )

        # Unique structures
        seen_structures = set()
        unique_structures = []
        for record in queryset:
            if record.structure_name not in seen_structures:
                unique_structures.append(record)
                seen_structures.add(record.structure_name)

        serializer = self.get_serializer(unique_structures, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        """Filter queryset based on query params"""
        queryset = super().get_queryset()
        for field in ['hq', 'factory', 'department', 'line', 'subline', 'station', 'structure_name']:
            value = self.request.query_params.get(field)
            if value:
                queryset = queryset.filter({f"{field}_id" if field != 'structure_name' else field: value})
        return queryset

    def _find_first_station(self, structure_data):
        """Helper to find first station in nested data"""
        departments = structure_data.get('departments', [])
        for department in departments:
            dept_id = department.get('id')

            # Handle case: station directly under department
            for station in department.get('stations', []):
                return {'department_id': dept_id, 'line_id': None,
                        'subline_id': None, 'station_id': station.get('id')}

            # Handle case: station directly under line
            for line in department.get('lines', []):
                line_id = line.get('id')
                for station in line.get('stations', []):
                    return {'department_id': dept_id, 'line_id': line_id,
                            'subline_id': None, 'station_id': station.get('id')}

                # Normal flow: line → subline → station
                for subline in line.get('sublines', []):
                    subline_id = subline.get('id')
                    stations = subline.get('stations', [])
                    if stations:
                        station_id = stations[0].get('id')
                        return {'department_id': dept_id, 'line_id': line_id,
                                'subline_id': subline_id, 'station_id': station_id}
        return None


from rest_framework.decorators import action

# ------------------ Mastertable Views ------------------
import pandas as pd
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser ,JSONParser
from io import BytesIO
import json
from datetime import datetime
from .models import MasterTable, Department
from .serializers import MasterTableSerializer

class MasterTableViewSet(viewsets.ModelViewSet):
    queryset = MasterTable.objects.all()
    serializer_class = MasterTableSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download Excel template with headers only"""
        try:
            # Create DataFrame with column headers matching the model fields
            headers = [
                'emp_id',
                'first_name', 
                'last_name',
                'department_name',  # We'll use department name in template
                'date_of_joining',
                'birth_date',
                'sex',
                'email',
                'phone'
            ]
            
            # Create empty DataFrame with headers
            df = pd.DataFrame(columns=headers)
            
            # Add example row to show format
            example_row = {
                'emp_id': 'EMP001',
                'first_name': 'John',
                'last_name': 'Doe',
                'department_name': 'IT Department',
                'date_of_joining': '2024-01-15',
                'birth_date': '1990-05-20',
                'sex': 'M',
                'email': 'john.doe@company.com',
                'phone': '+1234567890'
            }
            
            # Add example row and then clear it (keeps formatting)
            df.loc[0] = example_row
            df = df.iloc[0:0]  # Remove the example row, keep structure
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Employee Template', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Employee Template']
                
                # Add instructions in a separate sheet
                instructions_df = pd.DataFrame({
                    'Instructions': [
                        '1. Fill in employee details in the Employee Template sheet',
                        '2. emp_id: Must be unique (e.g., EMP001, EMP002)',
                        '3. first_name, last_name: Employee names',
                        '4. department_name: Exact department name (case sensitive)',
                        '5. date_of_joining: Format YYYY-MM-DD (e.g., 2024-01-15)',
                        '6. birth_date: Format YYYY-MM-DD (optional)',
                        '7. sex: M for Male, F for Female, O for Other',
                        '8. email: Must be unique and valid email format',
                        '9. phone: Include country code (e.g., +1234567890)',
                        '',
                        'Available Departments:'
                    ]
                })
                
                # Add available departments to instructions
                departments = Department.objects.all().values_list('department_name', flat=True)
                for dept in departments:
                    instructions_df = pd.concat([
                        instructions_df,
                        pd.DataFrame({'Instructions': [f'- {dept}']})
                    ], ignore_index=True)
                
                instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Failed to generate template: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        """Upload Excel file and create or update employee records"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file format. Please upload Excel file (.xlsx or .xls)'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            df = pd.read_excel(file, sheet_name='Employee Template')
            
            required_columns = ['emp_id', 'first_name', 'email', 'date_of_joining']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            df = df.dropna(subset=['emp_id'])
            
            if df.empty:
                return Response(
                    {'error': 'No valid employee data found in the file'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            created_employees = []
            updated_employees = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    department = None
                    if pd.notna(row.get('department_name')):
                        try:
                            department = Department.objects.get(
                                department_name=str(row['department_name']).strip()
                            )
                        except Department.DoesNotExist:
                            errors.append({
                                'row': index + 2,
                                'emp_id': row.get('emp_id', 'N/A'),
                                'error': f'Department "{row["department_name"]}" not found'
                            })
                            continue
                    
                    emp_id = str(row['emp_id']).strip()
                    
                    def clean_field(val):
                        if pd.isna(val):
                            return ''
                        if isinstance(val, str) and val.strip() == '':
                            return ''
                        return val
                    
                    employee_data = {
                        'emp_id': emp_id,
                        'first_name': clean_field(row.get('first_name', '')),
                        'last_name': clean_field(row.get('last_name', '')),
                        'department': department.department_id if department else None,
                        'email': str(row['email']).strip().lower(),
                        'date_of_joining': pd.to_datetime(row['date_of_joining']).date(),
                        'birth_date': pd.to_datetime(row['birth_date']).date() if pd.notna(row.get('birth_date')) else None,
                        'sex': str(row.get('sex', '')).upper() if pd.notna(row.get('sex')) else '',
                        'phone': clean_field(row.get('phone', '')),
                    }
                    
                    if employee_data['sex'] and employee_data['sex'] not in ['M', 'F', 'O', '']:
                        errors.append({
                            'row': index + 2,
                            'emp_id': emp_id,
                            'error': f'Invalid sex value "{employee_data["sex"]}". Use M, F, O, or leave blank'
                        })
                        continue
                    
                    try:
                        employee = MasterTable.objects.get(emp_id=emp_id)
                        
                        # Check if all relevant fields are filled (non-empty)
                        fields_to_check = ['first_name', 'last_name', 'department', 'email', 'date_of_joining', 'birth_date', 'sex', 'phone']
                        all_filled = all([
                            getattr(employee, f) not in [None, '', []]
                            for f in fields_to_check
                        ])
                        
                        if all_filled:
                            errors.append({
                                'row': index + 2,
                                'emp_id': emp_id,
                                'error': 'Same emp_id cannot be uploaded again as all fields are already filled.'
                            })
                            continue
                        
                        # Else update only blank fields with new data
                        updated_fields = []
                        for field in employee_data:
                            new_val = employee_data[field]
                            old_val = getattr(employee, field, None)
                            if (old_val in [None, '', []]) and new_val not in [None, '', []]:
                                setattr(employee, field, new_val)
                                updated_fields.append(field)
                        if updated_fields:
                            employee.save(update_fields=updated_fields)
                            updated_employees.append({
                                'emp_id': employee.emp_id,
                                'name': f"{employee.first_name} {employee.last_name}".strip(),
                                'email': employee.email
                            })
                    except MasterTable.DoesNotExist:
                        serializer = MasterTableSerializer(data=employee_data)
                        if serializer.is_valid():
                            employee = serializer.save()
                            created_employees.append({
                                'emp_id': employee.emp_id,
                                'name': f"{employee.first_name} {employee.last_name}".strip(),
                                'email': employee.email
                            })
                        else:
                            error_messages = []
                            for field, field_errors in serializer.errors.items():
                                if isinstance(field_errors, list):
                                    error_messages.extend([f"{field}: {error}" for error in field_errors])
                                else:
                                    error_messages.append(f"{field}: {field_errors}")
                            errors.append({
                                'row': index + 2,
                                'emp_id': emp_id,
                                'error': '; '.join(error_messages)
                            })
                except Exception as e:
                    errors.append({
                        'row': index + 2,
                        'emp_id': row.get('emp_id', 'N/A'),
                        'error': str(e)
                    })
            
            response_data = {
                'message': f'Upload completed. {len(created_employees)} employees created, {len(updated_employees)} employees updated.',
                'created_count': len(created_employees),
                'updated_count': len(updated_employees),
                'error_count': len(errors),
                'created_employees': created_employees,
                'updated_employees': updated_employees
            }
            
            if errors:
                response_data['errors'] = errors
            
            return Response(response_data, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<emp_id>[^/.]+)')
    def retrieve_by_employee_code(self, request, emp_id=None):
        try:
           employee = MasterTable.objects.get(emp_id=emp_id)
           serializer = self.get_serializer(employee)
           return Response(serializer.data)
        except MasterTable.DoesNotExist:
           return Response({"error": "Employee not found"}, status=404)


# Level 0

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.views import APIView
from .models import UserRegistration,HumanBodyQuestions,HumanBodyCheckSession,HumanBodyCheckSheet
from .serializers import UserRegistrationSerializer,UserWithBodyCheckSerializer,HumanBodyCheckSessionSerializer,HumanBodyQuestionsSerializer,UserUpdateSerializer

from rest_framework import filters
class UserRegistrationViewSet(viewsets.ModelViewSet):
    queryset = UserRegistration.objects.all()
    serializer_class = UserRegistrationSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['first_name', 'last_name', 'email', 'phone_number']
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    lookup_field = 'temp_id'
    
    def get_serializer_class(self):
        if self.request.method in ['PATCH', 'PUT']:
            return UserUpdateSerializer
        return UserRegistrationSerializer

class HumanBodyQuestionsViewSet(viewsets.ModelViewSet):
    queryset = HumanBodyQuestions.objects.all()
    serializer_class = HumanBodyQuestionsSerializer

    @action(detail=False, methods=['post'])
    def add_question(self, request):
        
        question_text = request.data.get('question_text', '').strip()
        if not question_text:
            return Response({"error": "question_text is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        question, created = HumanBodyQuestions.objects.get_or_create(
            question_text=question_text,
        )
        
        if created:
            serializer = self.get_serializer(question)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response({"error": "Question already exists"}, status=status.HTTP_400_BAD_REQUEST)


class BodyCheckSubmissionView(APIView):
    
    
    def get(self, request):
        
        temp_id = request.query_params.get('temp_id')
        if temp_id:
            
            try:
                user = UserRegistration.objects.get(temp_id=temp_id)
                
                session = HumanBodyCheckSession.objects.filter(user=user).order_by('-created_at').first()
                if not session:
                    
                    session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).order_by('-created_at').first()
            except UserRegistration.DoesNotExist:
                
                session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).order_by('-created_at').first()
            
            if session:
                serializer = HumanBodyCheckSessionSerializer(session)
                return Response([serializer.data])
            return Response([])
        
        
        sessions = HumanBodyCheckSession.objects.all()
        serializer = HumanBodyCheckSessionSerializer(sessions, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        
        temp_id = request.data.get('temp_id')
        check_data = request.data.get('checkData', {})
        
        if not temp_id:
            return Response({"error": "temp_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = UserRegistration.objects.get(temp_id=temp_id)
            
            
            existing_session = HumanBodyCheckSession.objects.filter(user=user).first()
            if existing_session:
                return Response(
                    {"error": "Body check already exists for this user"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            
            existing_temp_session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).first()
            if existing_temp_session:
                return Response(
                    {"error": "Body check already exists for this user"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            
            session = HumanBodyCheckSession.objects.create(temp_id=temp_id, user=user)
            
           
            for item_id, item_data in check_data.items():
                description = item_data.get('description', '')
                answer = item_data.get('status', 'pending')
                
                if description and answer != '':
                   
                    question, created = HumanBodyQuestions.objects.get_or_create(
                        question_text=description
                    )
                    
                   
                    HumanBodyCheckSheet.objects.create(
                        session=session,
                        question=question,
                        answer=answer
                    )
            
            serializer = HumanBodyCheckSessionSerializer(session)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except UserRegistration.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)


class UserBodyCheckListView(APIView):
    
    def get(self, request):
        users = UserRegistration.objects.all()
        serializer = UserWithBodyCheckSerializer(users, many=True)
        return Response(serializer.data)
    
    def patch(self, request):
        """Mark user as added to master table"""
        temp_id = request.data.get('temp_id')
        
        if not temp_id:
            return Response({"error": "temp_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = UserRegistration.objects.get(temp_id=temp_id)
            
            # Check if user has passed status
            latest_session = HumanBodyCheckSession.objects.filter(user=user).order_by('-created_at').first()
            if not latest_session:
                latest_session = HumanBodyCheckSession.objects.filter(temp_id=temp_id).order_by('-created_at').first()
            
            if not latest_session or latest_session.overall_status != 'pass':
                return Response(
                    {"error": "User must have passed status to be added to master table"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if user.is_added_to_master:
                return Response(
                    {"error": "User already added to master table"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            from django.utils import timezone
            user.is_added_to_master = True
            user.added_to_master_at = timezone.now()
            user.save()
            
            # Return updated user list
            users = UserRegistration.objects.all()
            serializer = UserWithBodyCheckSerializer(users, many=True)
            return Response(serializer.data)
            
        except UserRegistration.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)




# views.py
from rest_framework import viewsets
from .models import (
    Level, Days, SubTopic, SubTopicContent, TrainingContent, Evaluation
)
from .serializers import ( LevelSerializer, DaysSerializer,DaysWriteSerializer, 
     SubTopicListSerializer, SubTopicAdminSerializer, SubTopicSerializer, SubTopicContentSerializer,
    TrainingContentSerializer, EvaluationSerializer
)

def _int_or_none(v):
    try:
        return int(str(v).strip("/"))
    except (TypeError, ValueError):
        return None


class LevelViewSet(viewsets.ModelViewSet):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer


class DaysViewSet(viewsets.ModelViewSet):
    queryset = Days.objects.all().select_related("level")
    # serializer_class = DaysSerializer

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return DaysWriteSerializer
        return DaysSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        level = _int_or_none(self.request.query_params.get("level"))
        if level is not None:
            qs = qs.filter(level__level_id=level)
        return qs


from rest_framework import viewsets
from .models import Level, Days, SubTopic, SubTopicContent, TrainingContent, Evaluation
from .serializers import (
    LevelSerializer, DaysSerializer,
    SubTopicSerializer, SubTopicListSerializer, SubTopicAdminSerializer,
    SubTopicContentSerializer, TrainingContentSerializer, EvaluationSerializer
)

class SubTopicViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all().select_related("days", "level")
      
    def get_serializer_class(self):
        if self.action in ["list", "retrieve"]:
            return SubTopicListSerializer  # read shape for frontend
        return SubTopicAdminSerializer  # subtopic_name, days, level
     
    def get_queryset(self):
        qs = super().get_queryset()
        level = self.request.query_params.get("level")   # /subtopics/?level=1
        days_id = self.request.query_params.get("days")  # /subtopics/?days=3
        
        if level is not None:
            qs = qs.filter(level__level_id=level)
        if days_id is not None:
            qs = qs.filter(days__days_id=days_id)
        return qs
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicSerializer


class SubTopicContentViewSet(viewsets.ModelViewSet):
    queryset = SubTopicContent.objects.all()
    serializer_class = SubTopicContentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        subtopic = self.request.query_params.get("subtopic")  # /subtopic-contents/?subtopic=12
        if subtopic is not None:
            qs = qs.filter(subtopic__subtopic_id=subtopic)
        return qs
    
    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx
    

from rest_framework import viewsets, parsers


class TrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer

    parser_classes = [parsers.MultiPartParser, parsers.FormParser, parsers.JSONParser]

    # def get_queryset(self):
    #     qs = super().get_queryset()
    #     stc = self.request.query_params.get("subtopiccontent")  # /training-contents/?subtopiccontent=55
    #     if stc:
    #         qs = qs.filter(subtopiccontent__subtopiccontent_id=stc)
    #     return qs

    def get_queryset(self):
        qs = super().get_queryset()
        # stc = self.request.query_params.get("subtopiccontent") or self.request.query_params.get("subtopic_content")
        stc = _int_or_none(self.request.query_params.get("subtopiccontent") or self.request.query_params.get("subtopic_content"))
        # if stc is not None:
        #     try:
        #         stc = int(str(stc).strip("/"))
        #         qs = qs.filter(subtopiccontent__subtopiccontent_id=stc)
        #     except ValueError:
        #         pass
        # return qs
        if stc is not None:
            qs = qs.filter(subtopiccontent__subtopiccontent_id=stc)
        return qs


    def get_serializer_context(self):
        # so training_file returns absolute URL
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx


class EvaluationViewSet(viewsets.ModelViewSet):
    queryset = Evaluation.objects.all()
    serializer_class = EvaluationSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import ProductionPlan
from .serializers import ProductionPlanSerializer

class ProductionPlanViewSet(viewsets.ModelViewSet):
    queryset = ProductionPlan.objects.all()
    serializer_class = ProductionPlanSerializer

    def create(self, request, *args, **kwargs):
        # Handle bulk creation
        if isinstance(request.data, list):
            serializer = self.get_serializer(data=request.data, many=True)
            serializer.is_valid(raise_exception=True)
            self.perform_bulk_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return super().create(request, *args, **kwargs)

    def perform_bulk_create(self, serializer):
        serializer.save()

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params

        # Filtering based on new structure
        hq = params.get('hq')
        factory = params.get('factory')
        department = params.get('department')
        line = params.get('line')
        subline = params.get('subline')
        station = params.get('station')
        year = params.get('year')
        month = params.get('month')

        if hq:
            queryset = queryset.filter(hq_id=hq)
        if factory:
            queryset = queryset.filter(factory_id=factory)
        if department:
            queryset = queryset.filter(department_id=department)
        if line:
            queryset = queryset.filter(line_id=line)
        if subline:
            queryset = queryset.filter(subline_id=subline)
        if station:
            queryset = queryset.filter(station_id=station)
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)

        return queryset


from rest_framework import viewsets
from .models import QuestionPaper
from .serializers import QuestionPaperSerializer

class QuestionPaperViewSet(viewsets.ModelViewSet):
    queryset = QuestionPaper.objects.all().order_by("-created_at")
    serializer_class = QuestionPaperSerializer

    @action(detail=True, methods=["get"])
    def questions(self, request, pk=None):
        # Get the paper
        paper = self.get_object()
        paper_serializer = QuestionPaperSerializer(paper)

        # Get related questions
        questions = TemplateQuestion.objects.filter(question_paper_id=pk)
        question_serializer = TemplateQuestionSerializer(questions, many=True)

        # Return combined response
        return Response({
            "question_paper": paper_serializer.data,
            "questions": question_serializer.data
        })




from rest_framework import viewsets
from .models import StationLevelQuestionPaper
from .serializers import StationLevelQuestionPaperSerializer


class StationLevelQuestionPaperViewSet(viewsets.ModelViewSet):
    queryset = StationLevelQuestionPaper.objects.all()
    serializer_class = StationLevelQuestionPaperSerializer

import io
import pandas as pd
from django.http import HttpResponse
from django.core.exceptions import ValidationError as DjangoValidationError
from django.shortcuts import get_object_or_404
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError as DRFValidationError

from .models import TemplateQuestion, QuestionPaper
from .serializers import TemplateQuestionSerializer
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

class TemplateQuestionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Template Questions, including bulk upload & download template.
    """
    queryset = TemplateQuestion.objects.all()
    serializer_class = TemplateQuestionSerializer
    parser_classes = [MultiPartParser, FormParser]

 
    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """
        Generates and serves an Excel template with Department → Line → Subline → Station → Level header
        fetched from QuestionPaper instance.
        """
        # Get question_paper_id from query params
        question_paper_id = request.query_params.get('question_paper_id')
        
        if not question_paper_id:
            return Response(
                {'detail': 'question_paper_id is required as a query parameter.'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            question_paper_id = int(question_paper_id)
        except (ValueError, TypeError):
            return Response(
                {'detail': 'Invalid question_paper_id format.'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Fetch QuestionPaper with related fields
        try:
            question_paper = get_object_or_404(
                QuestionPaper.objects.select_related(
                    'department', 'line', 'subline', 'station', 'level'
                ),
                question_paper_id=question_paper_id
            )
        except Exception as e:
            return Response(
                {'detail': f'QuestionPaper not found: {str(e)}'}, 
                status=status.HTTP_404_NOT_FOUND
            )

        # Example static sample data (for demo)
        sample_data = [
            {
                "question": "What is the capital of France?",
                "option_a": "Paris",
                "option_b": "London",
                "option_c": "Berlin",
                "option_d": "Madrid",
                "correct_answer": "Paris",
            },
            {
                "question": "Which planet is known as the Red Planet?",
                "option_a": "Earth",
                "option_b": "Venus",
                "option_c": "Mars",
                "option_d": "Jupiter",
                "correct_answer": "Mars",
            },
        ]

        df = pd.DataFrame(sample_data)

        # Write DataFrame to memory
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Template_Questions")

        buffer.seek(0)

        # Load workbook for header insertion
        wb = load_workbook(buffer)
        ws = wb["Template_Questions"]

        # === Fetch header fields from QuestionPaper instance ===
        department = str(question_paper.department)
        line = str(question_paper.line)
        subline = str(question_paper.subline)
        station = str(question_paper.station)
        level = str(question_paper.level)

        # Build header text
        header_text = (
            f"Department: {department} | Line: {line} | Subline: {subline} | "
            f"Station: {station} | Level: {level}"
        )
       

        # Insert header row above the question table
        ws.insert_rows(1, amount=2)  # make space for header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        cell = ws.cell(row=1, column=1, value=header_text)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 25

        # Save final Excel
        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        # Include question paper name in filename for better identification
        filename = f"Template_Questions_{question_paper.question_paper_name.replace(' ', '_')}.xlsx"

        response = HttpResponse(
            final_buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response


    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, *args, **kwargs):
        """
        Handles bulk creation of Template Questions from an Excel file.
        """
        file_obj = request.FILES.get('file')
        question_paper_id = request.data.get('question_paper_id')

        if not question_paper_id:
            return Response({'detail': 'question_paper_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            question_paper_id = int(question_paper_id)
        except (ValueError, TypeError):
            return Response({'detail': 'Invalid question_paper_id format.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Verify that the QuestionPaper exists using its actual primary key field name
        try:
            # Using .pk is robust, but using the specific field name is also fine
            question_paper = get_object_or_404(QuestionPaper, pk=question_paper_id)
        except Exception as e:
            return Response(
                {'detail': f'QuestionPaper not found: {str(e)}'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            # Assuming the template has a header, skip first 2 rows. 
            # header=2 means the 3rd row is the header.
            df = pd.read_excel(file_obj, sheet_name='Template_Questions', header=2, engine='openpyxl')
            df = df.where(pd.notnull(df), None)
            df.dropna(subset=['question'], inplace=True)
        except Exception as e:
            return Response(
                {'detail': f"Error reading the Excel file. Ensure it contains a sheet named 'Template_Questions'. Error: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        required_columns = {
            'question', 'option_a', 'option_b', 'option_c',
            'option_d', 'correct_answer'
        }
        
        if not required_columns.issubset(df.columns):
            missing_cols = required_columns - set(df.columns)
            return Response({'detail': f'File is missing required columns: {", ".join(missing_cols)}'}, status=status.HTTP_400_BAD_REQUEST)

        questions_to_create = []
        errors = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            
            # === THE FIX IS HERE ===
            # Use the correct primary key of the question_paper object
            # Using .pk is the best practice as it works regardless of the field's name
            row_data['question_paper'] = question_paper.pk 
            
            serializer = self.get_serializer(data=row_data)
            
            try:
                serializer.is_valid(raise_exception=True)
                questions_to_create.append(serializer.validated_data)
            except (DRFValidationError, DjangoValidationError) as e:
                excel_row_number = index + 4 # 2 header rows + 1 table header row + 1 for 0-index
                error_detail = serializer.errors if hasattr(serializer, 'errors') and serializer.errors else str(e)
                errors.append({'row': excel_row_number, 'errors': error_detail})
        
        if errors:
            return Response({
                'status': 'Upload failed due to validation errors.',
                'created_count': 0,
                'error_count': len(errors),
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)

        if questions_to_create:
            model_instances = [TemplateQuestion(**data) for data in questions_to_create]
            TemplateQuestion.objects.bulk_create(model_instances)

        response_data = {
            'status': 'Upload successful.',
            'created_count': len(questions_to_create),
            'error_count': len(errors),
        }

        return Response(response_data, status=status.HTTP_201_CREATED)
   

    def get_queryset(self):
        queryset = super().get_queryset()
        question_paper_id = self.request.query_params.get("question_paper")
        if question_paper_id:
            queryset = queryset.filter(question_paper_id=question_paper_id)
        return queryset

    from rest_framework import viewsets
from rest_framework.parsers import JSONParser, FormParser, MultiPartParser
from .models import Hq, Factory, Department, Line, SubLine, Station
from .serializers import HqSerializer, FactorySerializer, DepartmentSerializer, LineSerializer, SubLineSerializer, StationSerializer

# ------------------ ViewSets ------------------
class HqViewSet(viewsets.ModelViewSet):
    queryset = Hq.objects.all()
    serializer_class = HqSerializer
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_serializer(self, *args, **kwargs):
        if hasattr(self.request, "data"):
            print("Incoming data:", self.request.data)
        return super().get_serializer(*args, **kwargs)

    def create(self, request, *args, **kwargs):
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            response = Response()
            origin = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Origin'] = origin
            response['Access-Control-Allow-Methods'] = 'POST, PUT, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'
            return response

        try:
            print("[HQ CREATE] headers:", dict(request.headers))
            print("[HQ CREATE] content_type:", request.content_type)
            print("[HQ CREATE] data:", request.data)
            print("[HQ CREATE] raw body:", request.body)
        except Exception:
            pass
        resp = super().create(request, *args, **kwargs)
        # Add CORS headers
        origin = request.headers.get('Origin', '*')
        resp['Access-Control-Allow-Origin'] = origin
        resp['Access-Control-Allow-Credentials'] = 'true'
        resp['Vary'] = 'Origin'
        return resp

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            response = Response()
            origin = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Origin'] = origin
            response['Access-Control-Allow-Methods'] = 'PUT, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'
            return response

        try:
            print(f"[HQ UPDATE partial={partial}] headers:", dict(request.headers))
            print("[HQ UPDATE] content_type:", request.content_type)
            print("[HQ UPDATE] data:", request.data)
            print("[HQ UPDATE] raw body:", request.body)
        except Exception:
            pass
        resp = super().update(request, *args, **kwargs)
        # Add CORS headers
        origin = request.headers.get('Origin', '*')
        resp['Access-Control-Allow-Origin'] = origin
        resp['Access-Control-Allow-Credentials'] = 'true'
        resp['Vary'] = 'Origin'
        return resp


class FactoryViewSet(viewsets.ModelViewSet):
    queryset = Factory.objects.all()
    serializer_class = FactorySerializer
    
    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Content-Type:", request.content_type)
        print("Raw data:", request.data)
        print("User:", request.user)
        print("Authenticated:", request.user.is_authenticated)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        # Make a mutable copy of the QueryDict
        data = request.data.copy()
        print("Data copy:", data)
        
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Factory created successfully!")
            return response
            
        except Exception as e:
            print("Error creating factory:", str(e))
            response = Response(
                {"error": "Failed to create factory", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response
                                            

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Department Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Department created successfully!")
            return response
            
        except Exception as e:
            print("Error creating department:", str(e))
            response = Response(
                {"error": "Failed to create department", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Line Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Line created successfully!")
            return response
            
        except Exception as e:
            print("Error creating line:", str(e))
            response = Response(
                {"error": "Failed to create line", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


class SubLineViewSet(viewsets.ModelViewSet):
    queryset = SubLine.objects.all()
    serializer_class = SubLineSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming SubLine Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("SubLine created successfully!")
            return response
            
        except Exception as e:
            print("Error creating subline:", str(e))
            response = Response(
                {"error": "Failed to create subline", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer

    def create(self, request, *args, **kwargs):
        print("\n=== Incoming Station Request ===")
        print("Method:", request.method)
        print("Headers:", dict(request.headers))
        print("Data:", request.data)
        
        # Handle CORS preflight
        if request.method == 'OPTIONS':
            print("Handling OPTIONS request")
            response = Response()
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-Requested-With'
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Access-Control-Max-Age'] = '86400'  # 24 hours
            return response
            
        try:
            # Call parent's create method
            response = super().create(request, *args, **kwargs)
            
            # Add CORS headers to the response
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            response['Vary'] = 'Origin'
            
            print("Station created successfully!")
            return response
            
        except Exception as e:
            print("Error creating station:", str(e))
            response = Response(
                {"error": "Failed to create station", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            return response


# ------------------ AR/VR Views ------------------
from rest_framework import viewsets
from .models import ARVRTrainingContent
from .serializers import ARVRTrainingContentSerializer

class ARVRTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = ARVRTrainingContent.objects.all()
    serializer_class = ARVRTrainingContentSerializer


# --------------------------
# Level 2 Process Dojo
# --------------------------

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import LevelWiseTrainingContent, TrainingTopic
from .serializers import LevelWiseTrainingContentSerializer, TrainingTopicSerializer

class LevelWiseTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = LevelWiseTrainingContent.objects.all()
    serializer_class = LevelWiseTrainingContentSerializer

    # Custom endpoint: filter by level, station, and topic
    @action(detail=False, methods=["get"])
    def by_level_station_topic(self, request):
        level_id = request.query_params.get("level_id")
        station_id = request.query_params.get("station_id")
        topic_id = request.query_params.get("topic_id")

        queryset = self.queryset
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if station_id:
            queryset = queryset.filter(station_id=station_id)
        if topic_id:
            queryset = queryset.filter(topic_id=topic_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = TrainingTopic.objects.all()
    serializer_class = TrainingTopicSerializer
    
    def get_queryset(self):
        """Filter topics by level and station from query parameters by default"""
        queryset = TrainingTopic.objects.all()
        level_id = self.request.query_params.get("level_id")
        station_id = self.request.query_params.get("station_id")
        
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if station_id:
            queryset = queryset.filter(station_id=station_id)
            
        return queryset
    
    # Custom endpoint: filter topics by level and/or station (keeping for backward compatibility)
    @action(detail=False, methods=["get"])
    def by_level_station(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

# hanchou & shokucho 



# your_app/views.py

from rest_framework import viewsets
from .models import HanContent, HanSubtopic, HanTrainingContent
from .serializers import (
    HanContentDetailSerializer,
    HanContentListSerializer,
    HanSubtopicSerializer,
    HanTrainingContentSerializer
)

class HanContentViewSet(viewsets.ModelViewSet):
    """
    Handles listing, creating, retrieving, updating, and deleting Main Topics.
    """
    # prefetch_related is a performance optimization that prevents many small database queries.
    queryset = HanContent.objects.prefetch_related('subtopics__materials').all()

    def get_serializer_class(self):
        """
        Chooses the serializer based on the action.
        - For retrieving a single item ('retrieve'), use the detailed serializer.
        - For listing all items ('list'), use the simple serializer.
        """
        if self.action == 'retrieve':
            return HanContentDetailSerializer
        return HanContentListSerializer


# --- MODIFIED VIEWSET FOR SUBTOPICS ---
class HanSubtopicViewSet(viewsets.ModelViewSet):
    serializer_class = HanSubtopicSerializer

    def get_queryset(self):
        """
        Filters subtopics based on a query parameter from the URL.
        Example: GET /api/subtopics/?han_content_id=5
        """
        queryset = HanSubtopic.objects.all()
        han_content_id = self.request.query_params.get('han_content_id')
        if han_content_id:
            queryset = queryset.filter(han_content_id=han_content_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/subtopics/ with body {"title": "...", "han_content": 5}
        """
        # The serializer will handle associating the parent, since the ID is in the data.
        serializer.save()

from django.http import FileResponse


class HanTrainingContentViewSet(viewsets.ModelViewSet):
    serializer_class = HanTrainingContentSerializer

    def get_queryset(self):
        """
        Filters materials based on a query parameter from the URL.
        Example: GET /api/materials/?han_subtopic_id=12
        """
        queryset = HanTrainingContent.objects.all()
        subtopic_id = self.request.query_params.get('han_subtopic_id')
        if subtopic_id:
            queryset = queryset.filter(han_subtopic_id=subtopic_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/materials/ with body {"description": "...", "han_subtopic": 12}
        """
        serializer.save()


def serve_han_material_file(request, pk):
    """
    Serves the protected media file for a HanTrainingContent object.
    """
    material = get_object_or_404(HanTrainingContent, pk=pk)

    if not material.training_file:
        raise Http404("No file found for this material.")

    try:
        return FileResponse(material.training_file.open('rb'), as_attachment=False)
    except FileNotFoundError:
        raise Http404("File does not exist on the server.")


from rest_framework import viewsets
from django.shortcuts import get_object_or_404
from django.http import FileResponse
from .models import ShoContent, ShoSubtopic, ShoTrainingContent
from .serializers import (
    ShoContentListSerializer,
    ShoContentDetailSerializer,
    ShoSubtopicSerializer,
    ShoTrainingContentSerializer
)


# --- SHO CONTENT VIEWSET ---
class ShoContentViewSet(viewsets.ModelViewSet):
    """
    Handles listing, creating, retrieving, updating, and deleting Main Topics.
    """
    queryset = ShoContent.objects.prefetch_related('sho_subtopics__sho_materials').all()

    def get_serializer_class(self):
        """
        Chooses the serializer based on the action.
        - For retrieving a single item ('retrieve'), use the detailed serializer.
        - For listing all items ('list'), use the simple serializer.
        """
        if self.action == 'retrieve':
            return ShoContentDetailSerializer
        return ShoContentListSerializer


# --- SHO SUBTOPIC VIEWSET ---
class ShoSubtopicViewSet(viewsets.ModelViewSet):
    serializer_class = ShoSubtopicSerializer

    def get_queryset(self):
        """
        Filters subtopics based on a query parameter from the URL.
        Example: GET /api/sho-subtopics/?sho_content_id=5
        """
        queryset = ShoSubtopic.objects.all()
        sho_content_id = self.request.query_params.get('sho_content_id')
        if sho_content_id:
            queryset = queryset.filter(sho_content_id=sho_content_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/sho-subtopics/ with body {"title": "...", "sho_content": 5}
        """
        serializer.save()


# --- SHO TRAINING CONTENT VIEWSET ---
class ShoTrainingContentViewSet(viewsets.ModelViewSet):
    serializer_class = ShoTrainingContentSerializer

    def get_queryset(self):
        """
        Filters materials based on a query parameter from the URL.
        Example: GET /api/sho-materials/?sho_subtopic_id=12
        """
        queryset = ShoTrainingContent.objects.all()
        subtopic_id = self.request.query_params.get('sho_subtopic_id')
        if subtopic_id:
            queryset = queryset.filter(sho_subtopic_id=subtopic_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/sho-materials/ with body {"sho_description": "...", "sho_subtopic": 12}
        """
        serializer.save()


def serve_sho_material_file(request, pk):
    """
    Serves the protected media file for a ShoTrainingContent object.
    """
    material = get_object_or_404(ShoTrainingContent, pk=pk)

    if not material.training_file:
        raise Http404("No file found for this material.")

    try:
        return FileResponse(material.training_file.open('rb'), as_attachment=False)
    except FileNotFoundError:
        raise Http404("File does not exist on the server.")







from rest_framework import viewsets
from .models import HanchouExamQuestion
from .serializers import HanchouExamQuestionSerializer

class HanchouExamQuestionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Hanchou Exam Questions, including bulk upload.
    """
    queryset = HanchouExamQuestion.objects.all()
    serializer_class = HanchouExamQuestionSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """
        Generates and serves an Excel template with sample data for Hanchou questions.
        """
        sample_data = [
            {
                'question': 'What is the largest mammal in the world?',
                'option_a': 'Elephant',
                'option_b': 'Blue Whale',
                'option_c': 'Giraffe',
                'option_d': 'Great White Shark',
                'correct_answer': 'Blue Whale'
            },
            {
                'question': 'Which element has the atomic number 1?',
                'option_a': 'Helium',
                'option_b': 'Oxygen',
                'option_c': 'Hydrogen',
                'option_d': 'Carbon',
                'correct_answer': 'Hydrogen'
            }
        ]
        
        df = pd.DataFrame(sample_data)
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Using a specific sheet name is good practice
            df.to_excel(writer, index=False, sheet_name='Hanchou_Questions')
            
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Hanchou_Question_Upload_Template.xlsx"'
        
        return response

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, *args, **kwargs):
        """
        Handles bulk creation of Hanchou questions from an Excel file.
        """
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # The sheet_name must match the one used in download_template
            df = pd.read_excel(file_obj, sheet_name='Hanchou_Questions', engine='openpyxl')
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            return Response({'detail': f"Error reading the Excel file. Ensure it contains a sheet named 'Hanchou_Questions'. Error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = {
            'question', 'option_a', 'option_b', 'option_c',
            'option_d', 'correct_answer'
        }
        
        if not required_columns.issubset(df.columns):
            missing_cols = required_columns - set(df.columns)
            return Response({'detail': f'File is missing required columns: {", ".join(missing_cols)}'}, status=status.HTTP_400_BAD_REQUEST)

        questions_to_create = []
        errors = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            serializer = self.get_serializer(data=row_data)
            
            try:
                # Validate using the model's clean() method
                instance = HanchouExamQuestion(**row_data)
                instance.clean()
                
                # Also run standard serializer validation
                serializer.is_valid(raise_exception=True)
                questions_to_create.append(serializer.validated_data)

            except (DRFValidationError, DjangoValidationError, TypeError) as e:
                error_detail = serializer.errors if hasattr(serializer, 'errors') and serializer.errors else str(e)
                errors.append({'row': index + 2, 'errors': error_detail})
        
        if questions_to_create:
            model_instances = [HanchouExamQuestion(**data) for data in questions_to_create]
            HanchouExamQuestion.objects.bulk_create(model_instances)

        response_data = {
            'status': 'Upload complete.',
            'created_count': len(questions_to_create),
            'error_count': len(errors),
            'errors': errors
        }

        return Response(response_data, status=status.HTTP_201_CREATED)
    

from rest_framework import viewsets, permissions, filters as drf_filters
from django_filters import rest_framework as filters
from .models import HanchouExamResult
from .serializers import HanchouExamResultSerializer

class HanchouExamResultFilter(filters.FilterSet):
    pay_code = filters.CharFilter(field_name="employee__pay_code", lookup_expr="iexact")
    name = filters.CharFilter(field_name="employee__name", lookup_expr="icontains")
    exam_date = filters.DateFromToRangeFilter(field_name="exam_date")
    submitted_at = filters.DateFromToRangeFilter(field_name="submitted_at")
    passed = filters.BooleanFilter()

    class Meta:
        model = HanchouExamResult
        fields = ["employee", "passed", "exam_date"]

class HanchouExamResultViewSet(viewsets.ModelViewSet):
    queryset = HanchouExamResult.objects.select_related("employee").all().order_by("-submitted_at", "-started_at")
    serializer_class = HanchouExamResultSerializer
    # permission_classes = [permissions.IsAuthenticated]  # adjust as needed

    filter_backends = [filters.DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = HanchouExamResultFilter
    search_fields = ["employee__name", "employee__pay_code", "employee__card_no", "remarks"]
    ordering_fields = ["submitted_at", "exam_date", "score", "total_questions", "duration_seconds"]
    ordering = ["-submitted_at", "-started_at"]

from rest_framework import viewsets
from .models import ShokuchouExamQuestion,ShokuchouExamResult
from .serializers import ShokuchouExamQuestionSerializer,ShokuchouExamResultSerializer

# Add these imports at the top of your views.py
import pandas as pd
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError

# Your existing model and serializer
from .models import ShokuchouExamQuestion
from .serializers import ShokuchouExamQuestionSerializer


class ShokuchouExamQuestionViewSet(viewsets.ModelViewSet):
    queryset = ShokuchouExamQuestion.objects.all()
    serializer_class = ShokuchouExamQuestionSerializer
    # Add parser classes for the ViewSet to handle file uploads
    parser_classes = [JSONParser, MultiPartParser, FormParser]



    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """
        Generates and serves an Excel template with sample data for bulk uploading questions.
        """
        # Define the sample data
        sample_data = [
            {
                'sho_question': 'What is the capital of Japan?',
                'sho_option_a': 'Seoul',
                'sho_option_b': 'Beijing',
                'sho_option_c': 'Tokyo',
                'sho_option_d': 'Bangkok',
                'sho_correct_answer': 'Tokyo'
            },
            {
                'sho_question': 'Which planet is known as the Red Planet?',
                'sho_option_a': 'Earth',
                'sho_option_b': 'Mars',
                'sho_option_c': 'Jupiter',
                'sho_option_d': 'Saturn',
                'sho_correct_answer': 'Mars'
            }
        ]
        
        # Create a pandas DataFrame
        df = pd.DataFrame(sample_data)
        
        # Use an in-memory buffer
        buffer = io.BytesIO()
        
        # Write the DataFrame to the buffer as an Excel file
        # index=False prevents pandas from writing row indices to the file
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Questions')
            
        # Set the buffer's pointer to the beginning
        buffer.seek(0)
        
        # Create the HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Question_Upload_Template.xlsx"'
        
        return response







    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, *args, **kwargs):
        """
        Handles bulk creation of questions from an Excel file upload.
        The Excel file must have a sheet named 'Questions' and columns:
        'sho_question', 'sho_option_a', 'sho_option_b', 'sho_option_c',
        'sho_option_d', 'sho_correct_answer'.
        """
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file_obj.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': 'Invalid file format. Please upload an Excel file (.xlsx, .xls).'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Explicitly read the 'Questions' sheet
            df = pd.read_excel(file_obj, sheet_name='Questions', engine='openpyxl')
            # Replace NaN values with None for proper serialization
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            # Catches errors like missing sheet or unreadable file
            return Response(
                {'detail': f"Error reading the Excel file. Make sure it contains a sheet named 'Questions'. Error: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        required_columns = {
            'sho_question', 'sho_option_a', 'sho_option_b', 'sho_option_c',
            'sho_option_d', 'sho_correct_answer'
        }
        
        if not required_columns.issubset(df.columns):
            missing_cols = required_columns - set(df.columns)
            return Response(
                {'detail': f'File is missing required columns: {", ".join(missing_cols)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        questions_to_create = []
        errors = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            serializer = self.get_serializer(data=row_data)
            
            try:
                # The model's clean() method is not called by default in DRF serializers.
                # We can simulate it by creating a model instance and calling full_clean().
                # This ensures our custom validation rule is checked.
                instance = ShokuchouExamQuestion(**row_data)
                instance.clean() # This will call our custom validation
                
                # We also run serializer validation for data types, max_length, etc.
                serializer.is_valid(raise_exception=True)
                questions_to_create.append(serializer.validated_data)

            except (DRFValidationError, DjangoValidationError, TypeError) as e:
                error_detail = serializer.errors if hasattr(serializer, 'errors') and serializer.errors else str(e)
                errors.append({'row': index + 2, 'errors': error_detail}) # +2 because index is 0-based and header is row 1
        
        # Now, create the actual model instances for bulk_create
        if questions_to_create:
            model_instances = [ShokuchouExamQuestion(**data) for data in questions_to_create]
            ShokuchouExamQuestion.objects.bulk_create(model_instances)

        response_data = {
            'status': 'Upload complete.',
            'created_count': len(questions_to_create),
            'error_count': len(errors),
            'errors': errors
        }

        return Response(response_data, status=status.HTTP_201_CREATED)






class ShokuchouExamResultFilter(filters.FilterSet):
    pay_code = filters.CharFilter(field_name="employee__pay_code", lookup_expr="iexact")
    name = filters.CharFilter(field_name="employee__name", lookup_expr="icontains")
    sho_submitted_at = filters.DateFromToRangeFilter(field_name="sho_submitted_at")
    sho_passed = filters.BooleanFilter()

    class Meta:
        model = ShokuchouExamResult
        fields = ["employee", "sho_passed", "sho_submitted_at"]


# --- SHO RESULT VIEWSET ---
class ShokuchouExamResultViewSet(viewsets.ModelViewSet):
    queryset = (
        ShokuchouExamResult.objects.select_related("employee")
        .all()
        .order_by("-sho_submitted_at", "-sho_started_at")
    )
    serializer_class = ShokuchouExamResultSerializer
    # permission_classes = [permissions.IsAuthenticated]  # enable if needed

    filter_backends = [filters.DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = ShokuchouExamResultFilter
    search_fields = ["employee__name", "employee__pay_code", "employee__card_no", "sho_remarks"]
    ordering_fields = [
        "sho_submitted_at",
        "sho_score",
        "sho_total_questions",
        "sho_duration_seconds",
    ]
    ordering = ["-sho_submitted_at", "-sho_started_at"]


class HanchouResultCertificatePDF(APIView):
    def get(self, request, pk):
        try:
            result = HanchouExamResult.objects.select_related('employee').get(pk=pk)
            
            employee_name = result.employee.name.strip()
            exam_name = result.exam_name.strip()

            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            # Define colors and margins
            main_color = (85/255, 26/255, 139/255)
            shadow_color = (200/255, 200/255, 200/255)
            border_color = (244/255, 145/255, 34/255)
            margin = 0.5 * inch

            # Draw the decorative borders
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- DRAW CERTIFICATE CONTENT ---
            
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            draw_reflected_text(p, width / 2.0, height - 2.4*inch, "HANCHOUE EXAM CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY PASSED THE")
            
            draw_reflected_text(p, width / 2.0, height - 5.9*inch, exam_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # Signature
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "TRAINER SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            buffer.seek(0)
            return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="Hanchou_Certificate_{employee_name}.pdf"'})

        except HanchouExamResult.DoesNotExist:
            return Response({"error": "Hanchou exam result not found"}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated employee for this result could not be found."}, status=status.HTTP_404_NOT_FOUND)



# Make sure you have all necessary imports at the top of your views.py file
import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch

# from .models import Score # Ensure Score model is imported

# This helper function should be defined or imported in your views.py
def draw_reflected_text(p, x, y, text, font_name, font_size, main_color, shadow_color, y_offset=2):
    """Draws centered text with a subtle shadow/reflection effect."""
    # Shadow
    p.setFont(font_name, font_size)
    p.setFillColorRGB(*shadow_color)
    p.drawCentredString(x + y_offset, y - y_offset, text)
    # Main Text
    p.setFillColorRGB(*main_color)
    p.drawCentredString(x, y, text)



# --- Place the draw_reflected_text function here ---
def draw_reflected_text(p, x, y, text, font, size, main_color, shadow_color, x_offset=1.5, y_offset=1.5):
    """Draws text with a simple shadow/reflection underneath."""
    # Draw shadow/reflection
    p.setFont(font, size)
    p.setFillColorRGB(*shadow_color)
    p.drawCentredString(x + x_offset, y - y_offset, text)
    
    # Draw main text
    p.setFillColorRGB(*main_color)
    p.drawCentredString(x, y, text)
# ----------------------------------------------------


class ShokuchouResultCertificatePDF(APIView):
    def get(self, request, pk):
        try:
            # CHANGED: Query ShokuchouExamResult instead of HanchouExamResult
            result = ShokuchouExamResult.objects.select_related('employee').get(pk=pk)
            
            employee_name = result.employee.name.strip()
            # CHANGED: Use sho_exam_name field
            exam_name = result.sho_exam_name.strip()

            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            # Define colors and margins (kept the same for consistent branding)
            main_color = (85/255, 26/255, 139/255)
            shadow_color = (200/255, 200/255, 200/255)
            border_color = (244/255, 145/255, 34/255)
            margin = 0.5 * inch

            # Draw the decorative borders
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- DRAW CERTIFICATE CONTENT ---
            
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            # CHANGED: Updated certificate title
            draw_reflected_text(p, width / 2.0, height - 2.4*inch, "SHOKUCHOU EXAM CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY PASSED THE")
            
            draw_reflected_text(p, width / 2.0, height - 5.9*inch, exam_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # Signature
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "TRAINER SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            buffer.seek(0)
            # CHANGED: Updated the filename for the download
            return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="Shokuchou_Certificate_{employee_name}.pdf"'})

        # CHANGED: Catch DoesNotExist for the correct model
        except ShokuchouExamResult.DoesNotExist:
            return Response({"error": "Shokuchou exam result not found"}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated employee for this result could not be found."}, status=status.HTTP_404_NOT_FOUND)


# 10 cycle

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from collections import defaultdict
from django.shortcuts import get_object_or_404

from .models import (
    TenCycleDayConfiguration,
    TenCycleTopics,
    TenCycleSubTopic,
    TenCyclePassingCriteria,
    OperatorPerformanceEvaluation,
    EvaluationSubTopicMarks,
    MasterTable
)
from .serializers import (
    TenCycleDayConfigurationSerializer,
    TenCycleTopicsSerializer,
    TenCycleSubTopicSerializer,
    TenCyclePassingCriteriaSerializer,
    OperatorPerformanceEvaluationSerializer,
    EvaluationSubMarksSerializer
)


class TenCycleSubTopicViewSet(viewsets.ModelViewSet):
    queryset = TenCycleSubTopic.objects.all()
    serializer_class = TenCycleSubTopicSerializer
    
    @action(detail=False, methods=['get'], url_path='by-topic')
    def by_topic(self, request):
        topic_id = request.query_params.get('topic_id')
        
        if not topic_id:
            return Response(
                {"error": "topic_id is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        queryset = self.queryset.filter(topic_id=topic_id, is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TenCycleDayConfigurationViewSet(viewsets.ModelViewSet):
    queryset = TenCycleDayConfiguration.objects.all()
    serializer_class = TenCycleDayConfigurationSerializer
    
    @action(detail=False, methods=['get'], url_path='get-configuration')
    def get_configuration(self, request):
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        filters = {
            'level_id': level_id,
            'department_id': department_id,
            'is_active': True
        }
        
        if station_id:
            filters['station_id'] = station_id
        else:
            filters['station__isnull'] = True
            
        queryset = self.queryset.filter(**filters)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TenCycleTopicsViewSet(viewsets.ModelViewSet):
    queryset = TenCycleTopics.objects.all()
    serializer_class = TenCycleTopicsSerializer

    @action(detail=True, methods=['get'])
    def subtopics(self, request, pk=None):
        topic = self.get_object()
        subtopics = topic.subtopics.all()
        serializer = TenCycleSubTopicSerializer(subtopics, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='get-configuration')
    def get_configuration(self, request):
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        filters = {
            'level_id': level_id,
            'department_id': department_id,
            'is_active': True
        }
        
        if station_id:
            filters['station_id'] = station_id
        else:
            filters['station__isnull'] = True
            
        queryset = self.queryset.filter(**filters)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class TenCyclePassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = TenCyclePassingCriteria.objects.all()
    serializer_class = TenCyclePassingCriteriaSerializer
    
    @action(detail=False, methods=['get'], url_path='get-configuration')
    def get_configuration(self, request):
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if station_id:
            try:
                criteria = self.queryset.get(
                    level_id=level_id,
                    department_id=department_id,
                    station_id=station_id,
                    is_active=True
                )
                serializer = self.get_serializer(criteria)
                return Response(serializer.data)
            except TenCyclePassingCriteria.DoesNotExist:
                pass
        
        try:
            criteria = self.queryset.get(
                level_id=level_id,
                department_id=department_id,
                station__isnull=True,
                is_active=True
            )
            serializer = self.get_serializer(criteria)
            return Response(serializer.data)
        except TenCyclePassingCriteria.DoesNotExist:
            return Response(
                {"error": "No passing criteria found for the given configuration"}, 
                status=status.HTTP_404_NOT_FOUND
            )

class TenCycleConfigurationViewSet(viewsets.ViewSet):
    @action(detail=False, methods=['get'], url_path='complete-configuration')
    def complete_configuration(self, request):
        """Get complete configuration including days, topics, subtopics, and passing criteria"""
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')
        
        if not level_id or not department_id:
            return Response(
                {"error": "level_id and department_id are required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        filters = {
            'level_id': level_id,
            'department_id': department_id,
            'is_active': True
        }
        
        if station_id:
            filters['station_id'] = station_id
        else:
            filters['station__isnull'] = True
        
        # Get days configuration
        days = TenCycleDayConfiguration.objects.filter(**filters)
        days_data = TenCycleDayConfigurationSerializer(days, many=True).data
        
        # Get topics configuration
        topics = TenCycleTopics.objects.filter(**filters)
        topics_data = TenCycleTopicsSerializer(topics, many=True).data
        
        # Get subtopics for each topic
        for topic_data in topics_data:
            topic_id = topic_data['id']
            subtopics = TenCycleSubTopic.objects.filter(topic_id=topic_id, is_active=True)
            topic_data['subtopics'] = TenCycleSubTopicSerializer(subtopics, many=True).data
        
        # Get passing criteria
        passing_criteria = None
        try:
            if station_id:
                criteria = TenCyclePassingCriteria.objects.get(
                    level_id=level_id,
                    department_id=department_id,
                    station_id=station_id,
                    is_active=True
                )
            else:
                criteria = TenCyclePassingCriteria.objects.get(
                    level_id=level_id,
                    department_id=department_id,
                    station__isnull=True,
                    is_active=True
                )
            passing_criteria = TenCyclePassingCriteriaSerializer(criteria).data
        except TenCyclePassingCriteria.DoesNotExist:
            # Set default passing criteria
            passing_criteria = {
                'passing_percentage': 60.0,
                'level': level_id,
                'department': department_id,
                'station': station_id
            }
        
        return Response({
            'days': days_data,
            'topics': topics_data,
            'passing_criteria': passing_criteria,
            'level_id': level_id,
            'department_id': department_id,
            'station_id': station_id
        })

# class OperatorPerformanceEvaluationViewSet(viewsets.ModelViewSet):
#     queryset = OperatorPerformanceEvaluation.objects.all()
#     serializer_class = OperatorPerformanceEvaluationSerializer

#     @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
#     def by_employee_code(self, request, employee_code=None):
#         employee = get_object_or_404(MasterTable, emp_id=employee_code)
#         evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)
#         serializer = self.get_serializer(evaluations, many=True)
#         return Response(serializer.data)


from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404

class OperatorPerformanceEvaluationViewSet(viewsets.ModelViewSet):
    queryset = OperatorPerformanceEvaluation.objects.all()
    serializer_class = OperatorPerformanceEvaluationSerializer

    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
    def by_employee_code(self, request, employee_code=None):
        employee = get_object_or_404(MasterTable, emp_id=employee_code)

        # Get filter parameters from the query string
        level_id = request.query_params.get('level_id')
        department_id = request.query_params.get('department_id')
        station_id = request.query_params.get('station_id')

        evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)

        if level_id:
            evaluations = evaluations.filter(level_id=level_id)
        if department_id:
            evaluations = evaluations.filter(department_id=department_id)
        if station_id:
            evaluations = evaluations.filter(station_id=station_id)


        serializer = self.get_serializer(evaluations, many=True)
        return Response(serializer.data)



class EvaluationSubTopicMarksViewSet(viewsets.ModelViewSet):
    queryset = EvaluationSubTopicMarks.objects.all()
    serializer_class = EvaluationSubMarksSerializer

    @action(detail=False, methods=['get'], url_path='by-employee-code/(?P<employee_code>[^/.]+)')
    def by_employee_code(self, request, employee_code=None):
      employee = get_object_or_404(MasterTable, emp_id=employee_code)
      evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)

      marks_qs = EvaluationSubTopicMarks.objects.filter(employee__in=evaluations)
      serializer = self.get_serializer(marks_qs, many=True)

      days_map = defaultdict(list)
      for mark_obj, mark_data in zip(marks_qs, serializer.data):
        days_map[mark_obj.day.day_name].append((mark_obj, mark_data))

      per_day_results = []
      total_score = 0
      total_possible_score = 0

      for day_name, marks_list in days_map.items():
        day_score = 0
        day_max_score = 0
        for mark_obj, mark_data in marks_list:
            subtopic_score = sum(mark_data.get(f'mark_{i}', 0) or 0 for i in range(1, 11))
            day_score += subtopic_score
            max_score = mark_obj.subtopic.score_required * 10
            day_max_score += max_score

        passing_percentage = 60.0  # or fetch dynamically 

        status_str = "Pass" if (day_max_score > 0 and (day_score / day_max_score) * 100 >= passing_percentage) else "Fail - Retraining Required"
        per_day_results.append({
            "day": day_name,
            "score": day_score,
            "max_score": day_max_score,
            "passing_percentage": passing_percentage,
            "status": status_str
        })

        total_score += day_score
        total_possible_score += day_max_score

      final_status = "Not Evaluated"
      if total_possible_score > 0:
        overall_percentage = (total_score / total_possible_score) * 100
        final_status = "Pass" if overall_percentage >= passing_percentage else "Fail - Retraining Required"

      
      return Response({
        'employee_code': employee_code,
        'evaluations': serializer.data,   # <-- this sends full detailed marks list
        'per_day_results': per_day_results,
        'total_score': total_score,
        'total_possible_score': total_possible_score,
        'final_status': final_status,
    })


    @action(detail=False, methods=['put'], url_path='update-mark/(?P<employee_code>[^/.]+)')
    def update_by_employee_code(self, request, employee_code=None):
        employee = get_object_or_404(MasterTable, emp_id=employee_code)
        employee_eval_qs = OperatorPerformanceEvaluation.objects.filter(employee=employee)
        evaluations = EvaluationSubTopicMarks.objects.filter(employee__in=employee_eval_qs)

        topic_id = request.data.get('subtopic')
        day_id = request.data.get('day')
        if not topic_id or not day_id:
            return Response({"error": "subtopic and day IDs are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            mark_instance = evaluations.get(subtopic_id=topic_id, day_id=day_id)
        except EvaluationSubTopicMarks.DoesNotExist:
            return Response({"error": "Matching mark record not found."}, status=404)

        serializer = EvaluationSubMarksSerializer(mark_instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    
    @action(detail=False, methods=['get'], url_path='daily-total/(?P<employee_code>[^/.]+)')
    def daily_total(self, request, employee_code=None):
        employee = get_object_or_404(MasterTable, emp_id=employee_code)
        evaluations = OperatorPerformanceEvaluation.objects.filter(employee=employee)
        marks_qs = EvaluationSubTopicMarks.objects.filter(employee__in=evaluations)
        serializer = self.get_serializer(marks_qs, many=True)

        days_map = defaultdict(list)
        for mark_obj, mark_data in zip(marks_qs, serializer.data):
            days_map[mark_obj.day.day_name].append(mark_data)

        per_day_results = []
        grand_total = 0
        grand_max_total = 0

        for day_name, marks_list in days_map.items():
            day_score = 0
            day_max_score = 0
            for mark in marks_list:
                day_score += mark.get('total_score', 0) or 0
                day_max_score += mark.get('max_possible_score', 0) or 0
            grand_total += day_score
            grand_max_total += day_max_score
            per_day_results.append({
                'day': day_name,
                'total_score': day_score,
                'max_possible_score': day_max_score,
                'percentage': round((day_score / day_max_score)*100, 2) if day_max_score > 0 else 0
            })

        final_percentage = round((grand_total / grand_max_total)*100, 2) if grand_max_total > 0 else 0

        if evaluations.exists():
            eval_obj = evaluations.first()
            try:
                pass_crit = TenCyclePassingCriteria.objects.get(
                    level=eval_obj.level,
                    department=eval_obj.department,
                    station=eval_obj.station
                )
            except TenCyclePassingCriteria.DoesNotExist:
                pass_crit = TenCyclePassingCriteria.objects.filter(
                    level=eval_obj.level,
                    department=eval_obj.department,
                    station__isnull=True
                ).first()
            passing_percentage = pass_crit.passing_percentage if pass_crit else 60.0
        else:
            passing_percentage = 60.0  # default fallback

        final_status = "Pass" if final_percentage >= passing_percentage else "Fail - Retraining Required"

        return Response({
            'employee_code': employee_code,
            'per_day_totals': per_day_results,
            'grand_total_score': grand_total,
            'grand_max_score': grand_max_total,
            'final_percentage': final_percentage,
            'passing_percentage': passing_percentage,
            'final_status': final_status
        })
        
#=================================  10 cycle   ================================#


# ======================== Machine Allocation Approval ============ #


from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .serializers import MachineAllocationApprovalSerializer
from .models import MachineAllocation


class MachineAllocationApprovalViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationApprovalSerializer

    @action(detail=True, methods=['put'], url_path='set-status')
    def set_status(self, request, pk=None):
        allocation = self.get_object()
        status_value = request.data.get('approval_status')

        if status_value not in dict(MachineAllocation.APPROVAL_STATUS_CHOICES):
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

        allocation.approval_status = status_value
        allocation.save()
        return Response({
            'status': 'success',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        })

    @action(detail=True, methods=['put'], url_path='reject')
    def reject(self, request, pk=None):
        allocation = self.get_object()
        allocation.approval_status = 'rejected'
        allocation.save()
        return Response({
            'status': 'rejected',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        }, status=status.HTTP_200_OK)



from .serializers import EmployeeWithStatusSerializer

class EmployeeMachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = ...  # your main MachineAllocation serializer

    @action(detail=False, methods=['get'], url_path='eligible-employees')
    def eligible_employees(self, request):
        machine_id = request.query_params.get('machine_id')
        if not machine_id:
            return Response({'error': 'machine_id is required'}, status=400)

        try:
            machine = Machine.objects.get(id=machine_id)
        except Machine.DoesNotExist:
            return Response({'error': 'Machine not found'}, status=404)

        # matching_skills = OperatorSkill.objects.filter(station__skill=machine.process) # skilmatrix table
        # employee_ids = matching_skills.values_list('operator_id', flat=True).distinct()
        # employees = MasterTable.objects.filter(id__in=employee_ids)

        # serializer = EmployeeWithStatusSerializer(employees, many=True, context={'machine_id': machine_id})
        # return Response(serializer.data)


# =========================== Machine Allocation Approval =============================== # 

from rest_framework import viewsets
from .models import OJTTopic
from .serializers import OJTTopicSerializer

class OJTTopicViewSet(viewsets.ModelViewSet):
    queryset = OJTTopic.objects.all().order_by("sl_no")
    serializer_class = OJTTopicSerializer





from rest_framework import viewsets
from .models import OJTDay
from .serializers import OJTDaySerializer

class OJTDayViewSet(viewsets.ModelViewSet):
    queryset = OJTDay.objects.all().order_by("id")
    serializer_class = OJTDaySerializer









from rest_framework import viewsets
from .models import OJTScore
from .serializers import OJTScoreSerializer

class OJTScoreViewSet(viewsets.ModelViewSet):
    queryset = OJTScore.objects.all()
    serializer_class = OJTScoreSerializer

from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import OJTScoreRange
from .serializers import OJTScoreRangeSerializer


class OJTScoreRangeViewSet(viewsets.ModelViewSet):
    queryset = OJTScoreRange.objects.all().order_by("id")
    serializer_class = OJTScoreRangeSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset

from rest_framework import viewsets
from .models import OJTPassingCriteria
from .serializers import OJTPassingCriteriaSerializer


class OJTPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = OJTPassingCriteria.objects.all()
    serializer_class = OJTPassingCriteriaSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import TraineeInfo
from .serializers import TraineeInfoSerializer


# from django.db.models import Sum
# from rest_framework import viewsets, status
# from rest_framework.response import Response
# from .models import TraineeInfo, OJTScore, OJTTopic, OJTScoreRange, OJTPassingCriteria
# from .serializers import TraineeInfoSerializer


# class TraineeInfoViewSet(viewsets.ModelViewSet):
#     """
#     API endpoint to create, update, list, and delete trainees with nested OJT scores.
#     Status is updated after saving scores.
#     """
#     queryset = TraineeInfo.objects.all()
#     serializer_class = TraineeInfoSerializer

#     def perform_recalculate_status(self, trainee):
#         """Recalculate trainee status based on scores and criteria."""
#         if not trainee.scores.exists():
#             return trainee.status

#         for score in trainee.scores.all():
#             department = score.topic.department
#             level = score.topic.level
#             day = score.day

#             try:
#                 score_range = OJTScoreRange.objects.get(department=department, level=level)
#             except OJTScoreRange.DoesNotExist:
#                 continue  # skip if no range defined

#             total_score = OJTScore.objects.filter(
#                 trainee=trainee,
#                 topic__department=department,
#                 topic__level=level,
#                 day=day
#             ).aggregate(total=Sum("score"))["total"] or 0

#             total_topics = OJTTopic.objects.filter(department=department, level=level).count()

#             if total_topics > 0:
#                 max_possible = total_topics * score_range.max_score
#                 percentage = (total_score / max_possible) * 100
#             else:
#                 percentage = 0

#             # Passing criteria lookup
#             criteria = (
#                 OJTPassingCriteria.objects.filter(department=department, level=level, day=day).first()
#                 or OJTPassingCriteria.objects.filter(department=department, level=level, day__isnull=True).first()
#             )

#             if criteria and percentage >= criteria.percentage:
#                 trainee.status = "Pass"
#             else:
#                 trainee.status = "Fail"

#             trainee.save(update_fields=["status"])
#         return trainee.status

#     def create(self, request, *args, **kwargs):
#         serializer = self.get_serializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
#         trainee = serializer.save()

#         # update status after save
#         self.perform_recalculate_status(trainee)

#         return Response(serializer.data, status=status.HTTP_201_CREATED)

#     def update(self, request, *args, **kwargs):
#         partial = kwargs.pop('partial', False)
#         instance = self.get_object()
#         serializer = self.get_serializer(instance, data=request.data, partial=partial)
#         serializer.is_valid(raise_exception=True)
#         trainee = serializer.save()

#         # update status after update
#         self.perform_recalculate_status(trainee)

#         return Response(serializer.data)



from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import TraineeInfo
from .serializers import TraineeInfoSerializer


from django.db.models import Sum
from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import TraineeInfo, OJTScore, OJTTopic, OJTScoreRange, OJTPassingCriteria
from .serializers import TraineeInfoSerializer
from .signals import run_after_delay,update_skill_matrix

class TraineeInfoViewSet(viewsets.ModelViewSet):
    """
    API endpoint to create, update, list, and delete trainees with nested OJT scores.
    Status is updated after saving scores.
    """
    queryset = TraineeInfo.objects.all()
    serializer_class = TraineeInfoSerializer

    def get_queryset(self):
        """
        Filters the queryset based on 'emp_id' and 'station_id' from the query parameters.
        """
        queryset = super().get_queryset()

        # Get the parameters sent by the React frontend
        emp_id = self.request.query_params.get('emp_id')
        station_id = self.request.query_params.get('station_id')

        # If an emp_id was provided in the URL, filter by it
        if emp_id:
            queryset = queryset.filter(emp_id=emp_id)

        # If a station_id was provided in the URL, filter by it
        if station_id:
            queryset = queryset.filter(station_id=station_id)

        return queryset

    # def perform_recalculate_status(self, trainee):
    #     """Recalculate trainee status based on scores and criteria."""
    #     if not trainee.scores.exists():
    #         return trainee.status

    #     for score in trainee.scores.all():
    #         department = score.topic.department
    #         level = score.topic.level
    #         day = score.day

    #         try:
    #             score_range = OJTScoreRange.objects.get(department=department, level=level)
    #         except OJTScoreRange.DoesNotExist:
    #             continue  # skip if no range defined

    #         total_score = OJTScore.objects.filter(
    #             trainee=trainee,
    #             topic__department=department,
    #             topic__level=level,
    #             day=day
    #         ).aggregate(total=Sum("score"))["total"] or 0

    #         total_topics = OJTTopic.objects.filter(department=department, level=level).count()

    #         if total_topics > 0:
    #             max_possible = total_topics * score_range.max_score
    #             percentage = (total_score / max_possible) * 100
    #         else:
    #             percentage = 0

    #         # Passing criteria lookup
    #         criteria = (
    #             OJTPassingCriteria.objects.filter(department=department, level=level, day=day).first()
    #             or OJTPassingCriteria.objects.filter(department=department, level=level, day__isnull=True).first()
    #         )

    #         if criteria and percentage >= criteria.percentage:
    #             trainee.status = "Pass"
    #         else:
    #             trainee.status = "Fail"

    #         trainee.save(update_fields=["status"])
    #     return trainee.status


    def perform_recalculate_status(self, trainee):
        """
        Recalculate trainee status.
        Status is 'Pass' only if ALL required OJT days meet the passing criteria.
        Status is 'Fail' if ANY required OJT day fails the passing criteria.
        Status is 'Pending' if not all required OJT days have been scored.
        """
        # Determine the Department and Level from the trainee's associated topics/scores.
        # Assuming the department/level is consistent across all topics for a trainee.
        first_score = trainee.scores.first()
        if not first_score:
            trainee.status = "Pending"
            trainee.save(update_fields=["status"])
            return trainee.status

        department = first_score.topic.department
        level = first_score.topic.level

        # 1. Get all REQUIRED OJT Days for this Department and Level
        required_days = OJTDay.objects.filter(department=department, level=level).distinct()
        
        # If no days are defined, assume 'Pending' or a special case might be needed.
        if not required_days.exists():
            trainee.status = "Pending"
            trainee.save(update_fields=["status"])
            return trainee.status

        # 2. Get the Score Range and Max Score per Topic
        try:
            score_range = OJTScoreRange.objects.get(department=department, level=level)
            max_topic_score = score_range.max_score
        except OJTScoreRange.DoesNotExist:
            # Cannot calculate status without a score range.
            trainee.status = "Pending"
            trainee.save(update_fields=["status"])
            return trainee.status

        all_days_passed = True
        
        for day in required_days:
            # 3. Get the Passing Criteria for the current Day (or general criteria)
            criteria = (
                OJTPassingCriteria.objects.filter(department=department, level=level, day=day).first()
                or OJTPassingCriteria.objects.filter(department=department, level=level, day__isnull=True).first()
            )

            # If no criteria, we can't determine pass/fail for this day.
            if not criteria:
                # Treat as 'Pending' if we can't determine pass/fail for a required day
                trainee.status = "Pending"
                trainee.save(update_fields=["status"])
                return trainee.status

            required_percentage = criteria.percentage

            # 4. Calculate total score and max possible score for this Day
            
            # Topics for this Day's OJT (assuming ALL topics apply to ALL days if not filtered)
            # A more robust system might need OJTTopic.objects.filter(department=department, level=level, day=day)
            # but based on your models, all topics are for a department/level, so we'll use that:
            relevant_topics = OJTTopic.objects.filter(department=department, level=level)
            total_topics = relevant_topics.count()

            # Sum of actual scores for this Trainee for this Day
            daily_actual_score = OJTScore.objects.filter(
                trainee=trainee,
                day=day,
                topic__in=relevant_topics  # Ensures we only sum scores for relevant topics
            ).aggregate(total=Sum("score"))["total"] or 0
            
            # Check how many topics for this day have been scored.
            # If the number of scores is less than the total topics, the day is incomplete.
            scores_count = OJTScore.objects.filter(trainee=trainee, day=day, topic__in=relevant_topics).count()
            
            # If not all scores are in, the overall status is "Pending" and we can stop.
            if scores_count < total_topics:
                trainee.status = "Pending"
                trainee.save(update_fields=["status"])
                return trainee.status
            
            # 5. Calculate Percentage
            max_possible_score = total_topics * max_topic_score
            
            if max_possible_score > 0:
                daily_percentage = (daily_actual_score / max_possible_score) * 100
            else:
                daily_percentage = 0 # Can happen if total_topics is 0
            
            # 6. Check Pass/Fail for this Day
            if daily_percentage < required_percentage:
                all_days_passed = False
                break  # Fail on the first failed day

        # 7. Set final Trainee Status
        if all_days_passed:
            trainee.status = "Pass"
        else:
            trainee.status = "Fail"

        trainee.save(update_fields=["status"])
        return trainee.status


    def trigger_skill_matrix_update(self, trainee):
        """Fires the skill matrix update after status is calculated."""
        # Only trigger if the final status is 'Pass'
        if trainee.status == "Pass":
            try:
                employee = MasterTable.objects.get(emp_id=trainee.emp_id)
                station_object = trainee.station
                
                # Since TraineeInfo doesn't directly hold the Level, we infer it from a score
                first_score = OJTScore.objects.filter(trainee=trainee).first()
                if not first_score: return

                level_object = first_score.topic.level
                
                # Trigger the check after the status is definitely saved
                run_after_delay(update_skill_matrix, 5, employee, station_object, level_object, True)
            except Exception as e:
                print(f"[ERROR] Failed to trigger skill matrix update: {e}")


    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        trainee = serializer.save()
        self.perform_recalculate_status(trainee)
        self.trigger_skill_matrix_update(trainee) # <--- ADDED HERE
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        trainee = serializer.save()
        self.perform_recalculate_status(trainee)
        self.trigger_skill_matrix_update(trainee) # <--- ADDED HERE
        return Response(serializer.data)






# views.py
from rest_framework.generics import ListAPIView
from .models import OJTTopic
from .serializers import OJTTopicSerializer

class OJTTopicListView(ListAPIView):
    serializer_class = OJTTopicSerializer

    def get_queryset(self):
        queryset = OJTTopic.objects.all()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset
    


from rest_framework.generics import ListAPIView
from .models import OJTDay
from .serializers import OJTDaySerializer

class OJTDayListView(ListAPIView):
    serializer_class = OJTDaySerializer

    def get_queryset(self):
        queryset = OJTDay.objects.all()
        department_id = self.request.query_params.get("department")
        level_id = self.request.query_params.get("level")

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)

        return queryset
    

from rest_framework import generics
from .models import TraineeInfo
from .serializers import TraineeInfoSerializer

class TraineeInfoListView(generics.ListAPIView):
    serializer_class = TraineeInfoSerializer

    def get_queryset(self):
        queryset = TraineeInfo.objects.all()
        trainer_id = self.request.query_params.get("trainer_id")
        station = self.request.query_params.get("station")

        if trainer_id:
            queryset = queryset.filter(trainer_id=trainer_id)
        if station:
            queryset = queryset.filter(station=station)

        return queryset
    

from rest_framework import viewsets
from .models import QuantityOJTScoreRange, QuantityPassingCriteria
from .serializers import QuantityOJTScoreRangeSerializer, QuantityPassingCriteriaSerializer


class QuantityOJTScoreRangeViewSet(viewsets.ModelViewSet):
    queryset = QuantityOJTScoreRange.objects.all()
    serializer_class = QuantityOJTScoreRangeSerializer


class QuantityPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = QuantityPassingCriteria.objects.all()
    serializer_class = QuantityPassingCriteriaSerializer

    
from rest_framework import viewsets
from .models import OJTLevel2Quantity, Level
from .serializers import OJTLevel2QuantitySerializer, LevelSerializer


class OJTLevel2QuantityViewSet(viewsets.ModelViewSet):
    """
    API endpoint for OJT Level 2 Quantity with nested evaluations.
    Supports filtering by trainee_id, level, and station.
    """
    queryset = OJTLevel2Quantity.objects.all()
    serializer_class = OJTLevel2QuantitySerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        trainee_id = self.request.query_params.get("trainee_id")
        level_id = self.request.query_params.get("level")
        station_name = self.request.query_params.get("station")

        if trainee_id:
            queryset = queryset.filter(trainee_id=trainee_id)
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if station_name:
            queryset = queryset.filter(station_name__iexact=station_name)  # case-insensitive match

        return queryset



# =================== Refreshment Training ============================= #

from rest_framework import viewsets
from .models import Training_category, Curriculum, CurriculumContent, Trainer_name, Venues, Schedule, MasterTable , RescheduleLog,EmployeeAttendance
from .serializers import Training_categorySerializer, CurriculumSerializer, CurriculumContentSerializer, Trainer_nameSerializer, VenuesSerializer, ScheduleSerializer, MasterTableSerializer, EmployeeAttendanceSerializer, RescheduleLogSerializer

class Training_categoryViewSet(viewsets.ModelViewSet):
    queryset = Training_category.objects.all()
    serializer_class = Training_categorySerializer

class CurriculumViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumSerializer

    def get_queryset(self):
        queryset = Curriculum.objects.all()
        category_id = self.request.query_params.get('category_id')
        if category_id is not None:
            queryset = queryset.filter(category_id=category_id)
        return queryset

class CurriculumContentViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumContentSerializer

    def get_queryset(self):
        queryset = CurriculumContent.objects.all()
        curriculum_id = self.request.query_params.get('curriculum')
        if curriculum_id is not None:
            queryset = queryset.filter(curriculum_id=curriculum_id)
        return queryset

class Trainer_nameViewSet(viewsets.ModelViewSet):
    queryset = Trainer_name.objects.all()
    serializer_class = Trainer_nameSerializer

class VenueViewSet(viewsets.ModelViewSet):
    queryset = Venues.objects.all()
    serializer_class = VenuesSerializer

class ScheduleViewSet(viewsets.ModelViewSet):
    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action

from .models import EmployeeAttendance, RescheduleLog
from .serializers import EmployeeAttendanceSerializer, RescheduleLogSerializer


class EmployeeAttendanceViewSet(viewsets.ModelViewSet):
    queryset = EmployeeAttendance.objects.all()
    serializer_class = EmployeeAttendanceSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        attendance_instance = serializer.save()

        print(f"Attendance created: status={attendance_instance.status}")

        if attendance_instance.status == 'rescheduled':
            if not (attendance_instance.reschedule_date and attendance_instance.reschedule_time and attendance_instance.reschedule_reason):
                print("Missing reschedule details; skipping RescheduleLog creation")
            else:
                try:
                    RescheduleLog.objects.create(
                        schedule=attendance_instance.schedule,
                        employee=attendance_instance.employee,
                        original_date=attendance_instance.schedule.date,
                        original_time=attendance_instance.schedule.time,
                        new_date=attendance_instance.reschedule_date,
                        new_time=attendance_instance.reschedule_time,
                        reason=attendance_instance.reschedule_reason,
                    )
                    print("RescheduleLog created")
                except Exception as e:
                    print("Error creating RescheduleLog:", e)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def by_schedule(self, request):
        schedule_id = request.query_params.get('schedule_id')
        if not schedule_id:
            return Response({"detail": "schedule_id query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        attendances = self.queryset.filter(schedule_id=schedule_id)
        serializer = self.get_serializer(attendances, many=True)
        return Response(serializer.data)




from rest_framework import viewsets
from rest_framework.response import Response

from .models import RescheduleLog
from .serializers import RescheduleLogSerializer


class RescheduleLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Viewset to retrieve Reschedule Logs (Read-only).
    Optional filtering by schedule or employee.
    """
    queryset = RescheduleLog.objects.all()
    serializer_class = RescheduleLogSerializer

    def list(self, request, *args, **kwargs):
        schedule_id = request.query_params.get('schedule_id')
        employee_id = request.query_params.get('employee_id')

        queryset = self.queryset

        if schedule_id:
            queryset = queryset.filter(schedule_id=schedule_id)

        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

# =================== Refreshment Training End ============================= #

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Department, Station, StationSetting
from .serializers import DepartmentselectSerializer, StationselectSerializer, StationSettingSerializer

class DepartmentListView(APIView):
    def get(self, request):
        departments = Department.objects.all()
        serializer = DepartmentselectSerializer(departments, many=True)
        return Response(serializer.data)

class StationsByDepartmentView(APIView):
    def get(self, request, department_id):
        stations = Station.objects.filter(subline_linedepartment_department_id=department_id)
        serializer = StationselectSerializer(stations, many=True)
        return Response(serializer.data)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import StationSetting, Department, Station
from .serializers import StationSettingSerializer

class StationSettingCreateView(APIView):
    def get(self, request):
        department_id = request.query_params.get('department_id')
        if not department_id:
            return Response(
                {"error": "department_id query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
                 
        try:
            department = Department.objects.get(department_id=department_id)
        except Department.DoesNotExist:
            return Response(
                {"error": "Department not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Group settings by station
        settings = StationSetting.objects.filter(department=department).select_related('station', 'department')
        
        # Group by station
        station_groups = {}
        for setting in settings:
            station_key = setting.station.station_id
            if station_key not in station_groups:
                station_groups[station_key] = {
                    'department_id': setting.department.department_id,
                    'department_name': setting.department.department_name,
                    'station_id': setting.station.station_id,
                    'station_name': setting.station.station_name,
                    'all_options': []
                }
            station_groups[station_key]['all_options'].append(setting.option)
        
        # Convert to list and remove duplicates from options
        result = []
        for station_data in station_groups.values():
            station_data['all_options'] = list(set(station_data['all_options']))
            result.append(station_data)
        
        return Response(result, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = StationSettingSerializer(data=request.data)
        if serializer.is_valid():
            created_settings = serializer.save()
            return Response({"success": True, "created": len(created_settings)}, status=201)
        return Response(serializer.errors, status=400)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction

# Import all the necessary models from your models.py file
from .models import TestSession, QuestionPaper, Station, MasterTable, Level


class StartTestSessionView(APIView):
    def post(self, request):
        try:
            print("Incoming Request Data:", request.data)

            # --- 1. Get data from the request ---
            test_name = request.data.get("test_name")
            assignments = request.data.get("assignments", [])
            question_paper_id = request.data.get("question_paper_id")
            level_id = request.data.get("level")
            skill_id = request.data.get("skill")  # Station id

            # --- 2. Basic validation ---
            if not test_name or not assignments:
                response_data = {"error": "Test name and assignments are required."}
                print("Response:", response_data)
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            # --- 3. Fetch related objects ---
            question_paper = None
            if question_paper_id:
                question_paper = get_object_or_404(QuestionPaper, question_paper_id=question_paper_id)

            level_obj = None
            if level_id:
                level_obj = get_object_or_404(Level, level_id=level_id)

            skill_obj = None
            if skill_id:
                skill_obj = get_object_or_404(Station, station_id=skill_id)

            created_sessions = []
            updated_sessions = []

            # --- 4. Use transaction so all succeed or all rollback ---
            with transaction.atomic():
                for item in assignments:
                    key_id = item.get("key_id")
                    employee_id = item.get("employee_id")

                    if not key_id or not employee_id:
                        raise ValueError("key_id and employee_id are required in each assignment.")

                    employee = get_object_or_404(MasterTable, emp_id=employee_id)

                    # ✅ Use update_or_create to avoid UNIQUE constraint error
                    session, created = TestSession.objects.update_or_create(
                        key_id=key_id,
                        defaults={
                            "test_name": test_name,
                            "employee": employee,
                            "level": level_obj,
                            "skill": skill_obj,
                            "question_paper": question_paper,
                        }
                    )

                    if created:
                        created_sessions.append(key_id)
                    else:
                        updated_sessions.append(key_id)

            response_data = {
                "status": "ok",
                "message": "Test sessions processed successfully.",
                "created_sessions": created_sessions,
                "updated_sessions": updated_sessions,
            }
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_201_CREATED)

        except ValueError as e:
            response_data = {"error": str(e)}
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            response_data = {"error": str(e)}
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


import logging
import traceback
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import TestSession, Score

# logger = logging.getLogger(__name__)
# EvaluationPassingCriteria

from decimal import Decimal
import traceback
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


from .models import TestSession, Score, QuestionPaper, MasterTable, TraineeInfo, EvaluationPassingCriteria, Station
from .serializers import ScoreSerializer, SimpleScoreSerializer, TestSessionSerializer
from django.core.cache import cache
from decimal import Decimal
import traceback
import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


import logging
import traceback
from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import TestSession, Score

class ScoreListView(APIView):
    def get(self, request):
        # Assuming you use caching for latest test session
        session_key = cache.get("latest_test_session")
        if not session_key:
            return Response([])

        scores = Score.objects.filter(test__key_id=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data)


class KeyIdToEmployeeNameMap(APIView):
    def get(self, request):
        mapping = TestSession.objects.select_related('employee').all()
        return Response({
            s.key_id: f"{s.employee.first_name} {s.employee.last_name}" 
            for s in mapping
        })


class PastTestSessionsView(APIView):
    def get(self, request):
        # Get distinct test names from Score model through the test relationship
        test_names = Score.objects.select_related('test').filter(
            test__isnull=False
        ).values_list('test__test_name', flat=True).distinct()
        
        return Response(list(test_names))


import traceback
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Score
from .serializers import ScoreSerializer  # Import your serializer

class ScoresByTestView(APIView):
    def get(self, request, name):
        try:
            # Filter scores and pre-fetch related objects for efficiency
            scores = (
                Score.objects
                .filter(test__test_name=name)
                .select_related(
                    'employee', 'employee__department',
                    'department',
                    'skill', 'skill__subline', 'skill__subline__line', 'skill__subline__line__department',
                    'level', 'test', 'test__department'
                )
            )

            # Use the serializer to handle data, including get_department logic
            serializer = ScoreSerializer(scores, many=True)

            # Debug: Print the serialized data (this will trigger the serializer's prints for department sources)
            print("Final data being sent to frontend:", serializer.data)

            return Response(serializer.data)

        except Exception as e:
            print(f"Error in ScoresByTestView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SkillListView(APIView):
    def get(self, request):
        skills = Station.objects.values_list('station_name', flat=True).distinct()
        return Response(list(skills))


class ResultSummaryAPIView(APIView):
    def get(self, request):
        try:
            scores = Score.objects.select_related('employee', 'level', 'skill')
            data = []
            for score in scores:
                # Use the percentage from the score model instead of recalculating
                percentage = score.percentage
                result = 'Pass' if score.percentage >= 80 else 'Retraining' if score.percentage >= 50 else 'Fail'

                data.append({
                    "employee_id": score.employee.emp_id if hasattr(score.employee, 'emp_id') else score.employee.id,
                    "name": f"{score.employee.first_name} {score.employee.last_name}",
                    "marks": score.marks,
                    "percentage": percentage,
                    "section": score.employee.section if hasattr(score.employee, 'section') else '',
                    "level_name": score.level.level_name if score.level and hasattr(score.level, 'level_name') else '',
                    "skill": score.skill.station_name if score.skill and hasattr(score.skill, 'station_name') else (score.skill.skill if score.skill else ''),
                    "result": result,
                })

            serializer = SimpleScoreSerializer(data, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            print(f"Error in ResultSummaryAPIView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class ResultSummaryAPIView(APIView):
    def get(self, request):
        scores = Score.objects.select_related('employee', 'level', 'skill')
        data = []
        for score in scores:
            percentage = round((score.marks / 10) * 100, 2)  # Adjust total marks accordingly
            result = 'Pass' if score.marks >= 8 else 'Retraining' if score.marks >= 5 else 'Fail'

            data.append({
                "employee_id": score.employee.id,
                "name": score.employee.name,
                "marks": score.marks,
                "percentage": percentage,
                "section": score.employee.section,  # assuming CharField
                "level_name": score.level.name if score.level else '',
                "skill": score.skill.skill if score.skill else '',  # Station.skill string
                "result": result,
            })

        serializer = SimpleScoreSerializer(data, many=True)
        return Response(serializer.data)
    


import logging
import traceback
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction
from .models import (
    QuestionPaper, Score, Station, TestSession, MasterTable, 
    Level, TemplateQuestion
)

# Configure logger
logger = logging.getLogger(__name__)

class SubmitWebTestAPIView(APIView):
    """
    Submit test answers from web/tablet exam (without remote).
    """
    def post(self, request):
        try:
            logger.info("SubmitWebTestAPIView called with data: %s", request.data)

            # Extract payload
            employee_id = request.data.get("employee_id")
            test_name = request.data.get("test_name")
            question_paper_id = request.data.get("question_paper_id")
            answers = request.data.get("answers", [])
            skill_id = request.data.get("skill_id")
            level_id = request.data.get("level_id")

            # Validate required fields
            if not employee_id or not test_name or not question_paper_id or not isinstance(answers, list):
                logger.warning("Validation failed: missing required fields.")
                return Response(
                    {"error": "employee_id, test_name, question_paper_id, and answers[] are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get employee
            logger.debug("Fetching employee_id=%s", employee_id)
            employee = get_object_or_404(MasterTable, emp_id=employee_id)

            # Get question paper
            logger.debug("Fetching question_paper_id=%s", question_paper_id)
            question_paper = get_object_or_404(QuestionPaper, question_paper_id=question_paper_id)

            # Handle skill (Station)
            skill = None
            if skill_id:
                try:
                    skill_id_int = int(skill_id)
                    logger.debug("Fetching skill by station_id=%s", skill_id_int)
                    skill = get_object_or_404(Station, station_id=skill_id_int)
                except (ValueError, TypeError):
                    logger.warning("Invalid skill_id: %s. Expected an integer.", skill_id)
                    return Response(
                        {"error": "skill_id must be a valid integer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Handle level
            level_obj = None
            if level_id:
                try:
                    level_id_int = int(level_id)
                    logger.debug("Fetching level_id=%s", level_id_int)
                    level_obj = get_object_or_404(Level, level_id=level_id_int)
                    logger.info("Level fetched: %s", level_obj.level_name)
                except (ValueError, TypeError):
                    logger.warning("Invalid level_id: %s. Expected an integer.", level_id)
                    return Response(
                        {"error": "level_id must be a valid integer."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Get all questions
            questions = list(
                TemplateQuestion.objects
                .filter(question_paper=question_paper)
                .order_by("id")
            )
            total_questions = len(questions)
            logger.info("Total questions found: %s", total_questions)

            if total_questions == 0:
                logger.error("No questions found for paper id=%s", question_paper_id)
                return Response({"error": "No questions found for this paper."}, status=status.HTTP_400_BAD_REQUEST)

            # Compare answers (frontend sends 0-3 or -1 for unanswered; correct_answer is text)
            correct_count = 0
            for i, submitted_ans in enumerate(answers):
                if i < total_questions:
                    question = questions[i]
                    logger.debug("Q%s: submitted=%s", i+1, submitted_ans)

                    # Map submitted answer (0-3 or -1) to letter
                    if isinstance(submitted_ans, int) and 0 <= submitted_ans <= 3:
                        submitted_letter = chr(65 + submitted_ans)  # 0=A, 1=B, etc.
                    else:
                        submitted_letter = None  # Unanswered or invalid
                        logger.debug("Q%s: skipped (unanswered or invalid)", i+1)

                    # Compute correct letter
                    options = [question.option_a, question.option_b, question.option_c, question.option_d]
                    try:
                        correct_idx = options.index(question.correct_answer)
                        correct_letter = chr(65 + correct_idx)  # A=0, etc.
                        if submitted_letter == correct_letter:
                            correct_count += 1
                            logger.debug("Q%s: correct (submitted=%s, correct=%s)", i+1, submitted_letter, correct_letter)
                        else:
                            logger.debug("Q%s: incorrect (submitted=%s, correct=%s)", i+1, submitted_letter, correct_letter)
                    except ValueError:
                        logger.warning("Q%s: correct_answer '%s' not in options %s", i+1, question.correct_answer, options)
                        continue

            percentage = round((correct_count / total_questions) * 100, 2) if total_questions > 0 else 0
            passed = percentage >= 80 # Assuming 80% is the passing mark
            logger.info("Scoring: %s/%s correct (%.2f%%), passed=%s",
                        correct_count, total_questions, percentage, passed)

            # Create or fetch TestSession
            test_session, _ = TestSession.objects.get_or_create(
                test_name=test_name,
                key_id=f"{employee_id}-{question_paper_id}",
                employee=employee,
                defaults={
                    "level": level_obj,
                    "skill": skill,
                    "question_paper": question_paper,
                }
            )

            # Save score linked to TestSession (FIX: Use test field, not test_name)
            with transaction.atomic():
                score, created = Score.objects.get_or_create(
                    employee=employee,
                    test=test_session,  # Use test_session object
                    defaults={
                        'marks': correct_count,
                        'percentage': percentage,
                        'passed': passed,
                        'skill': skill,
                        'level': level_obj,
                    }
                )

                if not created and correct_count > score.marks:
                    logger.info("Updating existing score for employee=%s, test_session=%s", employee_id, test_session.id)
                    score.marks = correct_count
                    score.percentage = percentage
                    score.passed = passed
                    score.skill = skill
                    score.level = level_obj
                    score.save()

            # Construct employee name
            employee_full_name = f"{employee.first_name or ''} {employee.last_name or ''}".strip() or employee.emp_id

            return Response({
                "employee": employee_full_name,
                "marks": correct_count,
                "total_questions": total_questions,
                "percentage": percentage,
                "passed": passed,
                "level_received": level_obj.level_name if level_obj else None,
                "message": "Score saved successfully"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            error_trace = traceback.format_exc()
            logger.error("Error in SubmitWebTestAPIView: %s\n%s", str(e), error_trace)
            return Response({
                "error": str(e),
                "traceback": error_trace
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import *
from .serializers import *
from rest_framework import viewsets


class KeyEventCreateView(APIView):
    def post(self, request):
        serializer = KeyEventSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Key event saved'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LatestKeyEventView(APIView):
    def get(self, request):
        try:
            latest_event = KeyEvent.objects.latest('timestamp')
            serializer = KeyEventSerializer(latest_event)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except KeyEvent.DoesNotExist:
            return Response({"message": "No key events yet."}, status=status.HTTP_404_NOT_FOUND)

        
# api/views.py
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import ConnectEventSerializer

@api_view(['POST'])
def connect_event_create(request):
    serializer = ConnectEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)




@api_view(['POST'])
def vote_event_create(request):
    serializer = VoteEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


from .models import EvaluationPassingCriteria, Level, Department
from .serializers import EvaluationPassingCriteriaSerializer, LevelSerializer, DepartmentSerializer

class EvaluationPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = EvaluationPassingCriteria.objects.select_related('level', 'department').all()
    serializer_class = EvaluationPassingCriteriaSerializer
    
    def create(self, request, *args, **kwargs):
        try:
            # Check if criteria already exists for this level-department combination
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).first()
            
            if existing:
                return Response(
                    {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error creating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if another criteria exists for this level-department combination
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).exclude(id=instance.id).first()
            
            if existing:
                return Response(
                    {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().update(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error updating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(
                {'detail': f'Error deleting criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 
        

# =================== Retraining start ============================= #


from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from collections import defaultdict
from django.shortcuts import get_object_or_404

from .models import (
    RetrainingSession, RetrainingConfig, MasterTable,
   
    OperatorPerformanceEvaluation, TenCyclePassingCriteria,
      
    TraineeInfo, OJTPassingCriteria, OJTScoreRange,
    
    Score, EvaluationPassingCriteria,
    Department, Level, Station
)
from .serializers import RetrainingSessionSerializer, RetrainingConfigSerializer

class RetrainingConfigViewSet(viewsets.ModelViewSet):
    queryset = RetrainingConfig.objects.all()
    serializer_class = RetrainingConfigSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        level_id = self.request.query_params.get('level_id')
        evaluation_type = self.request.query_params.get('evaluation_type')
        
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if evaluation_type:
            queryset = queryset.filter(evaluation_type=evaluation_type)
            
        return queryset
    
    
class RetrainingSessionViewSet(viewsets.ModelViewSet):
    queryset = RetrainingSession.objects.all().order_by("-created_at")
    serializer_class = RetrainingSessionSerializer

    def create(self, request, *args, **kwargs):
     """Schedule a new retraining session"""
     employee_id = request.data.get('employee')
     level_id = request.data.get('level')
     department_id = request.data.get('department')
     station_id = request.data.get('station')
     evaluation_type = request.data.get('evaluation_type')

     if not (employee_id and level_id and department_id and evaluation_type):
        return Response({
            "error": "Missing required fields: employee, level, department, evaluation_type"
        }, status=status.HTTP_400_BAD_REQUEST)

     # Get max allowed retraining sessions for this level and evaluation type
     config = RetrainingConfig.objects.filter(
        level_id=level_id, 
        evaluation_type=evaluation_type
     ).first()
     max_retraining_sessions = config.max_count if config else 2

     # Count existing retraining sessions for this specific combination
     filter_kwargs = {
        'employee_id': employee_id,
        'level_id': level_id,
        'department_id': department_id,
        'evaluation_type': evaluation_type
     }
    
     if station_id:
        filter_kwargs['station_id'] = station_id

     existing_sessions = RetrainingSession.objects.filter(**filter_kwargs)
     existing_count = existing_sessions.count()

     if existing_count >= max_retraining_sessions:
        return Response({
            'error': f'Maximum retraining sessions ({max_retraining_sessions}) reached for this employee and evaluation combination.'
        }, status=status.HTTP_400_BAD_REQUEST)

     request.data['attempt_no'] = existing_count + 2
    
     
    
     # Create the main session with scheduling info only
     response = super().create(request, *args, **kwargs)
     
     
    
     return response

    @action(detail=True, methods=['patch'], url_path='complete-session')
    def complete_session(self, request, pk=None):
        """Complete a retraining session with results and observations"""
        try:
            session = self.get_object()
            
            if session.status != 'Pending':
                return Response(
                    {'error': 'Only pending sessions can be completed'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update main session with completion data
            session.status = request.data.get('status', 'Completed')
            session.performance_percentage = request.data.get('performance_percentage')
            session.required_percentage = request.data.get('required_percentage')
            session.save()
            
            # Update session detail with observations and trainer info
            session_detail, created = RetrainingSessionDetail.objects.get_or_create(
                retraining_session=session
            )
            
            if 'observations_failure_points' in request.data:
                session_detail.observations_failure_points = request.data['observations_failure_points']
            
            if 'trainer_name' in request.data:
                session_detail.trainer_name = request.data['trainer_name']
            
            session_detail.save()
            
            # Return updated session with detail
            updated_serializer = self.get_serializer(session)
            return Response(updated_serializer.data)
            
        except RetrainingSession.DoesNotExist:
            return Response(
                {'error': 'Session not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['patch'], url_path='update-observations')
    def update_observations(self, request, pk=None):
        """Update only observations and trainer info (for partial updates)"""
        try:
            session = self.get_object()
            
            # Get or create session detail
            session_detail, created = RetrainingSessionDetail.objects.get_or_create(
                retraining_session=session
            )
            
            # Update only the fields provided
            if 'observations_failure_points' in request.data:
                session_detail.observations_failure_points = request.data['observations_failure_points']
            
            if 'trainer_name' in request.data:
                session_detail.trainer_name = request.data['trainer_name']
            
            session_detail.save()
            
            # Return updated session with detail
            updated_serializer = self.get_serializer(session)
            return Response(updated_serializer.data)
            
        except RetrainingSession.DoesNotExist:
            return Response(
                {'error': 'Session not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )

    

    @action(detail=False, methods=['get'], url_path='employee-sessions/(?P<employee_id>[^/.]+)')
    def get_employee_sessions(self, request, employee_id=None):
        """Get all retraining sessions for a specific employee, ordered by attempt number"""
        try:
            employee = MasterTable.objects.get(emp_id=employee_id)
            sessions = RetrainingSession.objects.filter(
                employee=employee
            ).select_related('session_detail').order_by(
                'evaluation_type', 'level', 'department', 'station', 'attempt_no'
            )
            
            serializer = self.get_serializer(sessions, many=True)
            return Response({
                'employee_id': employee_id,
                'employee_name': f"{employee.first_name or ''} {employee.last_name or ''}".strip(),
                'sessions': serializer.data
            })
            
        except MasterTable.DoesNotExist:
            return Response(
                {'error': 'Employee not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['get'], url_path='failed-employees')
    def get_failed_employees(self, request):
        """Get all failed employees from all evaluation types with retraining status"""
        failed_employees = []
        
        # Get failed employees from each evaluation system
        ten_cycle_failed = self._get_10cycle_failed_employees()
        failed_employees.extend(ten_cycle_failed)
        
        ojt_failed = self._get_ojt_failed_employees()
        failed_employees.extend(ojt_failed)
        
        evaluation_failed = self._get_evaluation_failed_employees()
        failed_employees.extend(evaluation_failed)
        
        print(f"10 Cycle failures: {len(ten_cycle_failed)}")      
        print(f"OJT failures: {len(ojt_failed)}")                  
        print(f"Evaluation failures: {len(evaluation_failed)}")  
        print(f"Total failures: {len(failed_employees)}") 
        
        # Apply filters
        department_id = request.query_params.get('department_id')
        evaluation_type = request.query_params.get('evaluation_type')
        status_filter = request.query_params.get('status')
        
        if department_id:
            failed_employees = [emp for emp in failed_employees if emp.get('department_id') == int(department_id)]
        
        if evaluation_type:
            failed_employees = [emp for emp in failed_employees if emp.get('evaluation_type') == evaluation_type]
            
        if status_filter:
            failed_employees = [emp for emp in failed_employees if emp.get('retraining_status') == status_filter]
        
        return Response({
            'count': len(failed_employees),
            'results': failed_employees
        })
        
        

    def _get_10cycle_failed_employees(self):
        # """Get failed employees from 10 Cycle evaluations"""
        # failed_employees = []
        
        # evaluations = OperatorPerformanceEvaluation.objects.filter(
        #     final_status='Fail - Retraining Required',
        #     is_completed=True
        # ).select_related('employee', 'department', 'station', 'level')
    
        failed_employees = []
    
        
        all_evaluations = OperatorPerformanceEvaluation.objects.all()
        print(f"Total OperatorPerformanceEvaluations: {all_evaluations.count()}")
    
        
        unique_statuses = OperatorPerformanceEvaluation.objects.values_list('final_status', flat=True).distinct()
        print(f"All final_status values in DB: {list(unique_statuses)}")
    
        
        completed_values = OperatorPerformanceEvaluation.objects.values_list('is_completed', flat=True).distinct()
        print(f"All is_completed values in DB: {list(completed_values)}")
    
        
        fail_evaluations = OperatorPerformanceEvaluation.objects.filter(
         final_status__icontains='Fail'
        )
        print(f"Evaluations containing 'Fail': {fail_evaluations.count()}")
        for eval in fail_evaluations:
         print(f"  - Status: '{eval.final_status}' | Completed: {eval.is_completed} | Employee: {eval.employee}")
    
        
        evaluations = OperatorPerformanceEvaluation.objects.filter(
          final_status='Fail - Retraining Required',
          is_completed=True
        ).select_related('employee', 'department', 'station', 'level')
    
        print(f"Evaluations matching exact criteria: {evaluations.count()}")
        
        for evaluation in evaluations:
            # Get passing criteria
            try:
                criteria = TenCyclePassingCriteria.objects.get(
                    level=evaluation.level,
                    department=evaluation.department,
                    station=evaluation.station,
                    is_active=True
                )
                required_percentage = criteria.passing_percentage
            except TenCyclePassingCriteria.DoesNotExist:
                required_percentage = 60.0
            
            # Get retraining sessions for this employee/evaluation combo
            existing_sessions = RetrainingSession.objects.filter(
                employee=evaluation.employee,
                level=evaluation.level,
                department=evaluation.department,
                station=evaluation.station,
                evaluation_type='10 Cycle'
            ).select_related('session_detail').order_by('-attempt_no')
            # ).order_by('-attempt_no')
            
            # Get max allowed attempts
            config = RetrainingConfig.objects.filter(
                level=evaluation.level, 
                evaluation_type='10 Cycle'
            ).first()
            max_attempts = config.max_count if config else 2
            
            # Determine retraining status
            retraining_status = self._determine_retraining_status(existing_sessions, max_attempts)
            
            gap = required_percentage - (evaluation.final_percentage or 0)
            
            failed_employees.append({
                'employee_pk': evaluation.employee.emp_id,
                'employee_id': evaluation.employee.emp_id,
                'employee_name': f"{evaluation.employee.first_name or ''} {evaluation.employee.last_name or ''}".strip(),
                'department_id': evaluation.department.department_id,
                'department_name': evaluation.department.department_name,
                'station_id': evaluation.station.station_id if evaluation.station else None,
                'station_name': evaluation.station.station_name if evaluation.station else 'N/A',
                'level_id': evaluation.level.level_id,
                'level_name': evaluation.level.level_name,
                'evaluation_type': '10 Cycle',
                'obtained_percentage': round(evaluation.final_percentage or 0, 2),
                'required_percentage': required_percentage,
                'performance_gap': round(gap, 2),
                'last_evaluation_date': evaluation.date.isoformat(),
                'existing_sessions_count': existing_sessions.count(),
                'max_attempts': max_attempts,
                'can_schedule_retraining': existing_sessions.count() < max_attempts,
                'retraining_status': retraining_status,
                # 'retraining_records': [
                #     {
                #         'id': session.id,
                #         'attempt_no': session.attempt_no,
                #         'scheduled_date': session.scheduled_date.isoformat(),
                #         'scheduled_time': session.scheduled_time.strftime('%H:%M'),
                #         'venue': session.venue,
                #         'status': session.status,
                #         'performance_percentage': session.performance_percentage
                #     } for session in existing_sessions
                # ]
                                'retraining_records': [
                       {
                          'id': session.id,
                           'attempt_no': session.attempt_no,
                           'scheduled_date': session.scheduled_date.isoformat(),
                            'scheduled_time': session.scheduled_time.strftime('%H:%M'),
                            'venue': session.venue,
                            'status': session.status,
                            'performance_percentage': session.performance_percentage,
                            'session_detail': {
                                'observations_failure_points': session.session_detail.observations_failure_points if hasattr(session, 'session_detail') else None,
                                'trainer_name': session.session_detail.trainer_name if hasattr(session, 'session_detail') else None,
                            } if hasattr(session, 'session_detail') else None
                        } for session in existing_sessions
                    ]
            })
        
        return failed_employees


    
    
    def _get_ojt_failed_employees(self):
     """Get failed employees from OJT evaluations"""
     failed_employees = []
    
     trainees = TraineeInfo.objects.filter(status='Fail')
    
     for trainee in trainees:
        # Get trainee's score info to determine department/level
        trainee_score = trainee.scores.select_related('topic_department', 'topic_level', 'day').first()
        if not trainee_score:
            continue

        department = trainee_score.topic.department
        level = trainee_score.topic.level
        day = trainee_score.day

        # Get passing criteria
        criteria = (
            OJTPassingCriteria.objects.filter(department=department, level=level, day=day).first()
            or OJTPassingCriteria.objects.filter(department=department, level=level, day__isnull=True).first()
        )
        required_percentage = criteria.percentage if criteria else 60.0

        # Calculate obtained percentage
        from django.db.models import Sum
        total_score = trainee.scores.aggregate(total=Sum("score"))["total"] or 0
        total_topics = trainee.scores.count()

        obtained_percentage = 0
        if total_topics > 0:
            try:
                score_range = OJTScoreRange.objects.filter(
                    department=department,
                    level=level
                ).first()
                max_score = score_range.max_score if score_range else 5
                obtained_percentage = (total_score / (total_topics * max_score)) * 100
            except:
                obtained_percentage = 0

        # Try to match trainee with MasterTable
        try:
            employee = MasterTable.objects.get(emp_id=trainee.emp_id)
            employee_pk = employee.pk
        except MasterTable.DoesNotExist:
            employee = None
            employee_pk = None

        # Get retraining sessions (only if mapped to MasterTable)
        if employee:
            existing_sessions = RetrainingSession.objects.filter(
                employee=employee,
                level=level,
                department=department,
                evaluation_type='OJT'
            ).select_related('session_detail').order_by('-attempt_no')
            # ).order_by('-attempt_no')

            config = RetrainingConfig.objects.filter(level=level, evaluation_type='OJT').first()
            max_attempts = config.max_count if config else 2

            retraining_status = self._determine_retraining_status(existing_sessions, max_attempts)
        else:
            existing_sessions = []
            max_attempts = 0
            retraining_status = "not_mapped"  # Or "N/A"

        gap = required_percentage - obtained_percentage

        failed_employees.append({
            'employee_pk': employee_pk, 
            'employee_id': trainee.emp_id,
            'employee_name': trainee.trainee_name,
            'department_id': department.department_id,
            'department_name': department.department_name,
            'station_id': None,
            'station_name': trainee.station,
            'level_id': level.level_id,
            'level_name': level.level_name,
            'evaluation_type': 'OJT',
            'obtained_percentage': round(obtained_percentage, 2),
            'required_percentage': required_percentage,
            'performance_gap': round(gap, 2),
            'last_evaluation_date': trainee.doj.isoformat() if trainee.doj else None,
            'existing_sessions_count': len(existing_sessions),
            'max_attempts': max_attempts,
            'can_schedule_retraining': employee is not None and len(existing_sessions) < max_attempts,
            'retraining_status': retraining_status,
            # 'retraining_records': [
            #     {
            #         'id': session.id,
            #         'attempt_no': session.attempt_no,
            #         'scheduled_date': session.scheduled_date.isoformat(),
            #         'scheduled_time': session.scheduled_time.strftime('%H:%M'),
            #         'venue': session.venue,
            #         'status': session.status,
            #         'performance_percentage': session.performance_percentage
            #     } for session in existing_sessions
            # ]
                            'retraining_records': [
                       {
                          'id': session.id,
                           'attempt_no': session.attempt_no,
                           'scheduled_date': session.scheduled_date.isoformat(),
                            'scheduled_time': session.scheduled_time.strftime('%H:%M'),
                            'venue': session.venue,
                            'status': session.status,
                            'performance_percentage': session.performance_percentage,
                            'session_detail': {
                                'observations_failure_points': session.session_detail.observations_failure_points if hasattr(session, 'session_detail') else None,
                                'trainer_name': session.session_detail.trainer_name if hasattr(session, 'session_detail') else None,
                            } if hasattr(session, 'session_detail') else None
                        } for session in existing_sessions
                    ]
        })
 
     return failed_employees


    def _get_evaluation_failed_employees(self):
        """Get failed employees from Evaluation tests"""
        failed_employees = []
        
        failed_scores = Score.objects.filter(passed=False).select_related(
            'employee', 'level', 'skill', 'test'
        )
        
        for score in failed_scores:
            if not score.employee or not score.level:
                continue
                
            # Get department
            department = None
            if hasattr(score.employee, 'department') and score.employee.department:
                department = score.employee.department
            elif score.skill and hasattr(score.skill, 'subline'):
                try:
                    department = score.skill.subline.line.department
                except AttributeError:
                    continue
            else:
                continue
            
            # Get passing criteria
            criteria = EvaluationPassingCriteria.objects.filter(
                level=score.level,
                department=department
            ).first()
            required_percentage = float(criteria.percentage) if criteria else 80.0
            
            # Get retraining sessions
            existing_sessions = RetrainingSession.objects.filter(
                employee=score.employee,
                level=score.level,
                department=department,
                evaluation_type='Evaluation'
            ).select_related('session_detail').order_by('-attempt_no')
            # ).order_by('-attempt_no')
            
            config = RetrainingConfig.objects.filter(level=score.level, evaluation_type='Evaluation').first()
            max_attempts = config.max_count if config else 2
            
            retraining_status = self._determine_retraining_status(existing_sessions, max_attempts)
            gap = required_percentage - score.percentage
            
            failed_employees.append({
                'employee_pk': score.employee.emp_id,
                'employee_id': score.employee.emp_id,
                'employee_name': f"{score.employee.first_name or ''} {score.employee.last_name or ''}".strip(),
                'department_id': department.department_id,
                'department_name': department.department_name,
                'station_id': score.skill.station_id if score.skill else None,
                'station_name': score.skill.station_name if score.skill else 'N/A',
                'level_id': score.level.level_id,
                'level_name': score.level.level_name,
                'evaluation_type': 'Evaluation',
                'obtained_percentage': round(score.percentage, 2),
                'required_percentage': required_percentage,
                'performance_gap': round(gap, 2),
                'last_evaluation_date': score.created_at.date().isoformat(),
                'existing_sessions_count': existing_sessions.count(),
                'max_attempts': max_attempts,
                'can_schedule_retraining': existing_sessions.count() < max_attempts,
                'retraining_status': retraining_status,
                # 'retraining_records': [
                #     {
                #         'id': session.id,
                #         'attempt_no': session.attempt_no,
                #         'scheduled_date': session.scheduled_date.isoformat(),
                #         'scheduled_time': session.scheduled_time.strftime('%H:%M'),
                #         'venue': session.venue,
                #         'status': session.status,
                #         'performance_percentage': session.performance_percentage
                #     } for session in existing_sessions
                # ]
                'retraining_records': [
                       {
                          'id': session.id,
                           'attempt_no': session.attempt_no,
                           'scheduled_date': session.scheduled_date.isoformat(),
                            'scheduled_time': session.scheduled_time.strftime('%H:%M'),
                            'venue': session.venue,
                            'status': session.status,
                            'performance_percentage': session.performance_percentage,
                            'session_detail': {
                                'observations_failure_points': session.session_detail.observations_failure_points if hasattr(session, 'session_detail') else None,
                                'trainer_name': session.session_detail.trainer_name if hasattr(session, 'session_detail') else None,
                            } if hasattr(session, 'session_detail') else None
                        } for session in existing_sessions
                    ]
            })
        
        return failed_employees

    def _determine_retraining_status(self, existing_sessions, max_attempts):
        """Determine the retraining status based on existing sessions"""
        if not existing_sessions.exists():
            return 'pending'
        
        latest_session = existing_sessions.first()
        
        if existing_sessions.count() >= max_attempts:
            # Check if last attempt was successful
            if latest_session.status == 'Completed' and latest_session.performance_percentage and latest_session.performance_percentage >= (latest_session.required_percentage or 60):
                return 'completed'
            else:
                return 'failed'  # Max attempts reached without success
        
        if latest_session.status == 'Pending':
            return 'scheduled'
        elif latest_session.status == 'Completed':
            if latest_session.performance_percentage and latest_session.performance_percentage >= (latest_session.required_percentage or 60):
                return 'completed'
            else:
                return 'pending'  # Can schedule another attempt
        elif latest_session.status == 'Missed':
            return 'pending'  # Can reschedule
        
        return 'pending'

    @action(detail=False, methods=['get'], url_path='summary')
    def get_summary(self, request):
        """Get summary statistics for retraining dashboard"""
        # Get all failed employees
        failed_response = self.get_failed_employees(request)
        all_failed = failed_response.data['results']
        
        # Calculate summary stats
        total_failed = len(all_failed)
        pending = len([emp for emp in all_failed if emp['retraining_status'] == 'pending'])
        scheduled = len([emp for emp in all_failed if emp['retraining_status'] == 'scheduled'])
        completed = len([emp for emp in all_failed if emp['retraining_status'] == 'completed'])
        failed = len([emp for emp in all_failed if emp['retraining_status'] == 'failed'])
        
        # Group by evaluation type
        by_evaluation_type = {}
        for emp in all_failed:
            eval_type = emp['evaluation_type']
            if eval_type not in by_evaluation_type:
                by_evaluation_type[eval_type] = {
                    'total': 0,
                    'pending': 0,
                    'scheduled': 0,
                    'completed': 0,
                    'failed': 0
                }
            by_evaluation_type[eval_type]['total'] += 1
            by_evaluation_type[eval_type][emp['retraining_status']] += 1
        
        # Group by department
        by_department = {}
        for emp in all_failed:
            dept_name = emp['department_name']
            if dept_name not in by_department:
                by_department[dept_name] = {
                    'total': 0,
                    'pending': 0,
                    'scheduled': 0,
                    'completed': 0,
                    'failed': 0
                }
            by_department[dept_name]['total'] += 1
            by_department[dept_name][emp['retraining_status']] += 1
        
        return Response({
            'overall_summary': {
                'total_failed_employees': total_failed,
                'pending_retraining': pending,
                'scheduled_retraining': scheduled,
                'completed_retraining': completed,
                'failed_retraining': failed
            },
            'by_evaluation_type': by_evaluation_type,
            'by_department': by_department
        })
    
# =================== Retraining end ============================= #

from rest_framework import viewsets
from .models import (
    QuantityScoreSetup,
    QuantityPassingCriteria,
    OJTLevel2Quantity,
    Level2QuantityOJTEvaluation,
)
from .serializers import (
    QuantityScoreSetupSerializer,
    QuantityPassingCriteriaSerializer,
    OJTLevel2QuantitySerializer,
    Level2QuantityOJTEvaluationSerializer,
)


# ----------------------------
# Quantity Score Ranges
# ----------------------------
class QuantityOJTScoreRangeViewSet(viewsets.ModelViewSet):
    queryset = QuantityScoreSetup.objects.all()
    serializer_class = QuantityScoreSetupSerializer


# ----------------------------
# Quantity Passing Criteria
# ----------------------------
class QuantityPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = QuantityPassingCriteria.objects.all()
    serializer_class = QuantityPassingCriteriaSerializer


# ----------------------------
# OJT Main Record (Trainee)
# ----------------------------

# ----------------------------
# Daily Evaluation
# ----------------------------
class Level2QuantityOJTEvaluationViewSet(viewsets.ModelViewSet):
    queryset = Level2QuantityOJTEvaluation.objects.all()
    serializer_class = Level2QuantityOJTEvaluationSerializer


from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import AssessmentMode
from .serializers import AssessmentModeSerializer

@api_view(['GET'])
def get_assessment_mode(request):
    """Get current assessment mode"""
    current_mode = AssessmentMode.get_current_mode()
    return Response({
        'mode': current_mode.mode,
        'updated_at': current_mode.updated_at
    })

@api_view(['POST'])
def toggle_assessment_mode(request):
    """Toggle between quality and quantity mode"""
    mode = request.data.get('mode')
    
    if mode not in ['quality', 'quantity']:
        return Response({
            'error': 'Invalid mode. Must be "quality" or "quantity"'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    current_mode = AssessmentMode.get_current_mode()
    current_mode.mode = mode
    current_mode.save()
    
    return Response({
        'mode': current_mode.mode,
        'updated_at': current_mode.updated_at,
        'message': f'Mode switched to {mode}'
    })


from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import LevelColour, Level
from .serializers import LevelColourSerializer

class LevelColourViewSet(viewsets.ModelViewSet):
    queryset = LevelColour.objects.all()
    serializer_class = LevelColourSerializer
    
    @action(detail=False, methods=['post'])
    def reset_to_defaults(self, request):
        """Reset all level colors to default values"""
        try:
            for level_id, color_code in DEFAULT_COLOURS.items():
                try:
                    level = Level.objects.get(level_id=level_id)
                    level_colour, created = LevelColour.objects.get_or_create(level=level)
                    level_colour.colour_code = color_code
                    level_colour.save()
                except Level.DoesNotExist:
                    continue
            
            # Return updated colors
            colours = LevelColour.objects.all()
            serializer = self.get_serializer(colours, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Update multiple level colors at once"""
        try:
            colors_data = request.data.get('colors', {})
            updated_colors = []
            
            for level_key, color_code in colors_data.items():
                # Extract level number from key like 'level1' -> 1
                level_num = int(level_key.replace('level', ''))
                
                try:
                    level = Level.objects.get(level_id=level_num)
                    level_colour, created = LevelColour.objects.get_or_create(level=level)
                    level_colour.colour_code = color_code
                    level_colour.save()
                    updated_colors.append(level_colour)
                except Level.DoesNotExist:
                    continue
            
            serializer = self.get_serializer(updated_colors, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import SkillMatrixDisplaySetting
from .serializers import SkillMatrixDisplaySettingSerializer

class SkillMatrixDisplaySettingViewSet(viewsets.ViewSet):
    """
    A simple ViewSet to get/set display shape.
    """

    def list(self, request):
        # Retrieve or create singleton instance
        setting, created = SkillMatrixDisplaySetting.objects.get_or_create(id=1)
        serializer = SkillMatrixDisplaySettingSerializer(setting)
        return Response(serializer.data)

    def update(self, request, pk=None):
        setting, created = SkillMatrixDisplaySetting.objects.get_or_create(id=1)
        serializer = SkillMatrixDisplaySettingSerializer(setting, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class DepartmentSubLineViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = SubLineSerializer
    
    def get_queryset(self):
        department_id = self.request.query_params.get('department_id')
        if department_id:
            return SubLine.objects.filter(line_department_department_id=department_id)
        return SubLine.objects.none()
    
    def list(self, request, *args, **kwargs):
        try:
            department_id = request.query_params.get('department_id')
            
            if not department_id:
                response = Response(
                    {"error": "department_id parameter is required"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            # Verify department exists
            try:
                department = Department.objects.get(department_id=department_id)
            except Department.DoesNotExist:
                response = Response(
                    {"error": "Department not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            
            response_data = {
                'department_id': department.department_id,
                'department_name': department.department_name,
                'sublines': serializer.data,
                'count': queryset.count()
            }
            
            response = Response(response_data, status=status.HTTP_200_OK)
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            
            return response
            
        except Exception as e:
            response = Response(
                {"error": "Failed to fetch sublines", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            return response


class DepartmentStationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StationSerializer
    
    def get_queryset(self):
        department_id = self.request.query_params.get('department_id')
        if department_id:
            return Station.objects.filter(subline_linedepartment_department_id=department_id)
        return Station.objects.none()
    
    def list(self, request, *args, **kwargs):
        try:
            department_id = request.query_params.get('department_id')
            
            if not department_id:
                response = Response(
                    {"error": "department_id parameter is required"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            # Verify department exists
            try:
                department = Department.objects.get(department_id=department_id)
            except Department.DoesNotExist:
                response = Response(
                    {"error": "Department not found"}, 
                    status=status.HTTP_404_NOT_FOUND
                )
                response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
                return response
            
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            
            response_data = {
                'department_id': department.department_id,
                'department_name': department.department_name,
                'stations': serializer.data,
                'count': queryset.count()
            }
            
            response = Response(response_data, status=status.HTTP_200_OK)
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            response['Access-Control-Allow-Credentials'] = 'true'
            
            return response
            
        except Exception as e:
            response = Response(
                {"error": "Failed to fetch stations", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            response['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
            return response


from rest_framework import viewsets
from .models import CompanyLogo
from .serializers import CompanyLogoSerializer

class CompanyLogoViewSet(viewsets.ModelViewSet):
    queryset = CompanyLogo.objects.all()
    serializer_class = CompanyLogoSerializer


@api_view(['GET'])
def get_lines_by_department(request, department_id):
    department = get_object_or_404(Department, department_id=department_id)
    lines = Line.objects.filter(
        department=department_id
    ).select_related('department')
    
    if not lines.exists():
        return Response({
            'message': f'No lines found under department: {department.department_name}',
            'department_name': department.department_name,
            'lines': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = LineReadSerializer(lines, many=True)
    
    return Response({
        'lines': serializer.data,
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def get_sublines_by_line(request, line_id):
    # Check if line exists
    line = get_object_or_404(Line, line_id=line_id)
    
    # Get all sublines under this line
    sublines = SubLine.objects.filter(
        line=line_id
    ).select_related('line', 'line__department')
    
    if not sublines.exists():
        return Response({
            'message': f'No sublines found under line: {line.line_name}',
            'line_name': line.line_name,
            'department_name': line.department.department_name,
            'sublines': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = SubLineReadSerializer(sublines, many=True)
    
    return Response({
        'line_id': line_id,
        'line_name': line.line_name,
        'sublines': serializer.data,
    }, status=status.HTTP_200_OK)

@api_view(['GET'])
def get_stations_by_subline(request, subline_id):
    
    # Check if subline exists
    subline = get_object_or_404(SubLine, subline_id=subline_id)
    
    # Get all stations under this subline
    stations = Station.objects.filter(
        subline=subline_id
    ).select_related('subline', 'subline__line', 'subline__line__department')
    
    if not stations.exists():
        return Response({
            'message': f'No stations found under subline: {subline.subline_name}',
            'subline_name': subline.subline_name,
            'line_name': subline.line.line_name,
            'department_name': subline.line.department.department_name,
            'stations': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = StationReadSerializer(stations, many=True)
    
    return Response({
        'subline_id': subline_id,
        'subline_name': subline.subline_name,
        'stations': serializer.data,
    }, status=status.HTTP_200_OK)



@api_view(['GET'])
def get_stations_by_department(request, department_id):
    department = get_object_or_404(Department, department_id=department_id)
    
    # Get all stations under this department
    stations = Station.objects.filter(
        subline__line__department_id=department_id
    ).select_related('subline', 'subline__line', 'subline__line__department')
    
    if not stations.exists():
        return Response({
            'message': f'No stations found under department: {department.department_name}',
            'department_name': department.department_name,
            'stations': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = StationReadSerializer(stations, many=True)
    
    return Response({
        'department_id': department_id,
        'department_name': department.department_name,
        'stations': serializer.data,
    }, status=status.HTTP_200_OK)



@api_view(['GET'])
def get_stations_by_line(request, line_id):
    
    line = get_object_or_404(Line, line_id=line_id)
    
    stations = Station.objects.filter(
        subline__line=line_id
    ).select_related('subline', 'subline__line', 'subline__line__department')
    
    if not stations.exists():
        return Response({
            'message': f'No stations found under line: {line.line_name}',
            'line_name': line.line_name,
            'department_name': line.department.department_name,
            'stations': [],
            'total_count': 0
        }, status=status.HTTP_200_OK)
    
    serializer = StationReadSerializer(stations, many=True)
    
    return Response({
        'line_id': line_id,
        'line_name': line.line_name,
        'stations': serializer.data,
    }, status=status.HTTP_200_OK)


from rest_framework.decorators import api_view

@api_view(['GET'])
def get_all_departments(request):
    departments = Department.objects.all()
    serializer = DepartmentReadSerializer(departments, many=True)
    return Response({
        'departments': serializer.data,
    }, status=status.HTTP_200_OK)





# ----------------------------------notification--------------------------#
from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, BasePermission
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta
from .models import Notification
from .serializers import (
    NotificationSerializer, NotificationCreateSerializer,
    NotificationUpdateSerializer, NotificationStatsSerializer
)


class NotificationPermission(BasePermission):
    """
    Custom permission for notifications:
    - Allow read access without authentication (for testing)
    - Require authentication for write operations
    """
    def has_permission(self, request, view):
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return True
        return request.user and request.user.is_authenticated


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing notifications - no authentication required for read
    """
    serializer_class = NotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Notification.objects.all().select_related(
            'recipient', 'employee', 'level', 'training_schedule',
            'machine_allocation', 'test_session', 'retraining_session',
            'human_body_check_session'
        )

    def get_serializer_class(self):
        if self.action == 'create':
            return NotificationCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return NotificationUpdateSerializer
        return NotificationSerializer
    

    def list(self, request, *args, **kwargs):
        logger.info(f"Received request for notifications with params: {request.query_params}")
        queryset = self.get_queryset()

        is_read = request.query_params.get('is_read')
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read.lower() == 'true')

        notification_types = request.query_params.getlist('notification_type')
        if notification_types:
            queryset = queryset.filter(notification_type__in=notification_types)

        notification_type = request.query_params.get('type')
        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)

        priority = request.query_params.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)

        days = request.query_params.get('days')
        if days:
            try:
                days_int = int(days)
                since_date = timezone.now() - timedelta(days=days_int)
                queryset = queryset.filter(created_at__gte=since_date)
            except ValueError:
                logger.warning(f"Invalid days parameter: {days}")

        queryset = queryset.order_by('-created_at')
        logger.info(f"Returning {queryset.count()} notifications")

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    
    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_read()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_unread(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_unread()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        count = queryset.count()
        for notification in queryset:
            notification.mark_as_read()
        return Response({'message': f'Marked {count} notifications as read', 'count': count})

    @action(detail=False, methods=['get'])
    def unread(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        queryset = self.get_queryset()
        total_count = queryset.count()
        unread_count = queryset.filter(is_read=False).count()
        read_count = total_count - unread_count
        recent_count = queryset.filter(created_at__gte=timezone.now() - timedelta(hours=24)).count()
        by_type = dict(queryset.values('notification_type').annotate(count=Count('id')).values_list('notification_type', 'count'))
        by_priority = dict(queryset.values('priority').annotate(count=Count('id')).values_list('priority', 'count'))

        stats_data = {
            'total_count': total_count,
            'unread_count': unread_count,
            'read_count': read_count,
            'recent_count': recent_count,
            'by_type': by_type,
            'by_priority': by_priority
        }
        serializer = NotificationStatsSerializer(stats_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def recent(self, request):
        since_date = timezone.now() - timedelta(hours=24)
        queryset = self.get_queryset().filter(created_at__gte=since_date)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ==================== API FUNCTION VIEWS ====================

@api_view(['GET'])
def notification_count(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

    count = Notification.objects.filter(
        Q(recipient=request.user) | Q(recipient_email=request.user.email),
        is_read=False
    ).count()
    return Response({'unread_count': count})


@api_view(['GET'])
def test_notifications(request):
    notifications = Notification.objects.all()[:10]
    serializer = NotificationSerializer(notifications, many=True)
    return Response({'count': notifications.count(), 'notifications': serializer.data, 'debug': 'This is a test endpoint'})


@api_view(['POST'])
def create_system_notification(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_403_FORBIDDEN)

    serializer = NotificationCreateSerializer(data=request.data)
    if serializer.is_valid():
        notification = serializer.save()
        response_serializer = NotificationSerializer(notification)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def create_test_notification(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        notification = Notification.objects.create(
            title="Test Notification",
            message="This is a test notification to verify the system is working.",
            notification_type="system_alert",
            recipient=recipient,
            priority="medium",
            metadata={"test": True}
        )
        serializer = NotificationSerializer(notification)
        return Response({'message': 'Test notification created successfully', 'notification': serializer.data}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_employee_notification(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        latest_employee = MasterTable.objects.last()

        if not latest_employee:
            return Response({'error': 'No employees found in the system'}, status=status.HTTP_400_BAD_REQUEST)

        notification = Notification.objects.create(
            title="New Employee Registered",
            message=f"New employee {latest_employee.first_name} {latest_employee.last_name} (Employee ID: {latest_employee.emp_id}) has been registered.",
            notification_type='employee_registration',
            recipient=recipient,
            employee=latest_employee,
            priority='medium',
            metadata={
                'emp_id': latest_employee.emp_id,
                'department': latest_employee.department.department_name if latest_employee.department else None
            }
        )
        serializer = NotificationSerializer(notification)
        return Response({'message': 'Employee notification created', 'notification': serializer.data}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_all_notification_types(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        latest_employee = MasterTable.objects.last()
        employee_name = f"{latest_employee.first_name} {latest_employee.last_name}" if latest_employee else "Test Employee"

        notification_types = [
            {'type': 'employee_registration', 'title': 'New Employee Registered', 'message': f'{employee_name} has been registered.', 'priority': 'medium'},
            {'type': 'level_exam_completed', 'title': 'Level Exam Completed', 'message': f'{employee_name} completed Level 2 evaluation.', 'priority': 'high'},
            {'type': 'training_scheduled', 'title': 'Training Scheduled', 'message': f'Training scheduled for {employee_name}.', 'priority': 'medium'},
            {'type': 'training_completed', 'title': 'Training Completed', 'message': f'{employee_name} completed training.', 'priority': 'medium'},
            {'type': 'training_reschedule', 'title': 'Training Rescheduled', 'message': f'Training for {employee_name} rescheduled.', 'priority': 'medium'},
            {'type': 'refresher_training_scheduled', 'title': 'Refresher Training Scheduled', 'message': f'Refresher training scheduled for {employee_name}.', 'priority': 'medium'},
            {'type': 'refresher_training_completed', 'title': 'Refresher Training Completed', 'message': f'{employee_name} completed refresher training.', 'priority': 'medium'},
            {'type': 'hanchou_exam_completed', 'title': 'Hanchou Exam Completed', 'message': f'{employee_name} completed Hanchou exam.', 'priority': 'high'},
            {'type': 'shokuchou_exam_completed', 'title': 'Shokuchou Exam Completed', 'message': f'{employee_name} completed Shokuchou exam.', 'priority': 'high'},
            {'type': 'ten_cycle_evaluation_completed', 'title': '10 Cycle Evaluation Completed', 'message': f'{employee_name} completed 10 Cycle evaluation.', 'priority': 'high'},
            {'type': 'ojt_completed', 'title': 'OJT Completed', 'message': f'{employee_name} completed OJT.', 'priority': 'medium'},
            {'type': 'ojt_quantity_completed', 'title': 'OJT Quantity Completed', 'message': f'{employee_name} completed OJT Quantity evaluation.', 'priority': 'medium'},
            {'type': 'machine_allocated', 'title': 'Machine Allocated', 'message': f'Machine allocated to {employee_name}.', 'priority': 'medium'},
            {'type': 'test_assigned', 'title': 'Test Assigned', 'message': f'Test assigned to {employee_name}.', 'priority': 'medium'},
            {'type': 'evaluation_completed', 'title': 'Evaluation Completed', 'message': f'{employee_name} completed an evaluation.', 'priority': 'high'},
            {'type': 'retraining_scheduled', 'title': 'Retraining Scheduled', 'message': f'Retraining scheduled for {employee_name}.', 'priority': 'medium'},
            {'type': 'retraining_completed', 'title': 'Retraining Completed', 'message': f'{employee_name} completed retraining.', 'priority': 'medium'},
            {'type': 'human_body_check_completed', 'title': 'Human Body Check Completed', 'message': f'{employee_name} completed human body check.', 'priority': 'medium'},
            {'type': 'milestone_reached', 'title': 'Milestone Reached', 'message': f'{employee_name} reached a milestone.', 'priority': 'high'},
            {'type': 'system_alert', 'title': 'System Alert', 'message': 'System maintenance at 2:00 AM.', 'priority': 'urgent'}
        ]

        created_notifications = []
        for notif_data in notification_types:
            n = Notification.objects.create(
                title=notif_data['title'],
                message=notif_data['message'],
                notification_type=notif_data['type'],
                recipient=recipient,
                employee=latest_employee if latest_employee else None,
                priority=notif_data['priority'],
                metadata={'test': True, 'employee': employee_name}
            )
            created_notifications.append(n)

        return Response({'message': f'Created {len(created_notifications)} notifications', 'count': len(created_notifications)}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
def delete_all_notifications(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

    try:
        deleted_count = Notification.objects.filter(
            Q(recipient=request.user) | Q(recipient_email=request.user.email)
        ).delete()[0]
        return Response({'message': f'Deleted {deleted_count} notifications', 'count': deleted_count}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class LevelOnePassedUsersView(generics.ListAPIView):
    """
    API view to retrieve a list of all scores for users who have 
    successfully passed a 'Level 1' assessment.

    The results are ordered by the most recent test date first.
    """
    serializer_class = LevelOnePassedScoreSerializer
    
    def get_queryset(self):
        """
        This view returns a list of all scores for
        users who have passed a test specifically at 'Level 1'.
        """
        queryset = Score.objects.filter(
            passed=True,
            level__level_name='Level 1'  # <-- CORRECTED LOOKUP HERE
        ).select_related(
            'employee', 'test', 'level', 'skill'
        ).order_by('-created_at')
        
        return queryset
    

# In views.py
from .models import HandoverSheet
from .serializers import HandoverSheetCreateSerializer


class HandoverSheetViewSet(viewsets.ModelViewSet):
    queryset = HandoverSheet.objects.all()
    serializer_class = HandoverSheetCreateSerializer

from django.shortcuts import get_object_or_404

class EmployeeHandoverView(generics.RetrieveAPIView):
    serializer_class = HandoverSheetCreateSerializer
    queryset = HandoverSheet.objects.all()

    lookup_field = "employee__emp_id"  # tell DRF to look up by emp_id
    lookup_url_kwarg = "emp_id"











from datetime import datetime, timedelta 
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction
from datetime import datetime
from django.db.models import Q
import calendar
from django.db.models import Sum, Avg, F
from django.db.models.functions import Coalesce

# Make sure to import your model and serializer
from .models import DailyProductionData
from .serializers import DailyProductionDataSerializer

class DailyProductionDataViewSet(viewsets.ModelViewSet):
    """
    Final, correct ViewSet for handling production data.
    All calculation logic is now handled here, not in signals.
    """
    queryset = DailyProductionData.objects.all()
    serializer_class = DailyProductionDataSerializer

    def get_queryset(self):
        """
        Overrides the default queryset to apply filters from URL parameters.
        """
        # Start with all objects
        queryset = super().get_queryset()

        # Get parameters from the URL
        factory_id = self.request.query_params.get('factory')
        department_id = self.request.query_params.get('department')

        # Apply filters if the parameters exist
        if factory_id:
            queryset = queryset.filter(factory_id=factory_id)

        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        # Return the final, correctly filtered data
        return queryset

    @action(detail=False, methods=['post'], url_path='save-plan-entry')
    @transaction.atomic
    def save_plan_entry(self, request):
        """
        Handles creating or updating a plan entry.
        It now performs all calculations itself, ignoring the pre_save signal.
        """
        data = request.data
        entry_id = data.get('id', None)

        try:
            # Step 1: Calculate all the totals into a separate dictionary
            # This is safer and doesn't modify the original request data.
            calculated_totals = {}
            calculated_totals['ctq_plan_total'] = data.get('ctq_plan_l1', 0) + data.get('ctq_plan_l2', 0) + data.get('ctq_plan_l3', 0) + data.get('ctq_plan_l4', 0)
            calculated_totals['ctq_actual_total'] = data.get('ctq_actual_l1', 0) + data.get('ctq_actual_l2', 0) + data.get('ctq_actual_l3', 0) + data.get('ctq_actual_l4', 0)
            calculated_totals['pdi_plan_total'] = data.get('pdi_plan_l1', 0) + data.get('pdi_plan_l2', 0) + data.get('pdi_plan_l3', 0) + data.get('pdi_plan_l4', 0)
            calculated_totals['pdi_actual_total'] = data.get('pdi_actual_l1', 0) + data.get('pdi_actual_l2', 0) + data.get('pdi_actual_l3', 0) + data.get('pdi_actual_l4', 0)
            calculated_totals['other_plan_total'] = data.get('other_plan_l1', 0) + data.get('other_plan_l2', 0) + data.get('other_plan_l3', 0) + data.get('other_plan_l4', 0)
            calculated_totals['other_actual_total'] = data.get('other_actual_l1', 0) + data.get('other_actual_l2', 0) + data.get('other_actual_l3', 0) + data.get('other_actual_l4', 0)
            calculated_totals['grand_total_plan'] = calculated_totals['ctq_plan_total'] + calculated_totals['pdi_plan_total'] + calculated_totals['other_plan_total']
            calculated_totals['grand_total_actual'] = calculated_totals['ctq_actual_total'] + calculated_totals['pdi_actual_total'] + calculated_totals['other_actual_total']

            # Step 2: Perform the overlap check
            start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
            end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
            entry_mode = data['entry_mode']
            
            overlap_condition = Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
            # filters = {'line_id': data.get('line'), 'station_id': data.get('station'), 'shop_floor_id': data.get('shop_floor'), 'factory_id': data.get('factory')}
            filters = {
                'Hq_id': data.get('Hq'),
                'factory_id': data.get('factory'),
                'department_id': data.get('department'),
                'line_id': data.get('line'),
                'subline_id': data.get('subline'),
                'station_id': data.get('station'),
            }
            filters = {k: v for k, v in filters.items() if v}
            overlapping_entries = DailyProductionData.objects.filter(overlap_condition, **filters)

            if entry_id:
                overlapping_entries = overlapping_entries.exclude(id=entry_id)

            if overlapping_entries.exists():
                existing_entry = overlapping_entries.first()
                if entry_mode == 'MONTHLY' and existing_entry.entry_mode != 'MONTHLY':
                    return Response({'error': "Cannot save: Weekly/Daily data already exists in this period."}, status=status.HTTP_409_CONFLICT)
                if entry_mode != 'MONTHLY' and existing_entry.entry_mode == 'MONTHLY':
                    return Response({'error': f"Cannot save: A Monthly entry already exists for this period."}, status=status.HTTP_409_CONFLICT)

            # Step 3: Serialize and Save
            if entry_id:
                instance = DailyProductionData.objects.get(id=entry_id)
                serializer = self.get_serializer(instance, data=data, partial=True)
            else:
                serializer = self.get_serializer(data=data)
            
            serializer.is_valid(raise_exception=True)
            # Pass the calculated totals as extra arguments to the save method
            saved_instance = serializer.save(**calculated_totals)
            
            # Create a new serializer from the final saved object to send back
            response_serializer = self.get_serializer(saved_instance)
            return Response(response_serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        


        # In views.py, add this method inside the DailyProductionDataViewSet class

    

    # --- ADD THIS ENTIRE NEW FUNCTION FOR THE WEEKLY VIEW ---
    @action(detail=False, methods=['get'], url_path='weekly-summary')
    def weekly_summary(self, request):
        """
        NOTE: This function aggregates all data for an entire MONTH, not a single week.
        It fetches level-based data for a given factory, year, and month.
        """
        # --- 1. Get Parameters ---
        factory_id = request.query_params.get('factory')
        month_str = request.query_params.get('month')
        year_str = request.query_params.get('year')

        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        # --- ADD THESE LINES ---
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')

        if not all([factory_id, month_str, year_str]):
            return Response(
                {'error': 'factory, month, and year are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # --- 2. Build Filters ---
        try:
            filters = {
                'factory_id': factory_id,
                'start_date__year': int(year_str),
                'start_date__month': int(month_str),
            }
            # --- EDIT THIS SECTION ---
            # Apply optional filters
            if Hq_id: filters['Hq_id'] = Hq_id
            if department_id: filters['department_id'] = department_id
            if subline_id: filters['subline_id'] = subline_id
            if station_id:
                filters['station_id'] = station_id
            elif line_id:
                filters['line_id'] = line_id
                
        except (ValueError, TypeError):
            return Response({'error': 'Invalid year or month format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Define Aggregation Fields (Corrected to be a DICTIONARY) ---
        aggregation_fields = {}
        field_names_to_sum = [
            'ctq_plan_l1', 'ctq_actual_l1', 'pdi_plan_l1', 'pdi_actual_l1', 'other_plan_l1', 'other_actual_l1',
            'ctq_plan_l2', 'ctq_actual_l2', 'pdi_plan_l2', 'pdi_actual_l2', 'other_plan_l2', 'other_actual_l2',
            'ctq_plan_l3', 'ctq_actual_l3', 'pdi_plan_l3', 'pdi_actual_l3', 'other_plan_l3', 'other_actual_l3',
            'ctq_plan_l4', 'ctq_actual_l4', 'pdi_plan_l4', 'pdi_actual_l4', 'other_plan_l4', 'other_actual_l4'
        ]
        for field in field_names_to_sum:
            aggregation_fields[field] = Sum(field)

        # --- 4. Query and Aggregate ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)
        
        # --- 5. Clean up None values and return ---
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)
        


    @action(detail=False, methods=['get'], url_path='trend-data')
    def trend_data(self, request):
        """
        A flexible endpoint to return trend data for different metrics.
        Accepts a 'data_key' to switch between production, manpower, etc.
        Groups data by 'monthly', 'weekly', or 'daily'.
        """
        # --- 1. Get Parameters ---
        factory_id = request.query_params.get('factory')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        group_by = request.query_params.get('group_by')
        data_key = request.query_params.get('data_key', 'production')

        if not all([factory_id, start_date_str, end_date_str, group_by]):
            return Response({'error': 'Required parameters missing'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 2. Determine which fields to aggregate based on data_key ---
        if data_key == 'production':
            plan_field = 'total_production_plan'
            actual_field = 'total_production_actual'
        elif data_key == 'manpower':
            plan_field = 'total_operators_required_plan'
            actual_field = 'total_operators_required_actual'
        else:
            return Response({'error': 'Invalid data_key'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Build Filters and get the base queryset ---
        filters = {
            'factory_id': factory_id,
            'start_date__lte': end_date,
            'end_date__gte': start_date,
        }
        # --- EDIT THIS SECTION ---
        # Add optional hierarchy filters
        Hq_id = request.query_params.get('hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        station_id = request.query_params.get('station')
        
        if Hq_id: filters['Hq_id'] = Hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id

        queryset = DailyProductionData.objects.filter(**filters)
        trend_data = []

        # --- 4. Aggregate and format data ---
        if group_by == 'monthly':
            for year in range(start_date.year, end_date.year + 1):
                for month in range(1, 13):
                    if (year == start_date.year and month < start_date.month) or \
                    (year == end_date.year and month > end_date.month):
                        continue
                    
                    aggregation = queryset.filter(start_date__year=year, start_date__month=month).aggregate(
                        plan_sum=Sum(plan_field),
                        actual_sum=Sum(actual_field)
                    )
                    trend_data.append({
                        "name": f"{calendar.month_name[month][:3]} '{str(year)[2:]}",
                        "planned": aggregation['plan_sum'] or 0,
                        "actual": aggregation['actual_sum'] or 0,
                    })

        elif group_by == 'weekly':
            # THIS IS THE NEW, CORRECTED WEEKLY LOGIC
            # Find all unique weekly records within the date range and return them.
            weekly_records = queryset.filter(entry_mode='WEEKLY').order_by('start_date')
            
            for record in weekly_records:
                trend_data.append({
                    "name": f"Week of {record.start_date.strftime('%b %d')}",
                    "planned": getattr(record, plan_field), # Use getattr to get field value by name
                    "actual": getattr(record, actual_field),
                })
        
        # (The 'daily' grouping is omitted as the frontend is no longer asking for it, but can be added back if needed)
        
        return Response(trend_data, status=status.HTTP_200_OK)

    

    @action(detail=False, methods=['get'], url_path='get-plan-data')
    def get_plan_data(self, request):
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        if not all([factory_id, start_date_str, end_date_str]):
            return Response({'error': 'Factory, start_date, and end_date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        
        filters = {
            'Hq_id': Hq_id,
            'factory_id': factory_id,
            'department_id': department_id,
            'line_id': line_id,
            'subline_id': subline_id,
            'station_id': station_id,
            'start_date': start_date,
            'end_date': end_date,
        }
        filters = {k: v for k, v in filters.items() if v and v != 'all'}
        
        try:
            plan_entry = DailyProductionData.objects.get(**filters)
            serializer = self.get_serializer(plan_entry)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except DailyProductionData.DoesNotExist:
            return Response({'message': 'No data found for this period.'}, status=status.HTTP_404_NOT_FOUND)
        except DailyProductionData.MultipleObjectsReturned:
            return Response({'error': 'Multiple entries found for this period. Please ensure unique data for the specified dates.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            

    @action(detail=False, methods=['get'], url_path='get-ending-team')
    def get_ending_team(self, request):
            """
            Calculates the ending team for a given period to be used as the
            starting team for the next period.
            """
            factory_id = request.query_params.get('factory')
            Hq_id = request.query_params.get('Hq')
            department_id = request.query_params.get('department')
            subline_id = request.query_params.get('subline')

            # shop_floor_id = request.query_params.get('shop_floor')
            line_id = request.query_params.get('line')
            station_id = request.query_params.get('station')
            # We will ask for the START date of the *previous* period
            target_date_str = request.query_params.get('target_date')

            if not all([factory_id, target_date_str]):
                return Response({'error': 'Factory and target_date are required.'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)
                
            filters = {
                'Hq_id': Hq_id,
                'factory_id': factory_id,
                'department_id': department_id,
                'line_id': line_id,
                'subline_id': subline_id,
                'station_id': station_id,
                'start_date__lte': target_date, 'end_date__gte': target_date,
            }
            filters = {k: v for k, v in filters.items() if v and v != 'all'}
            
            try:
                previous_entry = DailyProductionData.objects.get(**filters)
                
                # Perform the 'Ending Team' calculation
                starting_team = previous_entry.total_operators_available
                attrition_rate = previous_entry.attrition_rate
                
                # Use Decimal for precision
                from decimal import Decimal
                ending_team = Decimal(starting_team) * (Decimal(1) - (Decimal(attrition_rate) / Decimal(100)))
                
                # Return the calculated value, rounded to a whole number
                return Response({'ending_team': round(ending_team)}, status=status.HTTP_200_OK)
                
            except DailyProductionData.DoesNotExist:
                # If there's no previous data, we can't calculate, so we return 0
                return Response({'ending_team': 0}, status=status.HTTP_404_NOT_FOUND)
            

    @action(detail=False, methods=['get'], url_path='get-month-lock-status')
    def get_month_lock_status(self, request):
        """
        Investigates a month to see if it has any data and what its lock mode is.
        """
        factory_id = request.query_params.get('factory')
        target_date_str = request.query_params.get('target_date')
        
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
            month = target_date.month
            year = target_date.year
        except:
            return Response({'error': 'Invalid date'}, status=status.HTTP_400_BAD_REQUEST)

        filters = {
            'Hq_id': request.query_params.get('Hq'),
            'factory_id': factory_id,
            'department_id': request.query_params.get('department'),
            'line_id': request.query_params.get('line'),
            'subline_id': request.query_params.get('subline'),
            'station_id': request.query_params.get('station'),
            'start_date__year': year,
            'start_date__month': month,
        }
        filters = {k: v for k, v in filters.items() if v and v != 'all'}
        
        first_entry = DailyProductionData.objects.filter(**filters).first()
        
        if first_entry:
            return Response({'lock_mode': first_entry.entry_mode}, status=status.HTTP_200_OK)
        else:
            # Important to send 200 OK with null, so frontend knows the check was successful
            return Response({'lock_mode': None}, status=status.HTTP_200_OK)
        

    @action(detail=False, methods=['get'], url_path='get-period-summary')
    def get_period_summary(self, request):
        """
        This is our "Summary Chef".
        It aggregates all data for a specific period (month or week)
        to power the summary cards and stats components.
        """
        # --- 1. Get all the filter parameters from the control panel ---
        factory_id = request.query_params.get('factory')
        # shop_floor_id = request.query_params.get('shop_floor')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not all([factory_id, start_date_str, end_date_str]):
            return Response(
                {'error': 'Factory, start_date, and end_date are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response(
                {'error': 'Invalid date format. Use YYYY-MM-DD.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Build the filter dictionary dynamically ---
        # This correctly handles the "All" selections from the frontend.
        filters = {
            'factory_id': factory_id,
            # Find any record that overlaps with the selected date range
            'start_date__lte': end_date,
            'end_date__gte': start_date,
        }
        if Hq_id and Hq_id != 'all':
            filters['Hq_id'] = Hq_id
        if department_id and department_id != 'all':
            filters['department_id'] = department_id
        if subline_id and subline_id != 'all':
            filters['subline_id'] = subline_id
        if line_id and line_id != 'all':
            filters['line_id'] = line_id
        if station_id and station_id != 'all':
            filters['station_id'] = station_id
        # --- 3. Define all the fields we need to aggregate (Sum or Average) ---
        from django.db.models import Sum, Avg

        aggregation_fields = {
            # Main Totals
            'total_production_plan': Sum('total_production_plan'),
            'total_production_actual': Sum('total_production_actual'),
            'total_operators_available': Sum('total_operators_available'),
            'total_operators_required_plan': Sum('total_operators_required_plan'),
            'total_operators_required_actual': Sum('total_operators_required_actual'),
            # Rates should be averaged
            'attrition_rate': Avg('attrition_rate'),
            'absenteeism_rate': Avg('absenteeism_rate'),
            # Bifurcation fields
            'bifurcation_plan_l1': Sum('bifurcation_plan_l1'),
            'bifurcation_actual_l1': Sum('bifurcation_actual_l1'),
            'bifurcation_plan_l2': Sum('bifurcation_plan_l2'),
            'bifurcation_actual_l2': Sum('bifurcation_actual_l2'),
            'bifurcation_plan_l3': Sum('bifurcation_plan_l3'),
            'bifurcation_actual_l3': Sum('bifurcation_actual_l3'),
            'bifurcation_plan_l4': Sum('bifurcation_plan_l4'),
            'bifurcation_actual_l4': Sum('bifurcation_actual_l4'),
            # You can add all your CTQ, PDI, Other fields here if needed for stats
        }

        # --- 4. Perform the database query ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)

        # --- 5. Clean up the data and send the response ---
        # If the query finds nothing, aggregate returns None for all fields.
        # We need to replace None with 0 so the frontend doesn't crash.
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)
    

    # Add this code inside your DailyProductionDataViewSet class in views.py

    @action(detail=False, methods=['get'], url_path='monthly-summary')
    def monthly_summary(self, request):
        """
        Aggregates bifurcation and other key data for a specific calendar month.
        Used by the main dashboard cards.
        """
        # --- 1. Get query parameters ---
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        # shop_floor_id = request.query_params.get('shop_floor')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        month = request.query_params.get('month')
        year = request.query_params.get('year')

        if not all([factory_id, month, year]):
            return Response(
                {'error': 'Factory, month, and year are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Build the filter dictionary ---
        try:
            filters = {
                'factory_id': factory_id,
                'start_date__year': int(year),
                'start_date__month': int(month),
            }
            if Hq_id: filters['Hq_id'] = Hq_id
            if department_id: filters['department_id'] = department_id
            if subline_id: filters['subline_id'] = subline_id
            if line_id: filters['line_id'] = line_id
            if station_id: filters['station_id'] = station_id
        except (ValueError, TypeError):
            return Response({'error': 'Invalid year or month format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Define aggregation fields ---
        from django.db.models import Sum

        aggregation_fields = {
            'bifurcation_plan_l1': Sum('bifurcation_plan_l1'),
            'bifurcation_actual_l1': Sum('bifurcation_actual_l1'),
            'bifurcation_plan_l2': Sum('bifurcation_plan_l2'),
            'bifurcation_actual_l2': Sum('bifurcation_actual_l2'),
            'bifurcation_plan_l3': Sum('bifurcation_plan_l3'),
            'bifurcation_actual_l3': Sum('bifurcation_actual_l3'),
            'bifurcation_plan_l4': Sum('bifurcation_plan_l4'),
            'bifurcation_actual_l4': Sum('bifurcation_actual_l4'),
        }

        # --- 4. Query and Aggregate ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)
        
        # --- 5. Clean up None values and return ---
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)


    @action(detail=False, methods=['get'], url_path='aggregated-weekly-data')
    def aggregated_weekly_data(self, request):
        """
        Aggregates bifurcation and other key data for a 7-day period (a week).
        Used by the main dashboard cards in weekly view.
        """
        # --- 1. Get query parameters ---
        factory_id = request.query_params.get('factory')
        # shop_floor_id = request.query_params.get('shop_floor')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        subline_id = request.query_params.get('subline')
        line_id = request.query_params.get('line')
        station_id = request.query_params.get('station')
        start_date_str = request.query_params.get('start_date')

        if not all([factory_id, start_date_str]):
            return Response(
                {'error': 'Factory and start_date are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Calculate date range and build filters ---
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = start_date + timedelta(days=6)

            filters = {
                'factory_id': factory_id,
                'start_date__lte': end_date,
                'end_date__gte': start_date,
            }
            if Hq_id and Hq_id != 'all':
                filters['hq_id'] = Hq_id
            if department_id and department_id != 'all':
                filters['department_id'] = department_id
            if subline_id and subline_id != 'all':
                filters['subline_id'] = subline_id
            if line_id and line_id != 'all':
                filters['line_id'] = line_id
            if station_id and station_id != 'all':
                filters['station_id'] = station_id
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Define aggregation fields (same as monthly) ---
        from django.db.models import Sum

        aggregation_fields = {
            'bifurcation_plan_l1': Sum('bifurcation_plan_l1'),
            'bifurcation_actual_l1': Sum('bifurcation_actual_l1'),
            'bifurcation_plan_l2': Sum('bifurcation_plan_l2'),
            'bifurcation_actual_l2': Sum('bifurcation_actual_l2'),
            'bifurcation_plan_l3': Sum('bifurcation_plan_l3'),
            'bifurcation_actual_l3': Sum('bifurcation_actual_l3'),
            'bifurcation_plan_l4': Sum('bifurcation_plan_l4'),
            'bifurcation_actual_l4': Sum('bifurcation_actual_l4'),
        }

        # --- 4. Query and Aggregate ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)
        
        # --- 5. Clean up None values and return ---
        for key, value in summary_data.items():
            if value is None:
                summary_data[key] = 0
        
        return Response(summary_data, status=status.HTTP_200_OK)
    


    @action(detail=False, methods=['get'], url_path='gap-analysis')
    def gap_analysis(self, request):
        """
        Calculates the gap between plan and actual values for a given period.
        """
        # --- 1. Get and Validate Parameters ---
        factory_id = request.query_params.get('factory')
        hq_id = request.query_params.get('hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        station_id = request.query_params.get('station')
        month_name = request.query_params.get('month')
        year_str = request.query_params.get('year')

        if not all([factory_id, month_name, year_str]):
            return Response(
                {'error': 'Factory, month, and year are required parameters.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # --- 2. Convert Month Name to Number ---
        try:
            # Create a mapping from month name to month number
            month_map = {name: num for num, name in enumerate(calendar.month_name) if num}
            month_num = month_map.get(month_name.capitalize())
            if not month_num:
                raise ValueError("Invalid month name")
            
            year = int(year_str)
        except (ValueError, TypeError, AttributeError):
            return Response({'error': 'Invalid month or year format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 3. Build Database Filters ---
        filters = {
            'factory_id': factory_id,
            'start_date__year': year,
            'start_date__month': month_num,
        }
        # Add optional filters only if they are provided
        if hq_id: filters['Hq_id'] = hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id
        # You can add shop_floor and line here in the same way if needed
        
        # --- 4. Define All Aggregation Fields ---
        aggregation_fields = {
            'plan_production': Sum('total_production_plan'),
            'actual_production': Sum('total_production_actual'),
            'plan_operators': Sum('total_operators_required_plan'),
            'actual_operators': Sum('total_operators_required_actual'),
            'plan_ctq_l1': Sum('ctq_plan_l1'), 'actual_ctq_l1': Sum('ctq_actual_l1'),
            'plan_ctq_l2': Sum('ctq_plan_l2'), 'actual_ctq_l2': Sum('ctq_actual_l2'),
            'plan_ctq_l3': Sum('ctq_plan_l3'), 'actual_ctq_l3': Sum('ctq_actual_l3'),
            'plan_ctq_l4': Sum('ctq_plan_l4'), 'actual_ctq_l4': Sum('ctq_actual_l4'),
            # Add PDI, Other, etc. in the same pattern if needed by the frontend
        }

        # --- 5. Query the Database ---
        summary_data = DailyProductionData.objects.filter(**filters).aggregate(**aggregation_fields)

        # --- 6. Calculate Gaps and Structure the Response ---
        # Helper function to create the plan/actual/gap structure
        def create_metric(plan_key, actual_key):
            plan = summary_data.get(plan_key) or 0
            actual = summary_data.get(actual_key) or 0
            return {
                "plan": plan,
                "actual": actual,
                "gap": plan - actual
            }

        response_payload = {
            "production": create_metric('plan_production', 'actual_production'),
            "operators": create_metric('plan_operators', 'actual_operators'),
            "ctq_l1": create_metric('plan_ctq_l1', 'actual_ctq_l1'),
            "ctq_l2": create_metric('plan_ctq_l2', 'actual_ctq_l2'),
            "ctq_l3": create_metric('plan_ctq_l3', 'actual_ctq_l3'),
            "ctq_l4": create_metric('plan_ctq_l4', 'actual_ctq_l4'),
        }

        return Response(response_payload, status=status.HTTP_200_OK)
    


   
    @action(detail=False, methods=['get'], url_path='buffer-analysis')
    def buffer_analysis(self, request):
        """
        Performs a two-part analysis for a target period:
        1. Calculates the TOTAL production volume lost due to the overall operator gap.
        2. Calculates the NET AVAILABLE manpower and GAP for each of the 4 levels,
           factoring in attrition and absenteeism.
        """
        # --- 1. Get Simplified Parameters (Unchanged) ---
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        station_id = request.query_params.get('station')
        target_date_str = request.query_params.get('start_date')

        if not all([factory_id, target_date_str]):
            return Response({'error': 'Factory and start_date are required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 2. Find the Single Target Record (Unchanged) ---
        filters = {
            'factory_id': factory_id,
            'start_date__lte': target_date,
            'end_date__gte': target_date,
        }
        if Hq_id: filters['hq_id'] = Hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id

        matching_records = DailyProductionData.objects.filter(**filters)

        # From the matches, find the one with the shortest duration (most specific)
        # We annotate the query with the duration and order by it, ascending.
        target_record = matching_records.annotate(
            duration=F('end_date') - F('start_date')
        ).order_by('duration').first()

        # If no data exists, return a default structure for consistency
        if not target_record:
            return Response({
                "lost_production_due_to_operator_gap": 0,
                "manpower_analysis_by_level": []
            }, status=status.HTTP_200_OK)

        # --- 3. NEW: Calculate Total Lost Production Volume ---
        lost_production_volume = Decimal('0.0')

        # Prevent division by zero
        if target_record.total_operators_available > 0:
            # Calculate the average production per available operator ("$B$4" from the formula)
            avg_prod_per_operator = (
                Decimal(target_record.total_production_actual) /
                Decimal(target_record.total_operators_available)
            )

            # Check if there was an operator shortfall
            if target_record.total_operators_required_actual < target_record.total_operators_required_plan:
                # Calculate the number of missing operators
                operator_gap = (
                    target_record.total_operators_required_plan -
                    target_record.total_operators_required_actual
                )
                # Calculate the final lost volume
                lost_production_volume = operator_gap * avg_prod_per_operator

        # --- 4. Per-Level Manpower Gap Calculation (Your original logic) ---
        level_analysis_data = []
        attrition_rate = Decimal(target_record.attrition_rate or 0) / Decimal(100)
        absenteeism_rate = Decimal(target_record.absenteeism_rate or 0) / Decimal(100)

        for level in range(1, 5):
            available_field = f'bifurcation_actual_l{level}'
            required_field = f'bifurcation_plan_l{level}'

            operators_available = Decimal(getattr(target_record, available_field, 0))
            operators_required = Decimal(getattr(target_record, required_field, 0))

            net_available = operators_available - (operators_available * absenteeism_rate) - (operators_available * attrition_rate)
            net_available_rounded = int(net_available)
            gap = max(0, int(operators_required - Decimal(net_available_rounded)))

            level_analysis_data.append({
                "name": f"Level {level}",
                "operators_required": int(operators_required),
                "net_available": net_available_rounded,
                "gap": gap,
            })

        # --- 5. Combine results into a single structured response ---
        response_data = {
            "lost_production_due_to_operator_gap": round(lost_production_volume, 2),
            "manpower_analysis_by_level": level_analysis_data
        }

        return Response(response_data, status=status.HTTP_200_OK)
    

    
    # In views.py, replace the gap_volume_analysis function

    @action(detail=False, methods=['get'], url_path='gap-volume-analysis')
    def gap_volume_analysis(self, request):
        factory_id = request.query_params.get('factory')
        Hq_id = request.query_params.get('Hq')
        department_id = request.query_params.get('department')
        line_id = request.query_params.get('line')
        subline_id = request.query_params.get('subline')
        target_date_str = request.query_params.get('start_date') # Only need one date

        if not all([factory_id, target_date_str]):
            return Response({'error': 'Factory and start_date are required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format.'}, status=status.HTTP_400_BAD_REQUEST)

        filters = {'factory_id': factory_id, 'start_date__lte': target_date, 'end_date__gte': target_date}
        station_id = request.query_params.get('station')
        if Hq_id: filters['hq_id'] = Hq_id
        if department_id: filters['department_id'] = department_id
        if line_id: filters['line_id'] = line_id
        if subline_id: filters['subline_id'] = subline_id
        if station_id: filters['station_id'] = station_id

        # Find the single record for the target period
        target_record = DailyProductionData.objects.filter(**filters).first()

        if not target_record:
            return Response({'L1': 0, 'L2': 0, 'L3': 0, 'L4': 0}, status=status.HTTP_200_OK)

        gap_data = {}
        attrition_rate = Decimal(target_record.attrition_rate or 0) / Decimal(100)
        absenteeism_rate = Decimal(target_record.absenteeism_rate or 0) / Decimal(100)

        for level in range(1, 5):
            # Use bifurcation fields which are the sums of departments
            operators_available = Decimal(getattr(target_record, f'bifurcation_actual_l{level}', 0))
            operators_required = Decimal(getattr(target_record, f'bifurcation_plan_l{level}', 0))
            net_available = operators_available - (operators_available * absenteeism_rate) - (operators_available * attrition_rate)
            gap_volume = int(operators_required - net_available)
            gap_data[f'L{level}'] = max(0, gap_volume)
        
        return Response(gap_data, status=status.HTTP_200_OK)

    
def get_planning_data(request):
    # --- 1. Get and Validate Parameters ---
    Hq_id = request.GET.get('Hq')
    department_id = request.GET.get('department')
    line_id = request.GET.get('line')
    subline_id = request.GET.get('subline')
    station_id = request.GET.get('station')
    factory_id = request.GET.get('factory')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    group_by = request.GET.get('group_by') # No default, we need it to be explicit

    if not all([factory_id, start_date_str, end_date_str, group_by]):
        return JsonResponse({'success': False, 'error': 'Factory, start_date, end_date, and group_by are required.'}, status=400)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    # --- 2. Build Base Filters ---
    filters = {
        'factory_id': factory_id,
        'start_date__lte': end_date,
        'end_date__gte': start_date,
    }
    # Add optional filters here if needed (station, line, etc.)
    # station_id = request.GET.get('station')
    if station_id:
        filters['station_id'] = station_id
    if Hq_id: filters['hq_id'] = Hq_id
    if department_id: filters['department_id'] = department_id
    if line_id: filters['line_id'] = line_id
    if subline_id: filters['subline_id'] = subline_id
    if station_id: filters['station_id'] = station_id
    
    queryset = DailyProductionData.objects.filter(**filters)
    graph_data = []

    # --- 3. Aggregate data based on the group_by parameter ---
    if group_by == 'month':
        # THIS IS THE NEW, SIMPLER, AND CORRECT LOOP FOR MONTHS
        current_year = start_date.year
        current_month = start_date.month
        while (current_year < end_date.year) or (current_year == end_date.year and current_month <= end_date.month):
            month_name = calendar.month_name[current_month]
            aggregation = queryset.filter(
                start_date__year=current_year,
                start_date__month=current_month
            ).aggregate(
                total_plan_prod=Sum('total_production_plan'),
                total_actual_manpower=Sum('total_operators_required_actual'),
                total_operators_plan=Sum('total_operators_required_plan'),
                total_operators_actual=Sum('total_operators_required_actual')
            )
            graph_data.append({
                "label": month_name,
                "planned_production": aggregation['total_plan_prod'] or 0,
                "actual_manpower": aggregation['total_actual_manpower'] or 0,
                "total_operators_required_plan": aggregation['total_operators_plan'] or 0,
                "total_operators_required_actual": aggregation['total_operators_actual'] or 0,
            })
            # Advance to the next month
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

    elif group_by == 'week':
        current_week_start = start_date
        while current_week_start <= end_date:
            current_week_end = current_week_start + timedelta(days=6)
            week_label = f"Week of {current_week_start.strftime('%b %d')}"
            aggregation = queryset.filter(
                start_date__lte=current_week_end,
                end_date__gte=current_week_start
            ).aggregate(
                total_plan_prod=Sum('total_production_plan'),
                total_actual_manpower=Sum('total_operators_required_actual'),
                total_operators_plan=Sum('total_operators_required_plan'),
                total_operators_actual=Sum('total_operators_required_actual')
            )
            graph_data.append({
                "label": week_label,
                "planned_production": aggregation['total_plan_prod'] or 0,
                "actual_manpower": aggregation['total_actual_manpower'] or 0,
                "total_operators_required_plan": aggregation['total_operators_plan'] or 0,
                "total_operators_required_actual": aggregation['total_operators_actual'] or 0,
            })
            current_week_start += timedelta(days=7)

    response_payload = { "success": True, "data": graph_data }
    return JsonResponse(response_payload, status=200)





from django.http import JsonResponse
from django.db.models import Sum
from django.utils import timezone
import calendar
from datetime import datetime
from dateutil import parser

def get_operators_required_trend(request):
    factory_id = request.GET.get('factory')
    time_view = request.GET.get('time_view', 'Monthly')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if not factory_id or not start_date_str or not end_date_str:
        return JsonResponse({'error': 'Factory, start_date, and end_date are required.'}, status=400)

    try:
        start_date = parser.parse(start_date_str).date()
        end_date = parser.parse(end_date_str).date()
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date format.'}, status=400)

    if start_date > end_date:
        return JsonResponse({'error': 'start_date cannot be later than end_date.'}, status=400)

    filters = {'factory_id': factory_id}
    # --- EDIT THIS SECTION ---
    if request.GET.get('Hq'):
        filters['Hq_id'] = request.GET.get('Hq')
    if request.GET.get('department'):
        filters['department_id'] = request.GET.get('department')
    if request.GET.get('line'):
        filters['line_id'] = request.GET.get('line')
    if request.GET.get('subline'):
        filters['subline_id'] = request.GET.get('subline')
    if request.GET.get('station'):
        filters['station_id'] = request.GET.get('station')
    # if request.GET.get('shop_floor'): # Keep if still in use
    #     filters['shop_floor_id'] = request.GET.get('shop_floor')

    trend_data = []
    data_type = time_view

    if time_view == 'Weekly':
        # Existing weekly logic (working fine)
        from dateutil.rrule import rrule, WEEKLY
        weeks = list(rrule(WEEKLY, dtstart=start_date, until=end_date))
        for week_start in weeks:
            week_end = min(week_start.date() + timedelta(days=6), end_date)
            weekly_filters = filters.copy()
            weekly_filters['start_date__range'] = (week_start.date(), week_end)
            weekly_aggregation = DailyProductionData.objects.filter(**weekly_filters).aggregate(
                plan_sum=Sum('total_operators_required_plan'),
                actual_sum=Sum('total_operators_required_actual')
            )
            trend_data.append({
                'period': week_start.date().isoformat(),
                'total_operators_required_plan': weekly_aggregation['plan_sum'] or 0,
                'total_operators_required_actual': weekly_aggregation['actual_sum'] or 0,
            })
    else:  # Monthly
        year = start_date.year
        for month in range(1, 13):
            # Create filters for this specific month and year
            monthly_filters = filters.copy()
            monthly_filters['start_date__year'] = year
            monthly_filters['start_date__month'] = month

            # Aggregate data for just this month
            monthly_aggregation = DailyProductionData.objects.filter(**monthly_filters).aggregate(
                plan_sum=Sum('total_operators_required_plan'),
                actual_sum=Sum('total_operators_required_actual')
            )
            
            # Append this month's data to the list
            trend_data.append({
                'period': calendar.month_name[month],
                'year': year,
                'total_operators_required_plan': monthly_aggregation['plan_sum'] or 0,
                'total_operators_required_actual': monthly_aggregation['actual_sum'] or 0,
            })

    return JsonResponse({
        'data_type': data_type,
        'data': trend_data
    }, safe=False, status=200)





from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Sum, Avg
from .models import DailyProductionData
from datetime import datetime
from dateutil import parser
import calendar
from math import floor, ceil

@require_GET
def monthly_availability_analysis(request):
    """
    Monthly gap analysis endpoint for /production-data/gap-analysis/.
    Computes GAP, gap_message, and Lost Production for multiple production plan scenarios.
    """
    # Get query parameters
    factory_id = request.GET.get('factory')
    month = request.GET.get('month')  # e.g., "March"
    year = request.GET.get('year')  # e.g., "2025"
    # shop_floor_id = request.GET.get('shop_floor')
    line_id = request.GET.get('line')
    station_id = request.GET.get('station')
    Hq_id = request.GET.get('Hq')
    department_id = request.GET.get('department')
    subline_id = request.GET.get('subline')

    # Validate required parameters
    if not (factory_id and month and year):
        return JsonResponse(
            {"success": False, "error": "Missing required parameters: factory, month, year"},
            status=400
        )

    try:
        year = int(year)
        month_num = list(calendar.month_name).index(month.capitalize())
    except (ValueError, IndexError):
        return JsonResponse(
            {"success": False, "error": "Invalid month or year format"},
            status=400
        )

    # Determine date range for the month
    start_date = datetime(year, month_num, 1)
    _, last_day = calendar.monthrange(year, month_num)
    end_date = datetime(year, month_num, last_day)

    # Calculate dynamic productivity rate from previous month's data
    prev_month = start_date.replace(day=1, month=month_num - 1 if month_num > 1 else 12)
    prev_year = year if month_num > 1 else year - 1
    prev_data = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__year=prev_year,
        start_date__month=prev_month.month,
        entry_mode='MONTHLY'
    ).aggregate(
        total_production_actual=Sum('total_production_actual'),
        total_operators_actual=Sum('total_operators_required_actual')
    )
    productivity_rate = (
        prev_data['total_production_actual'] / prev_data['total_operators_actual']
        if prev_data['total_production_actual'] and prev_data['total_operators_actual']
        else 12  # Fallback to Excel's default
    )

    # Build queryset with filters
    queryset = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__gte=start_date,
        end_date__lte=end_date,
        entry_mode='MONTHLY'
    )
    if Hq_id:
        queryset = queryset.filter(Hq_id=Hq_id)
    if department_id:
        queryset = queryset.filter(department_id=department_id)
    if subline_id:
        queryset = queryset.filter(subline_id=subline_id)
    # if shop_floor_id:
    #     queryset = queryset.filter(shop_floor_id=shop_floor_id)
    if line_id:
        queryset = queryset.filter(line_id=line_id)
    if station_id:
        queryset = queryset.filter(station_id=station_id)

    # Process multiple scenarios
    data_points = []
    for record in queryset:
        # Calculate Excel-derived values
        ending_team = record.total_operators_available * (1 - record.attrition_rate / 100)
        operators_available = floor(ending_team * (1 - record.absenteeism_rate / 100))  # ROUNDDOWN
        gap = operators_available - record.total_operators_required_plan
        lost_production = abs(gap) * productivity_rate if gap < 0 else 0
        gap_message = (
            f"Shortage: Hire {abs(gap)}" if gap < 0 else
            f"Surplus of {gap}" if gap > 0 else
            "Balanced"
        )

        # Previous month's actual operators for context
        prev_data = DailyProductionData.objects.filter(
            factory_id=factory_id,
            start_date__year=prev_year,
            start_date__month=prev_month.month,
            entry_mode='MONTHLY'
        ).aggregate(total_actual=Sum('total_operators_required_actual'))

        data_points.append({
            "month": month,
            "year": year,
            "production_plan": {"value": record.total_production_plan},
            "operators_required": {"value": record.total_operators_required_plan},
            "operators_available": {"value": operators_available},
            "gap": {"value": gap},
            "gap_message": gap_message,
            "lost_production": {"value": round(lost_production)},
            "attrition_rate": float(record.attrition_rate),
            "previous_month_data": {
                "month": calendar.month_name[prev_month.month],
                "actual_operators": prev_data['total_actual'] or 0
            } if prev_data['total_actual'] else None
        })

    if not data_points:
        return JsonResponse(
            {"success": False, "error": "No data found for the selected filters"},
            status=404
        )

    return JsonResponse({
        "success": True,
        "operators_availability_graph": data_points[0] if len(data_points) == 1 else data_points
    })

@require_GET
def weekly_availability_summary(request):
    """
    Weekly summary endpoint for /production-data/date-range-summary/.
    Aggregates data for a week and computes GAP, gap_message, and Lost Production.
    """
    # Get query parameters
    factory_id = request.GET.get('factory')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    # shop_floor_id = request.GET.get('shop_floor')
    line_id = request.GET.get('line')
    station_id = request.GET.get('station')
    Hq_id = request.GET.get('Hq')
    department_id = request.GET.get('department')
    subline_id = request.GET.get('subline')

    # Validate required parameters
    if not (factory_id and start_date_str and end_date_str):
        return JsonResponse(
            {"success": False, "error": "Missing required parameters: factory, start_date, end_date"},
            status=400
        )

    try:
        start_date = parser.parse(start_date_str).date()
        end_date = parser.parse(end_date_str).date()
    except ValueError:
        return JsonResponse(
            {"success": False, "error": "Invalid date format"},
            status=400
        )

    # Calculate dynamic productivity rate from previous week's data
    prev_week_end = start_date - timedelta(days=1)
    prev_week_start = prev_week_end - timedelta(days=6)
    prev_data = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__gte=prev_week_start,
        end_date__lte=prev_week_end,
        entry_mode='WEEKLY'
    ).aggregate(
        total_production_actual=Sum('total_production_actual'),
        total_operators_actual=Sum('total_operators_required_actual')
    )
    productivity_rate = (
        prev_data['total_production_actual'] / prev_data['total_operators_actual']
        if prev_data['total_production_actual'] and prev_data['total_operators_actual']
        else 12  # Fallback to Excel's default
    )

    # Build queryset with filters
    queryset = DailyProductionData.objects.filter(
        factory_id=factory_id,
        start_date__gte=start_date,
        end_date__lte=end_date,
        entry_mode='WEEKLY'
    )
    if Hq_id:
        queryset = queryset.filter(Hq_id=Hq_id)
    if department_id:
        queryset = queryset.filter(department_id=department_id)
    if subline_id:
        queryset = queryset.filter(subline_id=subline_id)
    # if shop_floor_id:
    #     queryset = queryset.filter(shop_floor_id=shop_floor_id)
    if line_id:
        queryset = queryset.filter(line_id=line_id)
    if station_id:
        queryset = queryset.filter(station_id=station_id)

    # Aggregate data for the week
    aggregates = queryset.aggregate(
        total_production_plan=Sum('total_production_plan'),
        total_operators_required_plan=Sum('total_operators_required_plan'),
        total_operators_available=Sum('total_operators_available'),
        avg_attrition_rate=Avg('attrition_rate'),
        avg_absenteeism_rate=Avg('absenteeism_rate')
    )

    if not aggregates['total_production_plan']:
        return JsonResponse(
            {"success": False, "error": "No data found for the selected week"},
            status=404
        )

    # Calculate Excel-derived values
    ending_team = aggregates['total_operators_available'] * (1 - aggregates['avg_attrition_rate'] / 100)
    operators_available = floor(ending_team * (1 - aggregates['avg_absenteeism_rate'] / 100))  # ROUNDDOWN
    gap = operators_available - aggregates['total_operators_required_plan']
    lost_production = abs(gap) * productivity_rate if gap < 0 else 0
    gap_message = (
        f"Shortage: Hire {abs(gap)}" if gap < 0 else
        f"Surplus of {gap}" if gap > 0 else
        "Balanced"
    )

    return JsonResponse({
        "success": True,
        "total_production_plan": aggregates['total_production_plan'] or 0,
        "total_operators_required_plan": aggregates['total_operators_required_plan'] or 0,
        "total_operators_required_actual": operators_available,
        "attrition_rate": float(aggregates['avg_attrition_rate'] or 0),
        "gap": gap,
        "gap_message": gap_message,
        "lost_production": round(lost_production)
    })



from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import SkillMatrix
from .serializers import SkillMatrixSerializer

class SkillMatrixViewSet(viewsets.ModelViewSet):
    queryset = SkillMatrix.objects.all()
    serializer_class = SkillMatrixSerializer

    # Custom endpoint: filter by station_id
    @action(detail=False, methods=["get"])
    def by_station(self, request):
        station_id = request.query_params.get("station_id")
        if not station_id:
            return Response({"error": "station_id parameter is required"}, status=400)

        queryset = self.queryset.filter(hierarchy__station_id=station_id)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import HierarchyStructure

class HierarchyAllDepartmentsView(APIView):
    """
    Fetch hierarchy for all departments with flexible nesting:
    department → line → subline → station
    department → line → station
    department → subline → station
    department → station
    """

    def get(self, request, *args, **kwargs):
        try:
            structures = HierarchyStructure.objects.select_related(
                "department", "line", "subline", "station"
            )

            if not structures.exists():
                return Response(
                    {"error": "No hierarchy data found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            departments_data = {}

            for structure in structures:
                department = structure.department
                if not department:
                    continue  # skip if no department linked

                dept_id = department.department_id
                if dept_id not in departments_data:
                    departments_data[dept_id] = {
                        "department_id": dept_id,
                        "department_name": department.department_name,
                        "lines": {},
                        "sublines": {},
                        "stations": {},
                    }

                department_data = departments_data[dept_id]
                line = structure.line
                subline = structure.subline
                station = structure.station

                if line:
                    if line.line_id not in department_data["lines"]:
                        department_data["lines"][line.line_id] = {
                            "line_id": line.line_id,
                            "line_name": line.line_name,
                            "sublines": {},
                            "stations": {},
                        }

                    if subline:
                        if subline.subline_id not in department_data["lines"][line.line_id]["sublines"]:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id] = {
                                "subline_id": subline.subline_id,
                                "subline_name": subline.subline_name,
                                "stations": {},
                            }

                        if station:
                            department_data["lines"][line.line_id]["sublines"][subline.subline_id]["stations"][
                                station.station_id
                            ] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }
                    else:
                        if station:
                            department_data["lines"][line.line_id]["stations"][station.station_id] = {
                                "station_id": station.station_id,
                                "station_name": station.station_name,
                            }

                elif subline:
                    if subline.subline_id not in department_data["sublines"]:
                        department_data["sublines"][subline.subline_id] = {
                            "subline_id": subline.subline_id,
                            "subline_name": subline.subline_name,
                            "stations": {},
                        }

                    if station:
                        department_data["sublines"][subline.subline_id]["stations"][station.station_id] = {
                            "station_id": station.station_id,
                            "station_name": station.station_name,
                        }

                elif station:
                    department_data["stations"][station.station_id] = {
                        "station_id": station.station_id,
                        "station_name": station.station_name,
                    }

            # convert dicts to lists for clean JSON
            for dept_id, department_data in departments_data.items():
                department_data["lines"] = list(department_data["lines"].values())
                for line in department_data["lines"]:
                    line["sublines"] = list(line["sublines"].values())
                    line["stations"] = list(line["stations"].values())
                    for subline in line["sublines"]:
                        subline["stations"] = list(subline["stations"].values())

                department_data["sublines"] = list(department_data["sublines"].values())
                for subline in department_data["sublines"]:
                    subline["stations"] = list(subline["stations"].values())

                department_data["stations"] = list(department_data["stations"].values())

            return Response(list(departments_data.values()), status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404
from django.conf import settings
import os
import mimetypes
from .models import UserManualdocs
from .serializers import UserManualdocsSerializer

class UserManualdocsListCreateView(generics.ListCreateAPIView):
    queryset = UserManualdocs.objects.all()
    serializer_class = UserManualdocsSerializer
    parser_classes = [MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    'message': 'Document uploaded successfully!',
                    'data': serializer.data
                },
                status=status.HTTP_201_CREATED
            )
        return Response(
            {
                'message': 'Failed to upload document',
                'errors': serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )

class UserManualdocsDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = UserManualdocs.objects.all()
    serializer_class = UserManualdocsSerializer
    parser_classes = [MultiPartParser, FormParser]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(
            {'message': 'Document deleted successfully!'},
            status=status.HTTP_200_OK
        )

@api_view(['GET'])
def view_file(request, doc_id):
    try:
        doc = get_object_or_404(UserManualdocs, id=doc_id)
        
        if not doc.file:
            return Response(
                {'error': 'No file associated with this document'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        file_path = doc.file.path
        
        if not os.path.exists(file_path):
            return Response(
                {'error': 'File not found on server'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get the file's MIME type
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type is None:
            mime_type = 'application/octet-stream'
        
        # Read the file
        try:
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type=mime_type)
                
                # Set headers for inline viewing (opens in browser)
                filename = os.path.basename(file_path)
                response['Content-Disposition'] = f'inline; filename="{filename}"'
                
                return response
                
        except Exception as e:
            return Response(
                {'error': f'Error reading file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    except UserManualdocs.DoesNotExist:
        return Response(
            {'error': 'Document not found'},
            status=status.HTTP_404_NOT_FOUND
        )

@api_view(['GET'])
def download_file(request, doc_id):
    try:
        doc = get_object_or_404(UserManualdocs, id=doc_id)
        
        if not doc.file:
            return Response(
                {'error': 'No file associated with this document'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        file_path = doc.file.path
        
        if not os.path.exists(file_path):
            return Response(
                {'error': 'File not found on server'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get the file's MIME type
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type is None:
            mime_type = 'application/octet-stream'
        
        # Read the file
        try:
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type=mime_type)
                
                # Set headers for forced download
                filename = os.path.basename(file_path)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                return response
                
        except Exception as e:
            return Response(
                {'error': f'Error reading file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    except UserManualdocs.DoesNotExist:
        return Response(
            {'error': 'Document not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
from .models import EvaluationPassingCriteria, Level, Department
from .serializers import EvaluationPassingCriteriaSerializer, LevelSerializer, DepartmentSerializer

class EvaluationPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = EvaluationPassingCriteria.objects.select_related('level', 'department').all()
    serializer_class = EvaluationPassingCriteriaSerializer
    
    def create(self, request, *args, **kwargs):
        try:
            # Check if criteria already exists for this level-department combination
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).first()
            
            if existing:
                return Response(
                    {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error creating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if another criteria exists for this level-department combination
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).exclude(id=instance.id).first()
            
            if existing:
                return Response(
                    {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().update(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error updating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(
                {'detail': f'Error deleting criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 

from .models import EvaluationPassingCriteria, Level, Department
from .serializers import EvaluationPassingCriteriaSerializer, LevelSerializer, DepartmentSerializer

class EvaluationPassingCriteriaViewSet(viewsets.ModelViewSet):
    queryset = EvaluationPassingCriteria.objects.select_related('level', 'department').all()
    serializer_class = EvaluationPassingCriteriaSerializer
    
    def create(self, request, *args, **kwargs):
        try:
            # Check if criteria already exists for this level-department combination
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).first()
            
            if existing:
                return Response(
                    {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error creating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            level_id = request.data.get('level')
            department_id = request.data.get('department')
            
            if not level_id or not department_id:
                return Response(
                    {'detail': 'Level and Department are required.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if another criteria exists for this level-department combination
            existing = EvaluationPassingCriteria.objects.filter(
                level_id=level_id,
                department_id=department_id
            ).exclude(id=instance.id).first()
            
            if existing:
                return Response(
                    {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return super().update(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Another passing criteria already exists for this Level and Department combination.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error updating criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        try:
            return super().destroy(request, *args, **kwargs)
        except Exception as e:
            return Response(
                {'detail': f'Error deleting criteria: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 


class EndTestSessionView(APIView):
    def post(self, request):
        try:
            logger.info("📥 EndTestSessionView called with data: %s", request.data)
            submitted_data = request.data
            if not isinstance(submitted_data, dict):
                logger.warning("Invalid payload type: %s", type(submitted_data))
                return Response(
                    {"error": "Payload must be a dict of {key_id: [answers]}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            final_results = []
            logger.info("🔑 Submitted Key IDs: %s", submitted_data.keys())

            for key_id, answers in submitted_data.items():
                sessions = TestSession.objects.filter(key_id=key_id)
                if not sessions.exists():
                    final_results.append({"key_id": key_id, "error": "No TestSession found"})
                    continue

                for session in sessions:
                    if not session.question_paper or not session.employee:
                        error_msg = "No QuestionPaper assigned" if not session.question_paper else "No employee assigned"
                        final_results.append({"session_id": session.id, "error": error_msg})
                        continue

                    questions = list(session.question_paper.template_questions.order_by("id"))
                    total_questions = len(questions)
                    if total_questions == 0:
                        final_results.append({"session_id": session.id, "error": "No questions in QuestionPaper"})
                        continue

                    if len(answers) != total_questions:
                        final_results.append({
                            "session_id": session.id,
                            "error": f"Mismatch: {len(answers)} answers submitted, but {total_questions} questions expected"
                        })
                        continue

                    correct_count = 0
                    for i, submitted_ans in enumerate(answers):
                        question = questions[i]
                        if isinstance(submitted_ans, int) and 0 <= submitted_ans <= 3:
                            submitted_letter = chr(65 + submitted_ans)
                        else:
                            submitted_letter = None

                        options = [question.option_a, question.option_b, question.option_c, question.option_d]
                        try:
                            correct_idx = options.index(question.correct_answer)
                            correct_letter = chr(65 + correct_idx)
                            if submitted_letter == correct_letter:
                                correct_count += 1
                        except ValueError:
                            pass

                    percentage = round((correct_count / total_questions) * 100, 2)
                    passed = percentage >= 80

                    try:
                        with transaction.atomic():
                            existing_score = Score.objects.filter(
                                employee=session.employee,
                                test=session,
                                skill=session.skill
                            ).first()

                            if existing_score:
                                final_results.append({
                                    "session_id": session.id,
                                    "error": "Attempt already exists for this employee and station"
                                })
                                continue

                            # Create new score with proper department assignment
                            score = Score.objects.create(
                                employee=session.employee,
                                test=session,
                                marks=correct_count,
                                percentage=Decimal(str(percentage)),
                                passed=passed,
                                level=session.level,
                                skill=session.skill,
                                department=session.department or session.employee.department,  # Ensure department is set
                            )

                            # Get department name for response
                            department_name = "N/A"
                            if score.department:
                                department_name = score.department.department_name
                            elif session.employee.department:
                                department_name = session.employee.department.department_name

                        final_results.append({
                            "session_id": session.id,
                            "marks": correct_count,
                            "percentage": percentage,
                            "passed": passed,
                            "level": session.level.level_name if session.level else None,
                            "skill": session.skill.station_name if session.skill else None,
                            "department": department_name
                        })
                    except Exception as e:
                        logger.error("Error saving Score for session %s: %s", session.id, e)
                        final_results.append({"session_id": session.id, "error": str(e)})

            return Response({"results": final_results}, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error("Unhandled error in EndTestSessionView: %s", traceback.format_exc())
            return Response({"error": "Internal server error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ScoreListView(APIView):
    def get(self, request):
        # Assuming you use caching for latest test session
        session_key = cache.get("latest_test_session")
        if not session_key:
            return Response([])

        scores = Score.objects.filter(test__key_id=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data)


class KeyIdToEmployeeNameMap(APIView):
    def get(self, request):
        mapping = TestSession.objects.select_related('employee').all()
        return Response({
            s.key_id: f"{s.employee.first_name} {s.employee.last_name}" 
            for s in mapping
        })


class PastTestSessionsView(APIView):
    def get(self, request):
        # Get distinct test names from Score model through the test relationship
        test_names = Score.objects.select_related('test').filter(
            test__isnull=False
        ).values_list('test__test_name', flat=True).distinct()
        
        return Response(list(test_names))

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from app1.models import Score
from app1.serializers import ScoreSerializer
import traceback

class ScoresByTestView(APIView):
    def get(self, request, name):
        try:
            scores = (
                Score.objects
                .filter(test__test_name=name)
                .select_related(
                    'employee', 'employee__department',
                    'department',
                    'skill', 'skill__subline', 'skill__subline__line', 'skill__subline__line__department',
                    'level', 'test', 'test__department'
                )
            )
            serializer = ScoreSerializer(scores, many=True)
            print("Final data being sent to frontend:", serializer.data)
            return Response(serializer.data)
        except Exception as e:
            print(f"Error in ScoresByTestView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class SkillListView(APIView):
    def get(self, request):
        skills = Station.objects.values_list('station_name', flat=True).distinct()
        return Response(list(skills))


class ResultSummaryAPIView(APIView):
    def get(self, request):
        try:
            scores = Score.objects.select_related('employee', 'level', 'skill')
            data = []
            for score in scores:
                # Use the percentage from the score model instead of recalculating
                percentage = score.percentage
                result = 'Pass' if score.percentage >= 80 else 'Retraining' if score.percentage >= 50 else 'Fail'

                data.append({
                    "employee_id": score.employee.emp_id if hasattr(score.employee, 'emp_id') else score.employee.id,
                    "name": f"{score.employee.first_name} {score.employee.last_name}",
                    "marks": score.marks,
                    "percentage": percentage,
                    "section": score.employee.section if hasattr(score.employee, 'section') else '',
                    "level_name": score.level.level_name if score.level and hasattr(score.level, 'level_name') else '',
                    "skill": score.skill.station_name if score.skill and hasattr(score.skill, 'station_name') else (score.skill.skill if score.skill else ''),
                    "result": result,
                })

            serializer = SimpleScoreSerializer(data, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            print(f"Error in ResultSummaryAPIView: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from .models import MasterTable, QuestionPaper, Station, Level, TestSession, Score, TemplateQuestion

logger = logging.getLogger(__name__)

class SubmitWebTestAPIView(APIView):
    """
    Submit test answers from web/tablet exam (without remote).
    """
    def post(self, request):
        try:
            logger.info("SubmitWebTestAPIView called with data: %s", request.data)

            employee_id = request.data.get("employee_id")
            test_name = request.data.get("test_name")
            question_paper_id = request.data.get("question_paper_id")
            answers = request.data.get("answers", [])
            skill_id = request.data.get("skill_id")
            level_id = request.data.get("level_id") # FIX: Renamed variable for clarity

            # Validate required fields
            # FIX: Added level_id to the validation check
            if not all([employee_id, test_name, question_paper_id, level_id]) or not isinstance(answers, list):
                logger.warning("Validation failed: missing required fields.")
                return Response(
                    {"error": "employee_id, test_name, question_paper_id, level_id, and answers[] are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get employee, question paper, and level objects
            logger.debug("Fetching employee_id=%s", employee_id)
            employee = get_object_or_404(MasterTable, emp_id=employee_id) # FIX: Use emp_id as primary key from your model

            logger.debug("Fetching question_paper_id=%s", question_paper_id)
            question_paper = get_object_or_404(QuestionPaper, question_paper_id=question_paper_id)
            
            # FIX: Fetch the Level object using the provided level_id
            logger.debug("Fetching level_id=%s", level_id)
            level_object = get_object_or_404(Level, level_id=level_id)

            # Handle skill (station)
            skill_object = None
            if skill_id:
                logger.debug("Fetching skill (station) by id=%s", skill_id)
                skill_object = get_object_or_404(Station, station_id=skill_id)

            # FIX: Correct way to get all questions related to the question paper
            questions = list(question_paper.template_questions.all().order_by("id"))
            total_questions = len(questions)
            logger.info("Total questions found: %s", total_questions)

            if total_questions == 0:
                logger.error("No questions found for paper id=%s", question_paper_id)
                return Response({"error": "No questions found for this paper."}, status=status.HTTP_400_BAD_REQUEST)

            # Compare answers
            correct_count = 0
            # FIX: The correct answer is stored in 'correct_answer' text field in TemplateQuestion model.
            # We need to map it to an index (0=A, 1=B, 2=C, 3=D).
            for i, ans in enumerate(answers):
                if i < total_questions:
                    question = questions[i]
                    options = [question.option_a, question.option_b, question.option_c, question.option_d]
                    try:
                        correct_index = options.index(question.correct_answer)
                    except ValueError:
                        correct_index = -1 # Should not happen if data is clean
                    
                    logger.debug("Q%s: submitted=%s, correct_index=%s", i+1, ans, correct_index)
                    if ans == correct_index:
                        correct_count += 1

            # --- Get Dynamic Passing Criteria ---
            required_percentage = Decimal('80.00')  # Default
            try:
                if level_object and employee.department:
                    # passing_criteria = EvaluationPassingCriteria.objects.filter(
                    #     level=level_object,
                    #     department=employee.department
                    # ).first()
                    passing_criteria = EvaluationPassingCriteria.objects.filter(
    level=question_paper.level,
    department=question_paper.department
).first()
                    if passing_criteria:
                        required_percentage = passing_criteria.percentage
            except Exception as e:
                logger.error("Error retrieving passing criteria: %s", e)

            # --- Calculate percentage using Decimal for precision ---
            if total_questions > 0:
                percentage_decimal = Decimal(correct_count) / Decimal(total_questions) * Decimal('100')
                percentage_decimal = percentage_decimal.quantize(Decimal('0.01'))
                percentage = float(percentage_decimal)
            else:
                percentage_decimal = Decimal('0.00')
                percentage = 0.0

            passed = percentage_decimal >= required_percentage
            logger.info("Scoring: %s/%s correct (%.2f%%), passed=%s",
                        correct_count, total_questions, percentage, passed)

            # Create or update TestSession and Score
            with transaction.atomic():
                test_session, _ = TestSession.objects.get_or_create(
                    employee=employee,
                    question_paper=question_paper,
                    # FIX: Use a more robust unique key
                    key_id=f"web-{employee_id}-{question_paper_id}",
                    defaults={
                        "test_name": test_name,
                        "level": level_object, # FIX: Assign the level object
                        "skill": skill_object, # FIX: Assign the skill object
                    }
                )

                score, created = Score.objects.get_or_create(
                    employee=employee,
                    test=test_session,
                    defaults={
                        'marks': correct_count,
                        'percentage': percentage,
                        'passed': passed,
                        'skill': skill_object, # FIX: Assign the skill object
                        'level': level_object, # FIX: Assign the level object
                    }
                )

                if not created:
                    logger.info("Updating existing score for employee=%s", employee.emp_id)
                    score.marks = correct_count
                    score.percentage = percentage
                    score.passed = passed
                    score.skill = skill_object
                    score.level = level_object
                    score.save()

            return Response({
                "employee": f"{employee.first_name} {employee.last_name}",
                "marks": correct_count,
                "total_questions": total_questions,
                "percentage": percentage,
                "passed": passed,
                "level_received": level_object.level_name,
                "message": "Score saved successfully"
            }, status=status.HTTP_200_OK)

        except Exception as e:
            error_trace = traceback.format_exc()
            logger.error("Error in SubmitWebTestAPIView: %s\n%s", str(e), error_trace)
            return Response({
                "error": str(e),
                "traceback": error_trace if settings.DEBUG else "An internal server error occurred."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        


# =================== TrainingAttendance ============================= #


from django.utils import timezone
from django.db import transaction
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response

# Make sure all your models and serializers are imported
from .models import TrainingBatch, UserRegistration, TrainingAttendance, Days
from .serializers import TrainingBatchSerializer, BatchAttendanceDetailSerializer, TrainingAttendanceSerializer

# --- Active/Past Batch Views (Unchanged, but included for context) ---
class ActiveTrainingBatchListView(generics.ListAPIView):
    queryset = TrainingBatch.objects.filter(is_active=True)
    serializer_class = TrainingBatchSerializer

# class BatchAttendanceDetailView(APIView):
#     """
#     GET /api/attendance-detail/{batch_id}/
#     A powerful view that returns everything the frontend needs for the attendance page.
#     """
#     def get(self, request, batch_id, *args, **kwargs):
#         # 1. Get all users for the batch
#         users = UserRegistration.objects.filter(batch_id=batch_id)
#         if not users.exists():
#             return Response({"error": "No users found for this batch."}, status=status.HTTP_404_NOT_FOUND)

#         # 2. Get total days from Days model
#         total_days = Days.objects.count()
#         if total_days == 0:
#             return Response({"error": "No training days configured."}, status=status.HTTP_400_BAD_REQUEST)

#         # 3. Determine the next training day to mark
#         today = timezone.now().date()
#         next_day_to_mark = None
        
#         # Find the latest attendance record for this batch to see when they last marked
#         latest_attendance = TrainingAttendance.objects.filter(batch_id=batch_id).order_by('-day_number__days_id').first()

#         if not latest_attendance:
#             # No attendance has ever been marked, so start with first day
#             first_day = Days.objects.first()
#             next_day_to_mark = first_day.days_id if first_day else None
#         else:
#             last_marked_day = latest_attendance.day_number.days_id
#             last_marked_date = latest_attendance.attendance_date
            
#             # Get the next day after the last marked day
#             next_day = Days.objects.filter(days_id__gt=last_marked_day).first()
            
#             if not next_day:
#                 # Training is complete - no more days available
#                 next_day_to_mark = None
#             elif today > last_marked_date:
#                 # It's a new day, so unlock the next training day
#                 next_day_to_mark = next_day.days_id
#             else:
#                 # Attendance was already marked today, so no day is available to mark.
#                 next_day_to_mark = None
        
#         # Check if the batch is completed (all days are done)
#         last_day = Days.objects.last()
#         is_completed = False
#         if last_day:
#             is_completed = TrainingAttendance.objects.filter(batch_id=batch_id, day_number=last_day.days_id).exists()

#         # 4. Prepare the data payload
#         data = {
#             'batch_id': batch_id,
#             'next_training_day_to_mark': next_day_to_mark,
#             'is_completed': is_completed,
#             'users': users,
#             'total_days': total_days  # Add total days to response
#         }

#         # 5. Serialize the data
#         serializer = BatchAttendanceDetailSerializer(data, context={'batch_id': batch_id})
#         return Response(serializer.data, status=status.HTTP_200_OK)

from django.utils import timezone # Make sure timezone is imported

class BatchAttendanceDetailView(APIView):
    """
    GET /api/attendance-detail/{batch_id}/
    Returns all data needed for the attendance page.
    The concept of 'next_training_day_to_mark' is removed from the logic,
    as all days are now editable. It is kept in the response for compatibility but is null.
    """
    def get(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
        except TrainingBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

        users = UserRegistration.objects.filter(batch_id=batch_id)
        if not users.exists():
            return Response({"error": "No users found for this batch."}, status=status.HTTP_404_NOT_FOUND)

        data = {
            'batch_id': batch_id,
            'next_training_day_to_mark': None, # This is no longer used to lock UI
            'is_completed': not batch.is_active,
            'users': users
        }
        
        serializer = BatchAttendanceDetailSerializer(data, context={'batch_id': batch_id})
        return Response(serializer.data, status=status.HTTP_200_OK)


class PastTrainingBatchListView(generics.ListAPIView):
    queryset = TrainingBatch.objects.filter(is_active=False)
    serializer_class = TrainingBatchSerializer

class BulkAttendanceUpdateView(APIView):
    """
    POST /api/attendances/
    Creates or updates a list of attendance records. This view is designed to be
    flexible, accepting records for multiple users across multiple days in a single request.
    It uses a transaction to ensure all updates succeed or none do.
    """
    def post(self, request, *args, **kwargs):
        attendance_data = request.data
        if not isinstance(attendance_data, list):
            return Response({"error": "Expected a list of attendance objects."}, status=status.HTTP_400_BAD_REQUEST)
        if not attendance_data:
            return Response({"message": "No attendance data provided to update."}, status=status.HTTP_200_OK)

        target_batch_id = attendance_data[0].get('batch')
        if not target_batch_id:
             return Response({"error": "Batch ID is missing in payload."}, status=status.HTTP_400_BAD_REQUEST)

        created_records = []
        updated_records = []

        try:
            # Use a transaction to make the entire operation atomic
            with transaction.atomic():
                for item in attendance_data:
                    user_id = item.get('user')
                    day_id = item.get('day_number')
                    status_val = item.get('status')

                    if not all([user_id, day_id, status_val]):
                        raise ValueError(f"Invalid item in payload: {item}")

                    # Use update_or_create to handle both new and existing records
                    obj, created = TrainingAttendance.objects.update_or_create(
                        user_id=user_id,
                        batch_id=target_batch_id,
                        day_number_id=day_id,
                        defaults={
                            'status': status_val,
                            'attendance_date': timezone.now().date()
                        }
                    )
                    
                    if created:
                        created_records.append(TrainingAttendanceSerializer(obj).data)
                    else:
                        updated_records.append(TrainingAttendanceSerializer(obj).data)

            # --- Auto-Completion Logic ---
            # After saving, check if the batch should be completed.
            # We'll use Day 6 as the trigger.
            final_day_id = 6 # Assuming Day ID 6 is the final day. Adjust if necessary.
            if any(item.get('day_number') == final_day_id for item in attendance_data):
                all_users_in_batch_count = UserRegistration.objects.filter(batch_id=target_batch_id).count()
                final_day_attendance_count = TrainingAttendance.objects.filter(
                    batch_id=target_batch_id,
                    day_number_id=final_day_id
                ).count()

                if all_users_in_batch_count > 0 and all_users_in_batch_count == final_day_attendance_count:
                    # All users have the final day marked, so complete the batch.
                    TrainingBatch.objects.filter(batch_id=target_batch_id).update(is_active=False)
                    print(f"✅ Automatically completed batch {target_batch_id}")


        except (ValueError, Days.DoesNotExist, UserRegistration.DoesNotExist) as e:
            return Response({"error": f"Data validation error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"An unexpected error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            "message": f"Attendance saved successfully. {len(created_records)} created, {len(updated_records)} updated.",
            "created": created_records,
            "updated": updated_records
        }, status=status.HTTP_201_CREATED)

class CompleteTrainingBatchView(APIView):
    def post(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            if not batch.is_active:
                return Response({"message": "Batch is already completed."}, status=status.HTTP_200_OK)
            batch.is_active = False
            batch.save()
            return Response(TrainingBatchSerializer(batch).data, status=status.HTTP_200_OK)
        except TrainingBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)
# =================== TrainingAttendance End ============================= #
        






from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum
from .models import AdvanceManpowerDashboard
from .serializers import AdvanceManpowerDashboardSerializer


class AdvanceManpowerDashboardViewSet(viewsets.ModelViewSet):
    """
    This ViewSet handles the data for the Advance Manpower Dashboard.
    It now correctly filters by factory and department.
    """
    queryset = AdvanceManpowerDashboard.objects.all()
    serializer_class = AdvanceManpowerDashboardSerializer

    def get_queryset(self):
        """
        This method is automatically called by DRF for list views.
        We override it to apply our filters.
        """
        # Start with the base queryset
        queryset = super().get_queryset()

        # Get 'factory' and 'department' from the URL query parameters
        # e.g., /?factory=2&department=3
        factory_id = self.request.query_params.get('factory')
        department_id = self.request.query_params.get('department')

        # If a factory_id is provided in the URL, filter the queryset
        if factory_id:
            queryset = queryset.filter(factory_id=factory_id)

        # If a department_id is provided in the URL, filter the queryset further
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        # Return the final, filtered queryset
        return queryset
    

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """
        Generates an Excel template with name-based headers and example data.
        """
        headers = [
            'hq_name', 'factory_name', 'department_name', 'month', 'year',
            'total_stations', 'operators_required', 'operators_available',
            'buffer_manpower_required', 'buffer_manpower_available',
            'attrition_rate', 'absenteeism_rate'
        ]
        
        # Example data to guide the user
        dummy_data = [
            {
                'hq_name': 'Corporate HQ', 'factory_name': 'Main Plant', 'department_name': 'Assembly',
                'month': 1, 'year': 2024, 'total_stations': 50, 'operators_required': 120,
                'operators_available': 115, 'buffer_manpower_required': 12, 'buffer_manpower_available': 10,
                'attrition_rate': 2.5, 'absenteeism_rate': 3.1, 
                
            },
            {
                'hq_name': 'Corporate HQ', 'factory_name': 'Secondary Plant', 'department_name': '',
                'month': 1, 'year': 2024, 'total_stations': 25, 'operators_required': 60,
                'operators_available': 58, 'buffer_manpower_required': 6, 'buffer_manpower_available': 6,
                'attrition_rate': 1.8, 'absenteeism_rate': 2.5,
                
            },
        ]

        df = pd.DataFrame(dummy_data, columns=headers)
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='ManpowerData')
        
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Using a more descriptive filename
        response['Content-Disposition'] = 'attachment; filename="manpower_upload_template_with_examples.xlsx"'
        
        return response

    # --- UPDATED: Excel Upload Action to handle Names instead of IDs ---
    @action(detail=False, methods=['post'], url_path='upload-data')
    @transaction.atomic
    def upload_data(self, request):
        """
        Accepts an Excel file with names, looks up the corresponding IDs using a
        case-insensitive and whitespace-trimmed search, and creates or updates records.
        """
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Use fillna('') to handle empty cells gracefully
            df = pd.read_excel(file_obj, na_filter=True).fillna('')
        except Exception as e:
            return Response({"error": f"Error reading Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = {'factory_name', 'month', 'year'}
        if not required_columns.issubset(df.columns):
            missing = required_columns - set(df.columns)
            return Response({"error": f"Missing required columns: {', '.join(missing)}"}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        created_count = 0
        updated_count = 0

        for index, row in df.iterrows():
            row_num = index + 2  # For user-friendly error messages
            try:
                # --- Name to ID Lookup (Now more robust) ---
                # Get the value and strip any leading/trailing whitespace
                hq_name = str(row.get('hq_name', '')).strip()
                factory_name = str(row.get('factory_name', '')).strip()
                department_name = str(row.get('department_name', '')).strip()

                if not factory_name:
                    errors.append(f"Row {row_num}: 'factory_name' cannot be empty.")
                    continue

                # Look up Factory (required, case-insensitive)
                try:
                    factory = Factory.objects.get(factory_name__iexact=factory_name)
                except Factory.DoesNotExist:
                    errors.append(f"Row {row_num}: Factory with name '{factory_name}' not found.")
                    continue
                
                # Look up Hq (optional, case-insensitive)
                hq = None
                if hq_name:
                    try:
                        hq = Hq.objects.get(hq_name__iexact=hq_name)
                    except Hq.DoesNotExist:
                        errors.append(f"Row {row_num}: HQ with name '{hq_name}' not found.")
                        continue
                
                # Look up Department (optional, case-insensitive)
                department = None
                if department_name:
                    try:
                        # THIS IS THE KEY FIX: using __iexact
                        department = Department.objects.get(department_name__iexact=department_name)
                    except Department.DoesNotExist:
                        errors.append(f"Row {row_num}: Department with name '{department_name}' not found.")
                        continue
                
                # --- Prepare data for update_or_create ---
                lookup_fields = {
                    'factory': factory,
                    'month': int(row['month']),
                    'year': int(row['year']),
                    'hq': hq,
                    'department': department,
                }
                
                data_fields = {
                    'total_stations': int(row.get('total_stations', 0) or 0),
                    'operators_required': int(row.get('operators_required', 0) or 0),
                    'operators_available': int(row.get('operators_available', 0) or 0),
                    'buffer_manpower_required': int(row.get('buffer_manpower_required', 0) or 0),
                    'buffer_manpower_available': int(row.get('buffer_manpower_available', 0) or 0),
                    'attrition_rate': float(row.get('attrition_rate', 0.0) or 0.0),
                    'absenteeism_rate': float(row.get('absenteeism_rate', 0.0) or 0.0),
                }

                obj, created = AdvanceManpowerDashboard.objects.update_or_create(
                    defaults=data_fields,
                    **lookup_fields
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except (ValueError, TypeError) as e:
                errors.append(f"Row {row_num}: Data type error. Please check numeric fields. Details: {e}")
            except Exception as e:
                errors.append(f"Row {row_num}: An unexpected error occurred: {str(e)}")

        if errors:
            transaction.set_rollback(True)
            return Response({
                "status": "Error",
                "message": "Upload failed due to data errors. No records were saved.",
                "errors": errors
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "status": "Success",
            "message": "Data uploaded successfully.",
            "records_created": created_count,
            "records_updated": updated_count,
        }, status=status.HTTP_201_CREATED)
    

    
    @action(detail=False, methods=["get"])
    def department_wise_stations(self, request):
        data = (
            AdvanceManpowerDashboard.objects
            .values("department__id", "department__department_name")
            .annotate(total_stations=Sum("total_stations"))
        )
        return Response(data)

    @action(detail=False, methods=["get"])
    def department_month_year(self, request):
        department_id = request.query_params.get("department_id")
        month = request.query_params.get("month")
        year = request.query_params.get("year")

        qs = AdvanceManpowerDashboard.objects.all()

        if department_id:
            qs = qs.filter(department_id=department_id)
        if month:
            qs = qs.filter(month=month)
        if year:
            qs = qs.filter(year=year)

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)



# views.py (CORRECTED)

from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response

from .models import ManagementReview
from .serializers import (
    ManagementReviewSerializer, TrainingDataSerializer, DefectsDataSerializer,
    OperatorsChartSerializer, TrainingPlansChartSerializer, DefectsChartSerializer
)

class ManagementReviewViewSet(viewsets.ModelViewSet):
    # This viewset is fine, no changes needed here.
    queryset = ManagementReview.objects.all().select_related('hq', 'factory', 'department')
    serializer_class = ManagementReviewSerializer
    pagination_class = None

class CurrentMonthTrainingDataView(APIView):
    def get(self, request):
        current_time = timezone.now()
        try:
            data = ManagementReview.objects.get(
                year=current_time.year,   # FIX: Changed from month_year__year
                month=current_time.month  # FIX: Changed from month_year__month
            )
            serializer = TrainingDataSerializer(data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ManagementReview.DoesNotExist:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )

class CurrentMonthDefectsDataView(APIView):
    def get(self, request):
        current_time = timezone.now()
        data = (
            ManagementReview.objects
            .filter(
                year=current_time.year,   # FIX: Changed from month_year__year
                month=current_time.month  # FIX: Changed from month_year__month
            )
            .order_by('-id')
            .first()
        )
        if not data:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = DefectsDataSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)

class PreviousMonthDefectsDataView(APIView):
    def get(self, request):
        current_date = timezone.now().date().replace(day=1)
        if current_date.month == 1:
            prev_year = current_date.year - 1
            prev_month = 12
        else:
            prev_year = current_date.year
            prev_month = current_date.month - 1
        
        data = (
            ManagementReview.objects
            .filter(
                year=prev_year,   # FIX: Changed from month_year__year
                month=prev_month  # FIX: Changed from month_year__month
            )
            .order_by('-id')
            .first()
        )
        if not data:
            return Response(
                {"message": "No data found for previous month"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = DefectsDataSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)

class NextMonthDefectsDataView(APIView):
    def get(self, request):
        current_date = timezone.now().date().replace(day=1)
        if current_date.month == 12:
            next_year = current_date.year + 1
            next_month = 1
        else:
            next_year = current_date.year
            next_month = current_date.month + 1

        data = (
            ManagementReview.objects
            .filter(
                year=next_year,   # FIX: Changed from month_year__year
                month=next_month  # FIX: Changed from month_year__month
            )
            .order_by('-id')
            .first()
        )
        if not data:
            return Response(
                {"message": "No data found for next month"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = DefectsDataSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)

class OperatorsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            year=current_year  # FIX: Changed from month_year__year
        ).order_by('year', 'month')  # FIX: Changed from 'month_year'
        serializer = OperatorsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class TrainingPlansChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            year=current_year  # FIX: Changed from month_year__year
        ).order_by('year', 'month')  # FIX: Changed from 'month_year'
        serializer = TrainingPlansChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class DefectsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            year=current_year  # FIX: Changed from month_year__year
        ).order_by('year', 'month')  # FIX: Changed from 'month_year'
        serializer = DefectsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

# your_app/views.py
from django.http import HttpResponse, JsonResponse # <-- Import JsonResponse
from django.views.decorators.csrf import csrf_exempt # For API views, often helpful
from rest_framework.parsers import MultiPartParser
from rest_framework.views import APIView
import pandas as pd
from .serializers import ManagementReviewSerializer

# We can refactor the upload view into a DRF APIView for better structure
class ManagementReviewUploadAPIView(APIView):
    parser_classes = [MultiPartParser] # Handles multipart/form-data for file uploads

    def post(self, request, format=None):
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            return JsonResponse({"error": "No file was uploaded."}, status=400)

        if not excel_file.name.endswith(('.xls', '.xlsx')):
            return JsonResponse({"error": "Invalid file format. Please upload an Excel file."}, status=400)

        try:
            df = pd.read_excel(excel_file)
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            return JsonResponse({"error": f"Error reading Excel file: {e}"}, status=400)

        successful_uploads = 0
        failed_uploads = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            serializer = ManagementReviewSerializer(data=row_data)
            
            if serializer.is_valid():
                try:
                    serializer.save()
                    successful_uploads += 1
                except Exception as e:
                    failed_uploads.append({'row': index + 2, 'errors': str(e)})
            else:
                failed_uploads.append({'row': index + 2, 'errors': serializer.errors})
        
        if not failed_uploads:
            return JsonResponse({
                "message": "All rows uploaded successfully!",
                "successful_uploads": successful_uploads,
                "failed_uploads_count": len(failed_uploads),
                "errors": []
            }, status=201)
        else:
            return JsonResponse({
                "message": f"Upload completed with {len(failed_uploads)} errors.",
                "successful_uploads": successful_uploads,
                "failed_uploads_count": len(failed_uploads),
                "errors": failed_uploads
            }, status=207) # 207 Multi-Status is appropriate here

# The download view remains the same.
def download_sample_excel(request):
    columns = [
        "hq", "factory", "department", "month", "year",
        "new_operators_joined", "new_operators_trained",
        "total_training_plans", "total_trainings_actual",
        "total_defects_msil", "ctq_defects_msil",
        "total_defects_tier1", "ctq_defects_tier1",
        "total_internal_rejection", "ctq_internal_rejection"
    ]
    sample_data = [{
        "hq": "HQ-North", "factory": "Factory-A", "department": "Assembly",
        "month": 1, "year": 2023, "new_operators_joined": 10,
        "new_operators_trained": 8, "total_training_plans": 20,
        "total_trainings_actual": 18, "total_defects_msil": 5,
        "ctq_defects_msil": 2, "total_defects_tier1": 3,
        "ctq_defects_tier1": 1, "total_internal_rejection": 15,
        "ctq_internal_rejection": 4
    }]
    df = pd.DataFrame(sample_data, columns=columns)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="management_review_sample.xlsx"'
    df.to_excel(response, index=False)
    return response




from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView

from .models import (
    MasterTable,
    # OperatorSkill,
    SkillMatrix,
    Score,
    # MultiSkilling,
    # RefreshmentTraining
)

from .serializers import (
    CardEmployeeMasterSerializer,
    OperatorCardSkillSerializer,
    CardScoreSerializer,
    # CardMultiSkillingSerializer,
    # CardRefreshmentTrainingSerializer,
    CardHanchouExamResultSerializer,  # <-- 2. IMPORT the new serializer
    CardTrainingAttendanceSerializer,
    CardScheduleSerializer
)
class EmployeeCardDetailsView(APIView):
    def get(self, request):
        card_no = request.query_params.get('card_no')
        if not card_no:
            return Response({'error': 'card_no parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = MasterTable.objects.get(emp_id=card_no)
        except MasterTable.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        # Fetch and serialize all related data
        employee_data = CardEmployeeMasterSerializer(employee).data
        operator_skills = OperatorCardSkillSerializer(SkillMatrix.objects.filter(employee=employee), many=True).data
        scores = CardScoreSerializer(Score.objects.filter(employee=employee), many=True).data
        # multi_skilling = CardMultiSkillingSerializer(MultiSkilling.objects.filter(employee=employee), many=True).data
        # refreshment_training = CardRefreshmentTrainingSerializer(RefreshmentTraining.objects.filter(employee=employee), many=True).data
        hanchou_results = CardHanchouExamResultSerializer(
            HanchouExamResult.objects.filter(employee=employee),
            many=True
        ).data
        shokuchou_results = CardShokuchouExamResultSerializer(
            ShokuchouExamResult.objects.filter(employee=employee),
            many=True
        ).data

        # --- ★ UPDATED SCHEDULE LOGIC ★ ---
        
        # 1. Get ALL schedules for this employee, REMOVE the category filter
        all_schedules = Schedule.objects.filter(employees=employee)

        # 2. Use the renamed serializer
        serializer_context = {'employee': employee}
        scheduled_trainings_data = CardScheduleSerializer(
            all_schedules, 
            many=True, 
            context=serializer_context
        ).data

        # --- ★ END: REVISED LOGIC ★ ---




        attendances_data = [] # Default to an empty list
        # employee_name = employee.name
        employee_name = employee.first_name

        try:
            # Step 1: Find all UserInfo records that match the employee's name.
            # We use __iexact for a case-insensitive match, which is safer.
            # This might return multiple users if names are not unique.
            user_info_records = UserRegistration.objects.filter(first_name__iexact=employee_name)

            if user_info_records.exists():
                # Step 2: Use the found UserInfo records to query for their attendance.
                # The `user__in` lookup efficiently finds attendance for all matched users.
                attendances = TrainingAttendance.objects.filter(user__in=user_info_records)
                
                # Step 3: Serialize the data
                attendances_data = CardTrainingAttendanceSerializer(attendances, many=True).data

        except Exception as e:
            # If anything goes wrong, we'll log it and continue with empty attendance
            print(f"Error fetching attendance for {employee_name}: {e}")
            attendances_data = []

        # Construct full response
        response_data = {
            'employee': employee_data,
            'operator_skills': operator_skills,
            'scores': scores,
            # 'multi_skilling': multi_skilling,
            'scheduled_trainings': scheduled_trainings_data,
            'hanchou_results': hanchou_results, 
            'shokuchou_results': shokuchou_results,
            'attendance': attendances_data,
        }

        # Print to console
        print("==== Employee Card Details ====")
        import pprint
        pprint.pprint(response_data)  # pretty-print for readability
        print("================================")

        return Response(response_data)

# =================== end =============================


# =================== excel download master table =======================
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from .models import MasterTable
from .serializers import MasterTableSerializer

class EmployeeExcelViewSet(viewsets.ModelViewSet):
    queryset = MasterTable.objects.all().select_related('department')
    serializer_class = MasterTableSerializer
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        queryset = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "EMPLOYEE MASTER"

        # Headers and titles
        ws.merge_cells('A1:I1')
        company_cell = ws['A1']
        company_cell.value = "Company Name: NL Technologies"
        company_cell.font = Font(bold=True, size=10)
        company_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A2:I2')
        run_date_cell = ws['A2']
        run_date_cell.value = f"Run Date & Time: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        run_date_cell.font = Font(size=10)
        run_date_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A4:I4')
        title_cell = ws['A4']
        title_cell.value = "EMPLOYEE MASTER"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            'Srl. No.',
            'Employee ID',
            'First Name',
            'Last Name',
            'Department',
            'Date of Joining',
            'Birth Date',
            'Sex',
            'Email',
            'Phone',
        ]
        header_font = Font(bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            
        ws.row_dimensions[6].height = 30
        
        for row_num, employee in enumerate(queryset, 7):
            birth_date = employee.birth_date.strftime('%d/%m/%Y') if employee.birth_date else ''
            joining_date = employee.date_of_joining.strftime('%d/%m/%Y') if employee.date_of_joining else ''
            department_name = employee.department.department_name if employee.department else 'N/A'

            row_data = [
                row_num - 6,  # Serial number starting from 1
                employee.emp_id,
                employee.first_name,
                employee.last_name,
                department_name,
                joining_date,
                birth_date,
                employee.get_sex_display(),
                employee.email,
                employee.phone,
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_num in [1, 2, 7, 8] else "left", vertical="center")
                cell.font = Font(size=9)
        
        # Set column widths
        column_widths = {
            'A': 6, 'B': 15, 'C': 15, 'D': 15, 'E': 20, 'F': 15, 'G': 12, 'H': 6, 'I': 25, 'J': 15,
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'EMPLOYEE_MASTER_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        return response
    
# ==================== MultiSkilling Start ======================== #
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from django.db.models import Q
from .models import MasterTable, SkillMatrix, MultiSkilling, Level
from .serializers import MultiSkillingSerializer

class EmployeeSkillSearch(APIView):
    """
    Search employee by emp_id or name and return their current skills.
    Only returns employees who have at least one skill in SkillMatrix.
    """
    def get(self, request):
        query = request.GET.get("query", "").strip()
        if not query:
            return Response({"error": "query parameter is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Find all employees who match the search criteria
        employees_query = MasterTable.objects.filter(
            Q(emp_id__icontains=query) | 
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query)
        )
        
        # Get employees who have any skill matrix entries
        skilled_employee_ids = SkillMatrix.objects.filter(
            employee__in=employees_query
        ).values_list('employee__emp_id', flat=True).distinct()  
        
        employees_with_skills = employees_query.filter(emp_id__in=skilled_employee_ids) 
        
        result = []
        for emp in employees_with_skills:
            # Get all skills for this employee
            skills = SkillMatrix.objects.filter(employee=emp).select_related(
                "hierarchy__station", 
                "hierarchy__department", 
                "level"
            )
            
            skills_data = []
            for s in skills:
                skills_data.append({
                    "skill_id": s.id,
                    "station": s.hierarchy.station.station_name if s.hierarchy and s.hierarchy.station else None,
                    "department": s.hierarchy.department.department_name if s.hierarchy and s.hierarchy.department else None,
                    "level": s.level.level_name,
                    "updated_at": s.updated_at,
                })
            
            result.append({
                "emp_id": emp.emp_id,
                "first_name": emp.first_name,
                "last_name": emp.last_name,
                "department": emp.department.department_name if emp.department else None,
                "date_of_joining": emp.date_of_joining,
                "skills": skills_data,
            })
        
        return Response(result, status=status.HTTP_200_OK)




class MultiSkillingViewSet(viewsets.ModelViewSet):
    queryset = MultiSkilling.objects.all().select_related( 
       "employee",
       "department",
       "station",
       "skill_level"
    )
    serializer_class = MultiSkillingSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        emp_id = self.request.query_params.get("emp_id")
        if emp_id:
            qs = qs.filter(employee__emp_id=emp_id)
        return qs




from .models import Machine,MachineAllocation
from .serializers import MachineSerializer,MachineAllocationSerializer

class MachineViewSet(viewsets.ModelViewSet):
    queryset = Machine.objects.all().order_by("id")
    serializer_class = MachineSerializer
    @action(detail=False, methods=['get'])
    def by_department(self, request):
        """Get machines filtered by department"""
        department_id = request.query_params.get('department_id')
        if department_id:
            machines = self.queryset.filter(department_id=department_id)
            serializer = self.get_serializer(machines, many=True)
            return Response(serializer.data)
        return Response({"error": "department_id parameter required"}, 
                       status=status.HTTP_400_BAD_REQUEST)

class MachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all().order_by("-allocated_at")
    serializer_class = MachineAllocationSerializer
    @action(detail=False, methods=['get'])
    def eligible_employees(self, request):
        """Get employees eligible for a specific machine based on level"""
        machine_id = request.query_params.get('machine_id')
        department_id = request.query_params.get('department_id')
        
        if not machine_id:
            return Response({"error": "machine_id parameter required"},
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            machine = Machine.objects.get(id=machine_id)
        except Machine.DoesNotExist:
            return Response({"error": "Machine not found"},
                          status=status.HTTP_404_NOT_FOUND)
        
        # Get SkillMatrix entries for employees in the target department
        # Filter by hierarchy department if department_id is provided
        target_department = department_id or machine.department.id
        
        # Get already allocated employees for this machine
        allocated_employees = MachineAllocation.objects.filter(
            machine=machine
        ).values_list('employee_id', flat=True)
        
        # Get eligible employees from SkillMatrix
        eligible_employees_queryset = SkillMatrix.objects.filter(
            hierarchy__department_id=target_department
        ).exclude(id__in=allocated_employees)
        
        # If you want to filter by specific hierarchy/station, add this:
        # .filter(hierarchy__station=machine.process)  # if machine.process links to station
        
        serializer = EligibleEmployeeSerializer(
            eligible_employees_queryset,
            many=True,
            context={'machine_level': machine.level}
        )
        
        return Response({
            'machine_level': machine.level,
            'employees': serializer.data
        })

    @action(detail=False, methods=['post'])
    def update_pending_status(self, request):
        """Manually trigger update of pending allocations"""
        MachineAllocation.update_pending_allocations()
        return Response({"message": "Pending allocations updated successfully"})

    @action(detail=False, methods=['get'])
    def by_status(self, request):
        """Get allocations filtered by approval status"""
        status_filter = request.query_params.get('status')
        if status_filter:
            allocations = self.queryset.filter(approval_status=status_filter)
            serializer = self.get_serializer(allocations, many=True)
            return Response(serializer.data)
        return Response({"error": "status parameter required"}, 
                status=status.HTTP_400_BAD_REQUEST)

    def perform_update(self, serializer):
        """Override to add any additional logic during update"""
        allocation = serializer.save()
        
        # Log the allocation update if needed
        print(f"Updated allocation: {allocation}")
        
        return allocation
    def perform_create(self, serializer):
        """Override to add any additional logic during creation"""
        # Ensure we're working with SkillMatrix instance
        employee_id = self.request.data.get('employee')
        try:
            skill_matrix_employee = SkillMatrix.objects.get(id=employee_id)
        except SkillMatrix.DoesNotExist:
            raise ValidationError("Employee not found in skill matrix")
        
        allocation = serializer.save(employee=skill_matrix_employee)
        print(f"Created allocation: {allocation}")
        return allocation




class MachineAllocationApprovalViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationApprovalSerializer

    @action(detail=True, methods=['put'], url_path='set-status')
    def set_status(self, request, pk=None):
        allocation = self.get_object()
        status_value = request.data.get('approval_status')

        if status_value not in dict(MachineAllocation.APPROVAL_STATUS_CHOICES):
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

        allocation.approval_status = status_value
        allocation.save()
        return Response({
            'status': 'success',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        })

    @action(detail=True, methods=['put'], url_path='reject')
    def reject(self, request, pk=None):
        allocation = self.get_object()
        allocation.approval_status = 'rejected'
        allocation.save()
        return Response({
            'status': 'rejected',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        }, status=status.HTTP_200_OK)
    



from rest_framework import viewsets
from .models import EvaluationLevel2
from .serializers import EvaluationLevel2Serializer

class EvaluationLevel2ViewSet(viewsets.ModelViewSet):
    """
    ViewSet for EvaluationLevel2 with correct manual filtering.
    """
    queryset = EvaluationLevel2.objects.all().order_by('-created_at')
    serializer_class = EvaluationLevel2Serializer

    def get_queryset(self):
        """
        This method correctly filters by the provided query parameters.
        The fix is removing the unnecessary '__id' from ForeignKey lookups.
        """
        queryset = super().get_queryset() 
        
        employee_id = self.request.query_params.get('employee__emp_id', None)
        department_id = self.request.query_params.get('department', None)
        station = self.request.query_params.get('station_name', None)
        level_id = self.request.query_params.get('level', None)
        
        if employee_id is not None:
            queryset = queryset.filter(employee__emp_id=employee_id)
            
        if department_id is not None:
            # ✅ CORRECTED LINE: No '__id' needed for ForeignKey filtering by ID.
            queryset = queryset.filter(department=department_id)
            
        if station is not None:
            queryset = queryset.filter(station_name__iexact=station)
            
        if level_id is not None:
            # ✅ CORRECTED LINE: No '__id' needed here either.
            queryset = queryset.filter(level=level_id)
            
        return queryset
    
from rest_framework import viewsets, permissions
from .models import EvaluationCriterion
from .serializers import EvaluationCriterionSerializer

class EvaluationCriterionViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows criteria to be viewed.
    Provides a list of criteria filtered by the level ID in the URL.
    e.g., /api/levels/2/criteria/
    """
    serializer_class = EvaluationCriterionSerializer
    queryset = EvaluationCriterion.objects.all().order_by('level', 'display_order')

    def get_queryset(self):
        """
        This view should return a list of all the criteria
        for the level as determined by the level_pk portion of the URL.
        """
        level_id = self.kwargs.get('level_pk')
        return EvaluationCriterion.objects.filter(level_id=level_id, is_active=True)

# views.py   for productivity and quality sheet Level 1

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .models import ProductivityEvaluation, ProductivitySequence
from .serializers import ProductivityEvaluationSerializer, ProductivitySequenceSerializer

class ProductivityEvaluationViewSet(viewsets.ModelViewSet):
    queryset = ProductivityEvaluation.objects.all()
    serializer_class = ProductivityEvaluationSerializer

    
    def get_queryset(self):
        """Override to allow filtering by employee"""
        queryset = ProductivityEvaluation.objects.all()
        employee = self.request.query_params.get('employee', None)
        if employee is not None:
            queryset = queryset.filter(employee=employee)
        return queryset.order_by('-evaluation_date')  # Most recent first




    @action(detail=True, methods=["post"])
    def calculate_results(self, request, pk=None):
        try:
            evaluation = self.get_object()
            sequences = evaluation.sequences.all()
            
            # Calculate individual column totals
            e1_total = 0
            e2_total = 0
            e3_total = 0
            max_total = 0
            
            for seq in sequences:
                # Skip cycle time sequence (mt = 0)
                if 'cycle time' in seq.sequence_name.lower() and seq.mt == 0:
                    continue
                
                e1_total += seq.e1
                e2_total += seq.e2
                e3_total += seq.e3
                max_total += seq.mt
            
            # Determine final result based on best column performance
            final_marks = 0
            final_percentage = 0
            final_status = "FAIL"
            
            if e1_total >= 12:
                final_marks = e1_total
                final_percentage = (e1_total / 15) * 100
                final_status = "PASS"
            elif e2_total >= 12:
                final_marks = e2_total
                final_percentage = (e2_total / 15) * 100
                final_status = "PASS"
            elif e3_total >= 12:
                final_marks = e3_total
                final_percentage = (e3_total / 15) * 100
                final_status = "PASS"
            else:
                # If none passed, use the highest score
                final_marks = max(e1_total, e2_total, e3_total)
                final_percentage = (final_marks / 15) * 100
                final_status = "FAIL"
            
            evaluation.max_marks = 15
            evaluation.obtained_marks = final_marks
            evaluation.percentage = final_percentage
            evaluation.status = final_status
            evaluation.save()
            
            return Response({
                "employee": f"{evaluation.employee.first_name} {evaluation.employee.last_name}",
                "emp_id": evaluation.employee.emp_id,
                "designation": evaluation.employee.designation,
                "department": evaluation.employee.department.department_name if evaluation.employee.department else None,
                "date_of_joining": evaluation.employee.date_of_joining,
                "obtained_marks": evaluation.obtained_marks,
                "max_marks": evaluation.max_marks,
                "percentage": evaluation.percentage,
                "status": evaluation.status,
                "e1_total": e1_total,
                "e2_total": e2_total,
                "e3_total": e3_total
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing evaluations"""
        employee_id = request.data.get('employee')
        
        existing_evaluation = ProductivityEvaluation.objects.filter(employee=employee_id).first()
        
        if existing_evaluation:
            serializer = self.get_serializer(existing_evaluation, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return super().create(request, *args, **kwargs)

class ProductivitySequenceViewSet(viewsets.ModelViewSet):
    queryset = ProductivitySequence.objects.all()
    serializer_class = ProductivitySequenceSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing sequences"""
        evaluation_id = request.data.get('evaluation')
        sequence_name = request.data.get('sequence_name')
        
       
        existing_sequence = ProductivitySequence.objects.filter(
            evaluation=evaluation_id, 
            sequence_name=sequence_name
        ).first()
        
        if existing_sequence:
            
            serializer = self.get_serializer(existing_sequence, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return super().create(request, *args, **kwargs)


     





from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .models import QualityEvaluation, QualitySequence
from .serializers import QualityEvaluationSerializer, QualitySequenceSerializer

class QualityEvaluationViewSet(viewsets.ModelViewSet):
    queryset = QualityEvaluation.objects.all()
    serializer_class = QualityEvaluationSerializer

    def get_queryset(self):
        """Override to allow filtering by employee"""
        queryset = QualityEvaluation.objects.all()
        employee = self.request.query_params.get('employee', None)
        if employee is not None:
            queryset = queryset.filter(employee=employee)
        return queryset.order_by('-evaluation_date')  # Most recent first

    @action(detail=True, methods=["post"])
    def calculate_results(self, request, pk=None):
        try:
            evaluation = self.get_object()
            sequences = evaluation.qualitysequences.all()
            
           
            e1_total = 0
            e2_total = 0
            e3_total = 0
            max_total = 0
            
            for seq in sequences:
                
                if 'cycle time' in seq.sequence_name.lower() and seq.mt == 0:
                    continue
                
                e1_total += seq.e1
                e2_total += seq.e2
                e3_total += seq.e3
                max_total += seq.mt
            
            # Determine final result based on best column performance
            final_marks = 0
            final_percentage = 0
            final_status = "FAIL"
            
            if e1_total >= 12:
                final_marks = e1_total
                final_percentage = (e1_total / 15) * 100
                final_status = "PASS"
            elif e2_total >= 12:
                final_marks = e2_total
                final_percentage = (e2_total / 15) * 100
                final_status = "PASS"
            elif e3_total >= 12:
                final_marks = e3_total
                final_percentage = (e3_total / 15) * 100
                final_status = "PASS"
            else:
                # If none passed, use the highest score
                final_marks = max(e1_total, e2_total, e3_total)
                final_percentage = (final_marks / 15) * 100
                final_status = "FAIL"
            
            evaluation.max_marks = 15
            evaluation.obtained_marks = final_marks
            evaluation.percentage = final_percentage
            evaluation.status = final_status
            evaluation.save()
            
            return Response({
                "employee": f"{evaluation.employee.first_name} {evaluation.employee.last_name}",
                "emp_id": evaluation.employee.emp_id,
                "designation": evaluation.employee.designation,
                "department": evaluation.employee.department.department_name if evaluation.employee.department else None,
                "date_of_joining": evaluation.employee.date_of_joining,
                "obtained_marks": evaluation.obtained_marks,
                "max_marks": evaluation.max_marks,
                "percentage": evaluation.percentage,
                "status": evaluation.status,
                "e1_total": e1_total,
                "e2_total": e2_total,
                "e3_total": e3_total
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing evaluations"""
        employee_id = request.data.get('employee')
        
        # Check if evaluation already exists for this employee
        existing_evaluation = QualityEvaluation.objects.filter(employee=employee_id).first()
        
        if existing_evaluation:
            # Update existing evaluation
            serializer = self.get_serializer(existing_evaluation, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            # Create new evaluation
            return super().create(request, *args, **kwargs)

class QualitySequenceViewSet(viewsets.ModelViewSet):
    queryset = QualitySequence.objects.all()
    serializer_class = QualitySequenceSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """Override create to handle updating existing sequences"""
        evaluation_id = request.data.get('evaluation')
        sequence_name = request.data.get('sequence_name')
        
        
        existing_sequence = QualitySequence.objects.filter(
            evaluation=evaluation_id, 
            sequence_name=sequence_name
        ).first()
        
        if existing_sequence:
           
            serializer = self.get_serializer(existing_sequence, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
           
            return super().create(request, *args, **kwargs)


# level1revision start
from rest_framework import viewsets
from .models import Question # Import new model
from .serializers import QuestionSerializer # Import new serializer

# ... your existing views ...

class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    
    # This allows filtering like /api/questions/?subtopiccontent=5
    def get_queryset(self):
        queryset = super().get_queryset()
        subtopiccontent_id = self.request.query_params.get('subtopiccontent')
        if subtopiccontent_id:
            queryset = queryset.filter(subtopiccontent_id=subtopiccontent_id)
        return queryset
# end

# operatorobservance sheet 

from rest_framework import viewsets
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import Topic, OperatorObservanceSheet
from .serializers import TopicSerializer, OperatorObservanceSheetSerializer

class TopicViewSet(viewsets.ModelViewSet):
    queryset = Topic.objects.all().order_by('sr_no')
    serializer_class = TopicSerializer
    lookup_field = 'sr_no'
class OperatorObservanceSheetViewSet(viewsets.ModelViewSet):
    queryset = OperatorObservanceSheet.objects.all()
    serializer_class = OperatorObservanceSheetSerializer

# @api_view(['GET'])
# def get_sheet_by_operator(request, operator_name):
#     try:
#         sheet = OperatorObservanceSheet.objects.filter(operator_name=operator_name).last()
#         if not sheet:
#             return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
#         serializer = OperatorObservanceSheetSerializer(sheet)
#         return Response(serializer.data)
#     except Exception as e:
#         return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

# @api_view(['GET'])
# def get_sheet_by_operator_level_station(request, operator_name, level,process_name):
#     try:
#         sheet = OperatorObservanceSheet.objects.filter(
#             operator_name=operator_name,
#             level=level,
#             process_name=process_name
#         ).last()
#         if not sheet:
#             return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
#         serializer = OperatorObservanceSheetSerializer(sheet)
#         return Response(serializer.data)
#     except Exception as e:
#         return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_sheet_by_operator_level_station(request, operator_name, level, station_name):
    try:
        sheet = OperatorObservanceSheet.objects.filter(
            operator_name=operator_name,
            level=level,
            process_name=station_name
        ).last()
        if not sheet:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = OperatorObservanceSheetSerializer(sheet)
        return Response(serializer.data)
    except Exception as e:
        return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
